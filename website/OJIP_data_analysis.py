from flask import Blueprint, render_template, request, jsonify, redirect, url_for, send_file
import os, base64, io, zipfile, uuid
from typing import cast
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy.interpolate import UnivariateSpline, LSQUnivariateSpline, PchipInterpolator
from scipy.ndimage import gaussian_filter1d
from scipy.optimize import curve_fit
from scipy.signal import find_peaks
from . import UPLOAD_FOLDER, csrf
from werkzeug.utils import secure_filename
from .ojip_interpretation import interpret_ojip, generate_narrative, summarise_findings, compare_ojip_params

OJIP_data_analysis = Blueprint('OJIP_data_analysis', __name__)


# ─── helpers ────────────────────────────────────────────────────────────────

def _ms_factor(fluorometer):
    """Multiplier: native time unit → milliseconds."""
    if fluorometer == 'Aquapen':
        return 0.001        # µs → ms
    if fluorometer == 'FL6000':
        return 1000.0       # s  → ms
    return 1.0              # MULTI-COLOR-PAM and OJIPImaging already in ms


def _axis_cfg(fluorometer: str) -> tuple[str, str, str, float, set[str], dict[str, tuple[float, float]]]:
    """Return (x_axis_col_name, x_unit_label, y_unit_label, xmin_for_plot, allowed_extensions, search_ranges)."""
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        return ('time/ms', 'Time (ms)', 'Fluorescence intensity (V)', 1e-2, {'.csv', '.CSV'},
                dict(FJ=(0.1, 10), FI=(10, 100), FP=(100, 1000)))
    if fluorometer == 'Aquapen':
        return ('time_us', 'Time (μs)', 'Fluorescence intensity (a.u.)', 1e-1, {'.txt'},
                dict(FJ=(100, 10000), FI=(10000, 100000), FP=(100000, 1000000)))
    if fluorometer == 'FL6000':
        return ('time_s', 'Time (s)', 'Fluorescence intensity (a.u.)', 1e-5, {'.txt'},
                dict(FJ=(1e-4, 0.01), FI=(0.01, 0.1), FP=(0.1, 1.0)))
    if fluorometer == 'OJIPImaging':
        return ('time_ms', 'Time (ms)', 'Fluorescence (norm.)', 1e-2, {'.xlsx'},
                dict(FJ=(0.1, 10), FI=(9, 100), FP=(100, 1000)))
    raise ValueError(f'Unknown fluorometer: {fluorometer}')


def _bg_correct(sf, x_col, data_cols, mode='auto', n=1, bckg_map=None):
    """Correct an AquaPen/FluorPen transient for the pre-illumination background.

    The first FluorPen timepoint (~11 µs) is a dark/background reading, not the
    O-step. Taking it as F0 inflates Fv/Fm and drops the F0 marker below the
    visible curve. This subtracts and/or drops the leading point(s) before F0
    is chosen — consistently for both the single-file and batch paths (the old
    hard-coded ``iloc[1:]`` only ever ran on the single-file path, so batch/wide
    FluorPen files kept the 11 µs point as F0).

    mode : 'auto' | 'subtract' | 'drop' | 'keep'
        auto     – subtract the instrument ``Bckg`` (from ``bckg_map``) when
                   available and drop the first point; with no known ``Bckg``,
                   just drop the first point (the historical behaviour).
        subtract – subtract the mean of the first ``n`` points from every point,
                   then drop those ``n`` points.
        drop     – drop the first ``n`` points, no subtraction.
        keep     – return unchanged (the pre-fix behaviour).
    n : int            number of leading background points (subtract / drop).
    bckg_map : dict[str, float] | None   per-column instrument background (auto).

    Returns (sf_corrected, bg_applied). ``bg_applied`` is True when a background
    value was subtracted, so the transient is now in background-subtracted units
    and the caller may substitute the instrument ``Fo`` for F0.
    """
    mode = (mode or 'auto').lower()
    n = max(0, int(n or 0))
    bckg_map = bckg_map or {}
    if mode == 'keep':
        return sf.reset_index(drop=True), False

    bg_applied = False
    if mode == 'subtract':
        k = max(1, n)
        for col in data_cols:
            bg = float(np.nanmean(sf[col].iloc[:k]))
            if np.isfinite(bg):
                sf[col] = sf[col] - bg
                bg_applied = True
        drop = k
    elif mode == 'drop':
        drop = max(1, n)
    else:  # auto
        for col in data_cols:
            bg = bckg_map.get(col)
            if bg is not None and np.isfinite(bg):
                sf[col] = sf[col] - float(bg)
                bg_applied = True
        drop = 1

    sf_out = sf.iloc[drop:].reset_index(drop=True) if drop > 0 else sf.reset_index(drop=True)
    return sf_out, bg_applied


def _fit_start_index(dn: pd.DataFrame, trim_first: int = 0) -> int:
    """First row index to include when fitting.

    Skips the leading row *and* any non-positive (pre-illumination) timepoints:
    MULTI-COLOR-PAM files carry a long baseline at t ≤ 0 ms, and both the
    log10(time) fitters and the geomspace evaluation grid require strictly
    positive time (log10 of ≤ 0 is nan/-inf, which crashes the spline fit and
    blanks the reconstruction). Aquapen/FL6000 loaders already drop t=0, so
    for them the first positive point is index 0 and this reduces to the old
    ``1 + trim_first`` behaviour. The user ``trim_first`` is applied on top.
    """
    t = dn.iloc[:, 0].values.astype(float)
    pos = np.flatnonzero(t > 0)
    first_pos = int(pos[0]) if pos.size else 1
    return max(0, first_pos) + trim_first


def _fit_splines(double_norm_df: pd.DataFrame, x_col: str, kr: int,
                 trim_first: int = 0, trim_last: int = 0) -> tuple:
    """
    Spline-fit, smooth and compute derivatives for each sample column in double_norm_df.
    Returns (Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time_series).

    Gaussian smoothing sigma is adaptive: proportional to the number of data
    points so that short curves (e.g. 60-pt OJIPImaging) retain per-curve
    derivative structure, while long curves (≥200 pts) keep sigma=20 as before.

    trim_first / trim_last: exclude the first / last N data points from the
    spline fit (useful when the tail of OJIPImaging curves is poorly fitted).
    """
    dn = double_norm_df
    cols = dn.columns          # [time_col, file1, file2, ...]
    n_files = len(cols) - 1
    _t_start = _fit_start_index(dn, trim_first)
    _t_end   = len(dn) - trim_last if trim_last > 0 else len(dn)
    n_pts = _t_end - _t_start  # usable points after trim
    sigma = max(2, min(20, round(n_pts * 0.1)))
    # Adaptive knot reduction: for short curves (e.g. 60-pt OJIPImaging),
    # ensure at least ~12 total knots (~10 interior) so the cubic spline can
    # capture the multi-phase OJIP shape.  Long curves keep the user-chosen kr.
    effective_kr = max(1, min(kr, max(1, n_pts // 12)))

    log_time = pd.Series(
        np.geomspace(dn.iloc[_t_start:_t_start + 1, 0].values.astype(float)[0],
                     dn.iloc[_t_end - 1:_t_end, 0].values.astype(float)[0], num=n_pts),
        name=cols[0])

    Raw_recon_list, D1_list, D2_list, D3_list, Infl_list = [], [], [], [], []

    for i in range(1, n_files + 1):
        fname = cols[i]
        x = dn.iloc[_t_start:_t_end, 0].values
        y = dn.iloc[_t_start:_t_end, i].values
        knots = UnivariateSpline(x, x, s=0).get_knots()[::effective_kr]
        model = LSQUnivariateSpline(x, y, knots[1:-1], k=3)
        model.set_smoothing_factor(0.5)
        recon = gaussian_filter1d(model(log_time.values), sigma)
        d1 = gaussian_filter1d(np.gradient(recon), sigma)
        d2 = gaussian_filter1d(np.gradient(d1), sigma)
        d3 = gaussian_filter1d(np.gradient(d2), sigma)
        # Upward zero-crossings only: D2 goes from - to + (D3 > 0 at crossing)
        # These are the inflection points per Akinyemi et al. 2025
        zc   = np.where(np.diff(np.sign(d2)) > 0)[0]
        infl = pd.Series(log_time.values[zc]).reset_index(drop=True).rename(fname)
        Raw_recon_list.append(pd.Series(recon, name=fname))
        D1_list.append(pd.Series(d1, name=fname))
        D2_list.append(pd.Series(d2, name=fname))
        D3_list.append(pd.Series(d3, name=fname))
        Infl_list.append(infl)

    def _assemble(series_list, time_series, col_names):
        df = pd.concat([time_series] + series_list, axis=1)
        df.columns = col_names
        return df

    Raw_recon_DF = _assemble(Raw_recon_list, log_time, cols)
    D1_DF = _assemble(D1_list, log_time, cols)
    D2_DF = _assemble(D2_list, log_time, cols)
    D3_DF = _assemble(D3_list, log_time, cols)
    Infl_DF = pd.concat(Infl_list, axis=1)  # columns = sample names

    # Residuals (interpolate reconstructed→raw time axis)
    Resid_list = []
    for i in range(1, n_files + 1):
        interp = np.interp(np.array(dn.iloc[:, 0], dtype=float), np.array(log_time, dtype=float), np.array(Raw_recon_DF.iloc[:, i], dtype=float))
        Resid_list.append(pd.Series(np.array(dn.iloc[:, i], dtype=float) - interp, name=cols[i]))
    Resid_DF = pd.concat([dn.iloc[:, 0].reset_index(drop=True)] + Resid_list, axis=1)
    Resid_DF.columns = cols

    return Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time


def _fit_oj_polynomial(double_norm_df: pd.DataFrame, x_col: str, ms_factor: float,
                       oj_lo_ms: float = 0.5, oj_hi_ms: float = 5.0,
                       n_dense: int = 500, max_degree: int = 9) -> dict:
    """
    Fit a polynomial to the O-J region of double_norm data (0.5–5 ms by default).
    Inflection points = roots of d2 where d3 > 0 (Akinyemi et al. 2023).

    The polynomial degree adapts to the number of data points in the window:
    ``degree = min(max_degree, n_points - 2)`` so that sparse instruments
    (e.g. OJIPImaging with only 8 points in the O-J window) can still produce
    meaningful fits instead of returning empty.

    Returns {fname: {'poly_oj_time_ms': list, 'poly_oj_d2': list, 'poly_infl_ms': list}}.
    """
    dn = double_norm_df
    cols = dn.columns
    n_files = len(cols) - 1

    oj_lo = oj_lo_ms / ms_factor
    oj_hi = oj_hi_ms / ms_factor
    x_all = np.asarray(dn.iloc[:, 0].values, dtype=float)

    result: dict[str, dict] = {}
    for i in range(1, n_files + 1):
        fname = cols[i]
        y_all = np.array(pd.to_numeric(dn.iloc[:, i], errors='coerce'), dtype=float)

        mask  = (x_all >= oj_lo) & (x_all <= oj_hi) & np.isfinite(y_all)
        x_oj  = x_all[mask]
        y_oj  = y_all[mask]

        empty = {'poly_oj_time_ms': [], 'poly_oj_d2': [], 'poly_infl_ms': []}
        # Need at least 5 points for a meaningful polynomial with d2/d3 analysis
        if len(x_oj) < 5:
            result[fname] = empty
            continue

        # Adaptive degree: ensure we never exceed (n_points - 2) to avoid
        # overfitting, and cap at max_degree (default 9).
        degree = min(max_degree, len(x_oj) - 2)

        # Work in ms; normalise to [-1, 1] for numerical stability
        x_ms = x_oj * ms_factor
        xc   = (x_ms[0] + x_ms[-1]) / 2.0
        xs   = max((x_ms[-1] - x_ms[0]) / 2.0, 1e-12)
        x_n  = (x_ms - xc) / xs

        try:
            coeffs = np.polyfit(x_n, y_oj, degree)
        except Exception:
            result[fname] = empty
            continue

        p   = np.poly1d(coeffs)
        pd2 = p.deriv(2)
        pd3 = p.deriv(3)

        # Dense evaluation grid for smooth display
        x_dense_ms = np.linspace(x_ms[0], x_ms[-1], n_dense)
        x_dense_n  = (x_dense_ms - xc) / xs
        d2_dense   = pd2(x_dense_n)
        d3_dense   = pd3(x_dense_n)

        # Inflection points: zero-crossings of d2 where d3 > 0
        zc = np.where(np.diff(np.sign(d2_dense)))[0]
        inflections: list[float] = []
        for idx in zc:
            nxt = min(idx + 1, len(d3_dense) - 1)
            if (d3_dense[idx] + d3_dense[nxt]) / 2 > 0:
                t_infl = float(np.interp(
                    0,
                    [d2_dense[idx], d2_dense[min(idx + 1, len(d2_dense) - 1)]],
                    [x_dense_ms[idx], x_dense_ms[min(idx + 1, len(x_dense_ms) - 1)]],
                ))
                inflections.append(round(t_infl, 6))

        result[fname] = {
            'poly_oj_time_ms': [round(float(v), 6) for v in x_dense_ms],
            'poly_oj_d2':      [_safe(v) for v in d2_dense],
            'poly_infl_ms':    inflections,
        }
    return result


# ─── New FJ/FI detection fitting functions ──────────────────────────────────

def _fit_three_exponential(double_norm_df: pd.DataFrame, x_col: str,
                           ms_factor: float, n_dense: int = 500) -> dict:
    """Fit V(t) = A_OJ*(1-exp(-t/τ_OJ)) + A_JI*(1-exp(-t/τ_JI)) + A_IP*(1-exp(-t/τ_IP)).

    Three saturating exponentials decompose the OJIP rise into O-J, J-I, I-P
    kinetic phases (Boisvert, Joly & Carpentier 2006, FEBS J.).  Phase
    boundaries FJ and FI are identified at the rate-crossover points where
    the derivative of one component falls below the next.

    Returns ``{fname: {...}}`` or ``{fname: None}`` on fit failure.
    """
    dn   = double_norm_df
    cols = dn.columns
    x_all = np.asarray(dn.iloc[:, 0].values, dtype=float)

    def _three_exp(t, a1, tau1, a2, tau2, a3, tau3):
        return (a1 * (1.0 - np.exp(-t / tau1))
                + a2 * (1.0 - np.exp(-t / tau2))
                + a3 * (1.0 - np.exp(-t / tau3)))

    result: dict = {}
    for i in range(1, len(cols)):
        fname = cols[i]
        y_all = np.array(pd.to_numeric(dn.iloc[:, i], errors='coerce'), dtype=float)
        mask  = np.isfinite(y_all) & np.isfinite(x_all) & (x_all > 0)
        x_ms  = x_all[mask] * ms_factor
        y     = y_all[mask]

        if len(x_ms) < 8:
            result[fname] = None
            continue

        # Initial guesses and bounds
        p0 = [0.5, 0.5, 0.3, 5.0, 0.2, 50.0]
        bounds = ([0, 0.01, 0, 0.01, 0, 0.01],
                  [1.5, 500, 1.5, 500, 1.5, 500])
        try:
            popt, _ = curve_fit(_three_exp, x_ms, y, p0=p0, bounds=bounds,
                                maxfev=10000)
            a1, tau1, a2, tau2, a3, tau3 = popt

            # Enforce ordering: τ1 < τ2 < τ3 (swap if needed)
            params = sorted([(tau1, a1), (tau2, a2), (tau3, a3)])
            tau1, a1 = params[0]
            tau2, a2 = params[1]
            tau3, a3 = params[2]

            # Dense evaluation grid
            t_dense = np.geomspace(max(x_ms[0], 0.01), x_ms[-1], n_dense)
            fit_total = _three_exp(t_dense, a1, tau1, a2, tau2, a3, tau3)
            fit_oj = a1 * (1.0 - np.exp(-t_dense / tau1))
            fit_ji = a2 * (1.0 - np.exp(-t_dense / tau2))
            fit_ip = a3 * (1.0 - np.exp(-t_dense / tau3))

            # FJ/FI from rate crossovers: d(component)/dt
            rate_oj = (a1 / tau1) * np.exp(-t_dense / tau1)
            rate_ji = (a2 / tau2) * np.exp(-t_dense / tau2)
            rate_ip = (a3 / tau3) * np.exp(-t_dense / tau3)

            # FJ: where rate_oj falls below rate_ji
            diff_oj_ji = rate_oj - rate_ji
            zc_fj = np.where(np.diff(np.sign(diff_oj_ji)) < 0)[0]
            fj_ms = float(np.interp(0, [diff_oj_ji[zc_fj[0]], diff_oj_ji[zc_fj[0]+1]],
                                    [t_dense[zc_fj[0]], t_dense[zc_fj[0]+1]])) if len(zc_fj) > 0 else None

            # FI: where rate_ji falls below rate_ip
            diff_ji_ip = rate_ji - rate_ip
            zc_fi = np.where(np.diff(np.sign(diff_ji_ip)) < 0)[0]
            fi_ms = float(np.interp(0, [diff_ji_ip[zc_fi[0]], diff_ji_ip[zc_fi[0]+1]],
                                    [t_dense[zc_fi[0]], t_dense[zc_fi[0]+1]])) if len(zc_fi) > 0 else None

            # R² goodness of fit
            y_pred = _three_exp(x_ms, a1, tau1, a2, tau2, a3, tau3)
            ss_res = float(np.sum((y - y_pred)**2))
            ss_tot = float(np.sum((y - np.mean(y))**2))
            r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0

            result[fname] = {
                'A_OJ': round(float(a1), 6), 'A_JI': round(float(a2), 6),
                'A_IP': round(float(a3), 6),
                'tau_OJ_ms': round(float(tau1), 4), 'tau_JI_ms': round(float(tau2), 4),
                'tau_IP_ms': round(float(tau3), 4),
                'fj_ms': round(fj_ms, 4) if fj_ms is not None else None,
                'fi_ms': round(fi_ms, 4) if fi_ms is not None else None,
                'fit_t_ms':  [round(float(v), 6) for v in t_dense],
                'fit_total': [_safe(v) for v in fit_total],
                'fit_oj':    [_safe(v) for v in fit_oj],
                'fit_ji':    [_safe(v) for v in fit_ji],
                'fit_ip':    [_safe(v) for v in fit_ip],
                'r2': round(r2, 6),
            }
        except Exception:
            result[fname] = None
    return result


def _fit_piecewise_linear(double_norm_df: pd.DataFrame, x_col: str,
                          ms_factor: float, n_dense: int = 500) -> dict:
    """Fit a piecewise-linear model to V(log₁₀(t)) with 3 optimised breakpoints.

    The OJIP transient in log-time is approximately piecewise-linear with
    four segments separated at FJ, FI, FP.  Breakpoint positions are
    optimised to minimise total squared residuals; per-segment slopes are
    solved analytically (LSQ) for each candidate breakpoint set.

    Returns ``{fname: {...}}`` or ``{fname: None}`` on failure.
    """
    from scipy.optimize import minimize

    dn   = double_norm_df
    cols = dn.columns
    x_all = np.asarray(dn.iloc[:, 0].values, dtype=float)

    def _pw_residual(breakpoints_log, x_log, y):
        """Sum-of-squares for 4 linear segments defined by 3 sorted breakpoints."""
        bp = np.sort(breakpoints_log)
        edges = np.concatenate([[x_log[0] - 0.01], bp, [x_log[-1] + 0.01]])
        ss = 0.0
        for seg in range(4):
            mask = (x_log >= edges[seg]) & (x_log < edges[seg + 1])
            if seg == 3:  # last segment includes endpoint
                mask = (x_log >= edges[seg]) & (x_log <= edges[seg + 1])
            xs = x_log[mask]
            ys = y[mask]
            if len(xs) < 2:
                ss += 1e6  # penalty for empty segments
                continue
            # Analytic LSQ line fit
            coeffs = np.polyfit(xs, ys, 1)
            ss += float(np.sum((ys - np.polyval(coeffs, xs))**2))
        return ss

    result: dict = {}
    for i in range(1, len(cols)):
        fname = cols[i]
        y_all = np.array(pd.to_numeric(dn.iloc[:, i], errors='coerce'), dtype=float)
        mask  = np.isfinite(y_all) & np.isfinite(x_all) & (x_all > 0)
        x_ms  = x_all[mask] * ms_factor
        y     = y_all[mask]

        if len(x_ms) < 8:
            result[fname] = None
            continue

        x_log = np.log10(np.clip(x_ms, 1e-6, None))

        # Initial breakpoint guesses in log10(ms)
        bp0 = np.array([np.log10(2.0), np.log10(30.0), np.log10(300.0)])
        # Clamp to data range
        bp0 = np.clip(bp0, x_log[1], x_log[-2])

        try:
            res = minimize(_pw_residual, bp0, args=(x_log, y),
                           method='Nelder-Mead',
                           options={'maxiter': 5000, 'xatol': 1e-4, 'fatol': 1e-8})
            bp_opt = np.sort(res.x)
            bp_ms  = 10.0 ** bp_opt  # convert back to ms

            # Compute per-segment slopes on optimal breakpoints
            edges = np.concatenate([[x_log[0] - 0.01], bp_opt, [x_log[-1] + 0.01]])
            slopes = []
            for seg in range(4):
                seg_mask = (x_log >= edges[seg]) & (x_log < edges[seg + 1])
                if seg == 3:
                    seg_mask = (x_log >= edges[seg]) & (x_log <= edges[seg + 1])
                xs = x_log[seg_mask]
                ys = y[seg_mask]
                if len(xs) >= 2:
                    slopes.append(round(float(np.polyfit(xs, ys, 1)[0]), 6))
                else:
                    slopes.append(None)

            # Build fitted piecewise curve on dense grid for plotting
            x_log_dense = np.linspace(x_log[0], x_log[-1], n_dense)
            y_dense = np.zeros(n_dense)
            for seg in range(4):
                seg_mask_d = (x_log_dense >= edges[seg]) & (x_log_dense < edges[seg + 1])
                if seg == 3:
                    seg_mask_d = (x_log_dense >= edges[seg]) & (x_log_dense <= edges[seg + 1])
                seg_mask_data = (x_log >= edges[seg]) & (x_log < edges[seg + 1])
                if seg == 3:
                    seg_mask_data = (x_log >= edges[seg]) & (x_log <= edges[seg + 1])
                xs_d = x_log[seg_mask_data]
                ys_d = y[seg_mask_data]
                if len(xs_d) >= 2:
                    coeffs = np.polyfit(xs_d, ys_d, 1)
                    y_dense[seg_mask_d] = np.polyval(coeffs, x_log_dense[seg_mask_d])

            t_dense_ms = 10.0 ** x_log_dense

            # R² goodness of fit
            y_pred_all = np.zeros_like(y)
            for seg in range(4):
                seg_mask_a = (x_log >= edges[seg]) & (x_log < edges[seg + 1])
                if seg == 3:
                    seg_mask_a = (x_log >= edges[seg]) & (x_log <= edges[seg + 1])
                xs_a = x_log[seg_mask_a]
                ys_a = y[seg_mask_a]
                if len(xs_a) >= 2:
                    y_pred_all[seg_mask_a] = np.polyval(np.polyfit(xs_a, ys_a, 1), xs_a)
            ss_res = float(np.sum((y - y_pred_all)**2))
            ss_tot = float(np.sum((y - np.mean(y))**2))
            r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0

            result[fname] = {
                'breakpoints_ms': [round(float(v), 4) for v in bp_ms],
                'segment_slopes': slopes,
                'fit_t_ms':  [round(float(v), 6) for v in t_dense_ms],
                'fit_y':     [_safe(v) for v in y_dense],
                'r2': round(r2, 6),
            }
        except Exception:
            result[fname] = None
    return result


def _fit_d1_gaussians(D1_DF: pd.DataFrame, x_col: str, ms_factor: float,
                      n_dense: int = 500) -> dict:
    """Decompose D1 = dV/d(log₁₀ t) into 2–3 Gaussian peaks.

    The D1 curve of a typical OJIP transient shows 2–3 peaks corresponding
    to the O-J, J-I, and I-P rate maxima.  Fitting Gaussians in log₁₀(t_ms)
    space produces peak centres, widths, and amplitudes for each phase.  The
    valleys (local minima of the fitted sum) between adjacent peaks give the
    FJ and FI timings.

    Attempts 3 Gaussians first; falls back to 2 if the third is negligible
    (amplitude < 5 % of max) or if the 3-Gaussian fit fails.

    Returns ``{fname: {...}}`` or ``{fname: None}`` on failure.
    """
    t_all = np.asarray(D1_DF.iloc[:, 0].values, dtype=float)

    def _sum_gaussians(x, *params):
        n_g = len(params) // 3
        y = np.zeros_like(x, dtype=float)
        for g in range(n_g):
            a, c, s = params[3*g], params[3*g+1], params[3*g+2]
            y += a * np.exp(-(x - c)**2 / (2 * s**2))
        return y

    result: dict = {}
    for i in range(1, len(D1_DF.columns)):
        fname = D1_DF.columns[i]
        d1 = np.asarray(D1_DF.iloc[:, i].values, dtype=float)
        mask = np.isfinite(d1) & np.isfinite(t_all) & (t_all > 0)
        t_ms = t_all[mask] * ms_factor
        d1v  = d1[mask]

        if len(t_ms) < 8:
            result[fname] = None
            continue

        x_log = np.log10(np.clip(t_ms, 1e-6, None))
        x_dense = np.linspace(x_log[0], x_log[-1], n_dense)

        # Initial guesses: 3 Gaussians at ~O-J, ~J-I, ~I-P in log10(ms)
        d1_max = float(np.nanmax(d1v)) if np.any(d1v > 0) else 1.0
        p0_3 = [d1_max * 0.6, np.log10(0.3), 0.3,   # G1: O-J peak
                d1_max * 0.3, np.log10(3.0), 0.3,    # G2: J-I peak
                d1_max * 0.2, np.log10(30.0), 0.3]   # G3: I-P peak
        lb_3 = [0, x_log[0], 0.05] * 3
        ub_3 = [d1_max * 3, x_log[-1], 2.0] * 3

        fit_ok = False
        n_gaussians = 3

        try:
            popt, _ = curve_fit(_sum_gaussians, x_log, d1v, p0=p0_3,
                                bounds=(lb_3, ub_3), maxfev=10000)
            # Check if 3rd Gaussian is negligible
            amps = [abs(popt[0]), abs(popt[3]), abs(popt[6])]
            if amps[2] < 0.05 * max(amps):
                raise ValueError('Third Gaussian negligible')
            fit_ok = True
        except Exception:
            n_gaussians = 2
            # Fall back to 2 Gaussians
            p0_2 = [d1_max * 0.6, np.log10(0.3), 0.3,
                    d1_max * 0.3, np.log10(10.0), 0.5]
            lb_2 = [0, x_log[0], 0.05] * 2
            ub_2 = [d1_max * 3, x_log[-1], 2.0] * 2
            try:
                popt, _ = curve_fit(_sum_gaussians, x_log, d1v, p0=p0_2,
                                    bounds=(lb_2, ub_2), maxfev=10000)
                fit_ok = True
            except Exception:
                pass

        if not fit_ok:
            result[fname] = None
            continue

        # Sort Gaussians by centre position
        gauss_list = []
        for g in range(n_gaussians):
            gauss_list.append((popt[3*g+1], popt[3*g], popt[3*g+2]))  # (center, amp, sigma)
        gauss_list.sort(key=lambda x: x[0])

        # Dense evaluation
        fit_total = _sum_gaussians(x_dense, *popt)
        fit_components = []
        centers_ms, sigmas, amplitudes = [], [], []
        for center, amp, sigma in gauss_list:
            g_curve = amp * np.exp(-(x_dense - center)**2 / (2 * sigma**2))
            fit_components.append(g_curve)
            centers_ms.append(round(float(10**center), 4))
            sigmas.append(round(float(sigma), 4))
            amplitudes.append(round(float(amp), 6))

        # Find valleys (FJ, FI) as local minima of fitted sum between peaks
        valleys_ms = []
        for v in range(len(gauss_list) - 1):
            c1, c2 = gauss_list[v][0], gauss_list[v + 1][0]
            vmask = (x_dense >= c1) & (x_dense <= c2)
            if vmask.any():
                v_idx = np.argmin(fit_total[vmask])
                valleys_ms.append(round(float(10**x_dense[vmask][v_idx]), 4))

        fj_ms = valleys_ms[0] if len(valleys_ms) >= 1 else None
        fi_ms = valleys_ms[1] if len(valleys_ms) >= 2 else None

        # R²
        y_pred = _sum_gaussians(x_log, *popt)
        ss_res = float(np.sum((d1v - y_pred)**2))
        ss_tot = float(np.sum((d1v - np.mean(d1v))**2))
        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0

        t_dense_ms = [round(float(10**v), 6) for v in x_dense]
        entry = {
            'n_gaussians': n_gaussians,
            'centers_ms': centers_ms, 'sigmas': sigmas, 'amplitudes': amplitudes,
            'fj_ms': fj_ms, 'fi_ms': fi_ms,
            'fit_t_ms':  t_dense_ms,
            'fit_total': [_safe(v) for v in fit_total],
            'r2': round(r2, 6),
        }
        for g_idx, g_curve in enumerate(fit_components):
            entry[f'fit_g{g_idx+1}'] = [_safe(v) for v in g_curve]
        result[fname] = entry
    return result


def _select_d1_min_pos(d1_values, t_values, expect=None, after=None,
                       prom_frac: float = 0.15):
    """Choose one D1 local minimum (plateau) within a search window.

    D1 = dV/d(log t).  Local minima of D1 correspond to plateaus where the
    fluorescence rise rate is slowest — i.e. the J and I steps.  This is
    complementary to the D2 trough approach: D2 troughs mark inflection
    points (curvature reversal), while D1 minima mark rate minima (plateaus).

    Uses the same prominence + log-time proximity scoring as
    ``_select_trough_pos()``.

    Returns ``(pos, confidence)``: *pos* is the positional index into the
    segment (−1 if unusable); *confidence* ∈ [0, 1].
    """
    d1 = np.asarray(d1_values, dtype=float)
    t  = np.asarray(t_values,  dtype=float)
    if d1.size == 0 or np.all(np.isnan(d1)):
        return -1, 0.0

    troughs, props = find_peaks(-d1, prominence=0)
    if troughs.size == 0:
        return int(np.nanargmin(d1)), 0.0

    proms = props['prominences']
    pmax  = float(proms.max()) or 1.0

    if expect is not None:
        _NOISE_FLOOR = 0.03
        keep  = proms >= _NOISE_FLOOR * pmax
        cand  = troughs[keep]
        cprom = proms[keep]

        if after is not None:
            mask = t[cand] > after
            if mask.any():
                cand, cprom = cand[mask], cprom[mask]

        if cand.size == 0:
            return int(np.nanargmin(d1)), 0.0

        _SIGMA = 0.5
        _ALPHA = 0.3
        log_t   = np.log10(np.clip(t[cand], 1e-12, None))
        log_exp = np.log10(max(float(expect), 1e-12))
        log_dist = np.abs(log_t - log_exp)
        proximity = np.exp(-log_dist**2 / (2 * _SIGMA**2))

        prom_norm = cprom / pmax
        score = _ALPHA * prom_norm + (1 - _ALPHA) * proximity
        j = int(np.argmax(score))

        cand_max = float(cprom.max()) or 1.0
        prom_conf = float(cprom[j] / cand_max)
        confidence = prom_conf * float(proximity[j])
        return int(cand[j]), confidence
    else:
        keep  = proms >= prom_frac * pmax
        cand  = troughs[keep]
        cprom = proms[keep]

        if after is not None:
            mask = t[cand] > after
            if mask.any():
                cand, cprom = cand[mask], cprom[mask]

        if cand.size == 0:
            return int(np.nanargmin(d1)), 0.0

        j = int(np.argmax(cprom))
        cand_max = float(cprom.max()) or 1.0
        return int(cand[j]), float(cprom[j] / cand_max)


def _select_trough_pos(d2_values, t_values, expect=None, after=None,
                       prom_frac: float = 0.15):
    """Choose one D2 local minimum (trough) within a search window, per-curve.

    A phase timing (FJ/FI/FP) marks a genuine deceleration trough — D2 declines
    and then rises again. The plain argmin of the window instead returns whatever
    is lowest, which on non-standard transients is often a value on a monotonic
    slope at the window edge, bleeding in from the neighbouring phase (e.g. a fast
    curve whose FP deceleration drags the FI window's minimum out to ~100 ms even
    though a real, shallower FI trough sits near 25 ms).

    Two selection modes:

    **When ``expect`` is given** (FJ, FI): a combined score balances prominence
    and proximity to the expected phase time.  The noise floor is lowered (3 %
    instead of ``prom_frac``) so weak-but-correctly-positioned troughs — e.g. the
    true J step at ~2 ms whose prominence is dwarfed by the nearby K step at
    ~0.3 ms — are not discarded.  A Gaussian proximity kernel in log₁₀-time
    (σ = 0.5 decades ≈ factor of 3) is combined with normalized prominence:

        score = α × prom_norm + (1 − α) × exp(−log_dist² / 2σ²)

    with α = 0.3 (prominence counts 30 %, proximity 70 %).  Confidence is the
    product of the prominence ratio and the proximity factor, so a trough that is
    the sole candidate but far from ``expect`` gets reduced confidence.

    **When ``expect`` is None** (FP): the original logic — prominence floor at
    ``prom_frac``, then the most prominent trough wins.

    In both modes:
      • ``after`` (a time) enforces the biological ordering FJ < FI < FP by
        dropping troughs at/before the previously-assigned phase.
      • Falls back to the plain minimum when the window has no interior trough.

    Returns (pos, confidence): pos is the positional index into the segment (−1
    if unusable); confidence ∈ [0, 1].
    """
    d2 = np.asarray(d2_values, dtype=float)
    t  = np.asarray(t_values,  dtype=float)
    if d2.size == 0 or np.all(np.isnan(d2)):
        return -1, 0.0

    troughs, props = find_peaks(-d2, prominence=0)
    if troughs.size == 0:
        # No interior trough: genuinely monotonic window — fall back to argmin.
        return int(np.nanargmin(d2)), 0.0

    proms = props['prominences']
    pmax  = float(proms.max()) or 1.0

    if expect is not None:
        # ── Weighted scoring: prominence × proximity ──────────────────────
        # Use a low noise floor so weak-but-near-expected troughs survive.
        _NOISE_FLOOR = 0.03
        keep  = proms >= _NOISE_FLOOR * pmax
        cand  = troughs[keep]
        cprom = proms[keep]

        if after is not None:
            mask = t[cand] > after
            if mask.any():
                cand, cprom = cand[mask], cprom[mask]

        if cand.size == 0:
            return int(np.nanargmin(d2)), 0.0

        # Gaussian proximity in log₁₀-time (σ = 0.5 decades).
        _SIGMA = 0.5
        _ALPHA = 0.3   # prominence weight (0.3 prom + 0.7 proximity)
        log_t   = np.log10(np.clip(t[cand], 1e-12, None))
        log_exp = np.log10(max(float(expect), 1e-12))
        log_dist = np.abs(log_t - log_exp)
        proximity = np.exp(-log_dist**2 / (2 * _SIGMA**2))

        prom_norm = cprom / pmax
        score = _ALPHA * prom_norm + (1 - _ALPHA) * proximity
        j = int(np.argmax(score))

        # Confidence: prominence ratio × proximity factor.
        cand_max = float(cprom.max()) or 1.0
        prom_conf = float(cprom[j] / cand_max)
        confidence = prom_conf * float(proximity[j])
        return int(cand[j]), confidence
    else:
        # ── No expected position (FP): most prominent trough wins ─────────
        keep  = proms >= prom_frac * pmax
        cand  = troughs[keep]
        cprom = proms[keep]

        if after is not None:
            mask = t[cand] > after
            if mask.any():
                cand, cprom = cand[mask], cprom[mask]

        if cand.size == 0:
            return int(np.nanargmin(d2)), 0.0

        j = int(np.argmax(cprom))
        cand_max = float(cprom.max()) or 1.0
        return int(cand[j]), float(cprom[j] / cand_max)


_DETECTION_METHODS = frozenset({
    'polynomial', 'd1_minima', 'three_exp', 'piecewise', 'gaussian_d1',
})


def _find_fjfifp(D2_DF, D3_DF, x_col, ranges, file_names, Infl_DF,
                 fj_expect=None, fi_expect=None,
                 poly_oj=None, poly_oi=None, ms_factor=1.0,
                 method='d2_trough',
                 D1_DF=None,
                 three_exp_results=None,
                 piecewise_results=None,
                 gaussian_d1_results=None):
    """
    Identify FJ/FI/FP timings using one of several detection strategies.

    ``method`` selects the detection algorithm:

    - ``'d2_trough'`` — prominence-ranked D2 local minima (default; used by
      spline/logspline/pchip).
    - ``'polynomial'`` — polynomial inflection points (Akinyemi et al. 2023).
    - ``'d1_minima'`` — D1 local minima (plateaus in dV/d(log t)).
    - ``'three_exp'`` — pre-computed 3-exponential rate crossovers.
    - ``'piecewise'`` — pre-computed piecewise-linear breakpoints.
    - ``'gaussian_d1'`` — pre-computed Gaussian D1 deconvolution valleys.

    All methods fall back to D2 trough for FP (and for FJ/FI when the
    primary method fails).

    Returns eight objects: six pd.Series (FJ/FI/FP_deriv and _infl),
    ``conf`` (dict of three pd.Series), and ``method_extras`` (dict of
    per-curve method-specific visualisation data).
    """
    t = D2_DF[x_col]

    def range_idx(lo, hi):
        return t.sub(lo).abs().idxmin(), t.sub(hi).abs().idxmin()

    windows = {
        'FJ': range_idx(*ranges['FJ']),
        'FI': range_idx(*ranges['FI']),
        'FP': range_idx(*ranges['FP']),
    }
    expect_map = {'FJ': fj_expect, 'FI': fi_expect, 'FP': None}

    # ── Polynomial inflection helper (unchanged) ─────────────────────────
    def _pick_poly_infl(poly_dict, fname, expect_native, ms):
        if poly_dict is None:
            return None, 0.0
        entry = poly_dict.get(fname)
        if not entry:
            return None, 0.0
        infls = entry.get('poly_infl_ms', [])
        if not infls:
            return None, 0.0
        infls = np.array(infls, dtype=float)
        if expect_native is not None and ms > 0:
            expect_ms = float(expect_native) * ms
            log_dist = np.abs(np.log10(np.clip(infls, 1e-12, None))
                              - np.log10(max(expect_ms, 1e-12)))
            j = int(np.argmin(log_dist))
            conf = float(np.exp(-log_dist[j]**2 / (2 * 0.5**2)))
        else:
            j = 0
            conf = 0.5
        native_t = infls[j] / ms
        return native_t, conf

    # ── Helper: read pre-computed FJ/FI from a method results dict ───────
    def _read_precomputed(res_dict, fname, key, ms):
        """Read FJ or FI from a pre-computed results dict.

        Returns (native_time, confidence) or (None, 0.0).
        """
        if res_dict is None:
            return None, 0.0
        entry = res_dict.get(fname)
        if entry is None:
            return None, 0.0
        field = 'fj_ms' if key == 'FJ' else ('fi_ms' if key == 'FI' else None)
        if field is None:
            return None, 0.0
        val_ms = entry.get(field)
        if val_ms is None:
            return None, 0.0
        # Confidence from R² of the fit (rough proxy)
        r2 = entry.get('r2', 0.5)
        conf = max(0.0, min(1.0, float(r2)))
        return float(val_ms) / ms, conf

    def _read_piecewise(res_dict, fname, key, ms):
        """Read FJ/FI/FP from piecewise breakpoints."""
        if res_dict is None:
            return None, 0.0
        entry = res_dict.get(fname)
        if entry is None:
            return None, 0.0
        bp = entry.get('breakpoints_ms', [])
        idx = {'FJ': 0, 'FI': 1, 'FP': 2}.get(key)
        if idx is None or idx >= len(bp):
            return None, 0.0
        r2 = entry.get('r2', 0.5)
        conf = max(0.0, min(1.0, float(r2)))
        return float(bp[idx]) / ms, conf

    # ── D2 trough fallback ───────────────────────────────────────────────
    def _d2_trough_fallback(fname, key, prev_t, d2_col):
        lo, hi = windows[key]
        d2_seg = d2_col.loc[lo:hi]
        t_seg  = t.loc[lo:hi]
        pos, cf = _select_trough_pos(d2_seg.values, t_seg.values,
                                     expect=expect_map[key], after=prev_t)
        if pos < 0:
            raise ValueError(
                'Could not identify phase timing — check data integrity.')
        return float(t_seg.iloc[pos]), cf

    # ── Main loop ────────────────────────────────────────────────────────
    results = {k: [] for k in ('FJ', 'FI', 'FP')}
    confs   = {k: [] for k in ('FJ', 'FI', 'FP')}
    method_extras: dict = {}

    for i in range(1, len(D2_DF.columns)):
        fname  = D2_DF.columns[i]
        d2_col = D2_DF.iloc[:, i]
        prev_t = None

        for key in ('FJ', 'FI', 'FP'):
            tv, cf = None, 0.0

            # ── Method-specific primary detection ────────────────────────
            if method == 'polynomial':
                poly_src = poly_oj if key == 'FJ' else (poly_oi if key == 'FI' else None)
                tv, cf = _pick_poly_infl(poly_src, fname, expect_map[key], ms_factor)

            elif method == 'd1_minima' and D1_DF is not None and key != 'FP':
                lo, hi = windows[key]
                d1_col = D1_DF.iloc[:, i]
                d1_seg = d1_col.loc[lo:hi]
                t_seg  = t.loc[lo:hi]
                pos, cf = _select_d1_min_pos(d1_seg.values, t_seg.values,
                                             expect=expect_map[key], after=prev_t)
                if pos >= 0:
                    tv = float(t_seg.iloc[pos])

            elif method == 'three_exp' and key != 'FP':
                tv, cf = _read_precomputed(three_exp_results, fname, key, ms_factor)

            elif method == 'piecewise':
                tv, cf = _read_piecewise(piecewise_results, fname, key, ms_factor)

            elif method == 'gaussian_d1' and key != 'FP':
                tv, cf = _read_precomputed(gaussian_d1_results, fname, key, ms_factor)

            # ── Ordering check + D2 trough fallback ──────────────────────
            if tv is not None and prev_t is not None and tv <= prev_t:
                tv = None  # ordering violated → fall back

            if tv is None:
                tv, cf = _d2_trough_fallback(fname, key, prev_t, d2_col)

            results[key].append(tv)
            confs[key].append(cf)
            prev_t = tv

        # Store method-specific visualisation data per curve
        if method == 'three_exp' and three_exp_results:
            entry = three_exp_results.get(fname)
            if entry:
                method_extras[fname] = entry
        elif method == 'piecewise' and piecewise_results:
            entry = piecewise_results.get(fname)
            if entry:
                method_extras[fname] = entry
        elif method == 'gaussian_d1' and gaussian_d1_results:
            entry = gaussian_d1_results.get(fname)
            if entry:
                method_extras[fname] = entry

    FJ_deriv = pd.Series(results['FJ'], index=file_names)
    FI_deriv = pd.Series(results['FI'], index=file_names)
    FP_deriv = pd.Series(results['FP'], index=file_names)
    conf = {k: pd.Series(confs[k], index=file_names) for k in ('FJ', 'FI', 'FP')}

    def nearest_inflect(deriv_ser):
        return pd.Series({
            col: Infl_DF[col][Infl_DF[col] > val].min()
            for col, val in deriv_ser.items()
        })

    return (FJ_deriv, FI_deriv, FP_deriv,
            nearest_inflect(FJ_deriv), nearest_inflect(FI_deriv), nearest_inflect(FP_deriv),
            conf, method_extras)


def _calc_areas_fm_timing(Summary_file, data_cols, FJ_idx, FI_idx, ms_factor,
                          F50ms_idx, F100ms_idx, F200ms_idx, F300ms_idx):
    """
    Compute complementary areas (OJ, JI, IP, OP) and FM timing per sample.
    Returns (AREAOJ, AREAJI, AREAIP, AREAOP, FM_timings_series) as pd.Series indexed by data_cols.
    """
    aoj, aji, aip, aop, fm_t = [], [], [], [], {}

    for i, col in enumerate(data_cols, start=1):
        F50ms_v = Summary_file.iloc[F50ms_idx, i]
        F100ms_v = Summary_file.iloc[F100ms_idx, i]
        F200ms_v = Summary_file.iloc[F200ms_idx, i]
        F300ms_v = Summary_file.iloc[F300ms_idx, i]

        if float(F100ms_v) < float(F50ms_v):
            Fm_val = Summary_file.iloc[F100ms_idx:, i].max()
            Fm_idx = Summary_file.iloc[F100ms_idx:, i].sub(Fm_val).abs().idxmin()
            if float(F200ms_v) < float(F100ms_v):
                Fm_val = Summary_file.iloc[F200ms_idx:, i].max()
                Fm_idx = Summary_file.iloc[F200ms_idx:, i].sub(Fm_val).abs().idxmin()
                if float(F300ms_v) < float(F200ms_v):
                    Fm_val = Summary_file.iloc[F300ms_idx:, i].max()
                    Fm_idx = Summary_file.iloc[F300ms_idx:, i].sub(Fm_val).abs().idxmin()
        else:
            Fm_val = Summary_file.iloc[F50ms_idx:, i].max()
            Fm_idx = Summary_file.iloc[F50ms_idx:, i].sub(Fm_val).abs().idxmin()

        fm_t[col] = float(Summary_file.iloc[Fm_idx, 0]) * ms_factor

        x = Summary_file.iloc[:, 0] * ms_factor   # convert to ms so areas are in [ms × F]
        y = Summary_file.iloc[:, i]
        FM_raw = y.iloc[:Fm_idx].max()

        aoj.append(max(x.iloc[:FJ_idx]) * FM_raw - np.trapz(y.iloc[:FJ_idx], x.iloc[:FJ_idx]))
        aji.append((max(x.iloc[FJ_idx:FI_idx]) - min(x.iloc[FJ_idx:FI_idx])) * FM_raw
                   - np.trapz(y.iloc[FJ_idx:FI_idx], x.iloc[FJ_idx:FI_idx]))
        aip.append((max(x.iloc[FI_idx:Fm_idx]) - min(x.iloc[FI_idx:Fm_idx])) * FM_raw
                   - np.trapz(y.iloc[FI_idx:Fm_idx], x.iloc[FI_idx:Fm_idx]))
        aop.append(max(x.iloc[:Fm_idx]) * FM_raw - np.trapz(y.iloc[:Fm_idx], x.iloc[:Fm_idx]))

    return (pd.Series(aoj, index=data_cols), pd.Series(aji, index=data_cols),
            pd.Series(aip, index=data_cols), pd.Series(aop, index=data_cols),
            pd.Series(fm_t))


# ─── O-J densify model fitters ───────────────────────────────────────────
# Each returns (predict_fn, fit_info_dict) or raises RuntimeError/ValueError.

def _fit_oj_exponential(t_oj, y_oj, params):
    """F(t) = A * (1 - exp(-t/tau)).  The p=0 (separate PSII units) case."""
    tau_ms = params.get('tau_ms') if params else None

    def _model(t, A, tau):
        return A * (1.0 - np.exp(-t / tau))

    if tau_ms is not None and tau_ms > 0:
        def _fixed(t, A):
            return A * (1.0 - np.exp(-t / tau_ms))
        popt, _ = curve_fit(_fixed, t_oj, y_oj,
                            p0=[0.5], bounds=([0.01], [2.0]), maxfev=2000)
        A_fit, tau_fit = float(popt[0]), float(tau_ms)
    else:
        popt, _ = curve_fit(_model, t_oj, y_oj, p0=[0.5, 1.0],
                            bounds=([0.01, 0.01], [2.0, 50.0]), maxfev=2000)
        A_fit, tau_fit = float(popt[0]), float(popt[1])

    def predict(t):
        return _model(t, A_fit, tau_fit)

    return predict, {'model': 'exponential',
                     'tau_ms': round(tau_fit, 4), 'A': round(A_fit, 4)}


def _fit_oj_biexponential(t_oj, y_oj, params):
    """F(t) = A1*(1-exp(-t/tau1)) + A2*(1-exp(-t/tau2)).

    Phenomenological two-component O-J rise (fast ~0.3 ms, slow ~2 ms).
    """
    params = params or {}
    tau1_user = params.get('tau1_ms')
    tau2_user = params.get('tau2_ms')

    def _model(t, A1, tau1, A2, tau2):
        return A1 * (1.0 - np.exp(-t / tau1)) + A2 * (1.0 - np.exp(-t / tau2))

    if (tau1_user and tau1_user > 0) and (tau2_user and tau2_user > 0):
        def _fixed(t, A1, A2):
            return _model(t, A1, tau1_user, A2, tau2_user)
        popt, _ = curve_fit(_fixed, t_oj, y_oj, p0=[0.3, 0.3],
                            bounds=([0.001, 0.001], [2.0, 2.0]), maxfev=3000)
        A1_fit, A2_fit = float(popt[0]), float(popt[1])
        tau1_fit, tau2_fit = float(tau1_user), float(tau2_user)
    else:
        popt, _ = curve_fit(_model, t_oj, y_oj, p0=[0.3, 0.3, 0.3, 2.0],
                            bounds=([0.001, 0.01, 0.001, 0.1],
                                    [2.0, 5.0, 2.0, 50.0]), maxfev=5000)
        A1_fit, tau1_fit, A2_fit, tau2_fit = [float(v) for v in popt]

    # Canonical order: tau1 < tau2
    if tau1_fit > tau2_fit:
        A1_fit, tau1_fit, A2_fit, tau2_fit = A2_fit, tau2_fit, A1_fit, tau1_fit

    def predict(t):
        return _model(t, A1_fit, tau1_fit, A2_fit, tau2_fit)

    return predict, {'model': 'biexponential',
                     'A1': round(A1_fit, 4), 'tau1_ms': round(tau1_fit, 4),
                     'A2': round(A2_fit, 4), 'tau2_ms': round(tau2_fit, 4)}


def _fit_oj_connectivity(t_oj, y_oj, params):
    """Joliot connectivity model (exciton-radical-pair equilibrium).

    C(t)     = 1 - exp(-k_L * t)             # fraction closed RCs
    C_eff(t) = C(t) * exp(-k_ox * t)         # with QA⁻ reoxidation
    V(t)     = C_eff*(1-p) / (1 - p*C_eff)   # relative variable fluorescence
    F(t)     = A * V(t)

    Ref: Guo et al. 2024, Plants 13(3):452; Joliot & Joliot 1964.
    When p=0, k_ox=0 this degenerates to the simple exponential.
    k_ox defaults to 0 (fixed) — negligible in the short O-J window.
    """
    params = params or {}
    p_user   = params.get('p')
    kL_user  = params.get('k_L')
    kox_user = params.get('k_ox', 0.0)  # default fixed at 0

    def _vt(t, k_L, p, k_ox):
        C = 1.0 - np.exp(-k_L * t)
        C_eff = C * np.exp(-k_ox * t)
        denom = 1.0 - p * C_eff
        denom = np.where(np.abs(denom) < 1e-12, 1e-12, denom)
        return C_eff * (1.0 - p) / denom

    def _model(t, A, k_L, p, k_ox):
        return A * _vt(t, k_L, p, k_ox)

    # Build free/fixed parameter sets:  [A, k_L, p, k_ox]
    p0_all = [0.5, 2.0, 0.4, 0.0]
    lo_all = [0.01, 0.1, 0.0, 0.0]
    hi_all = [2.0, 50.0, 0.85, 5.0]
    param_names = ['A', 'k_L', 'p', 'k_ox']

    fixed = {}
    if kox_user is not None:
        fixed['k_ox'] = float(kox_user)
    if p_user is not None:
        fixed['p'] = float(p_user)
    if kL_user is not None:
        fixed['k_L'] = float(kL_user)

    free_idx = [i for i, n in enumerate(param_names) if n not in fixed]

    if not free_idx:
        # Everything fixed — just fit A
        free_idx = [0]
        fixed.pop('A', None)

    free_p0 = [p0_all[i] for i in free_idx]
    free_lo = [lo_all[i] for i in free_idx]
    free_hi = [hi_all[i] for i in free_idx]

    def _partial(t, *free_vals):
        vals = list(p0_all)
        for i, fi in enumerate(free_idx):
            vals[fi] = free_vals[i]
        for name, fv in fixed.items():
            vals[param_names.index(name)] = fv
        return _model(t, *vals)

    popt, _ = curve_fit(_partial, t_oj, y_oj, p0=free_p0,
                        bounds=(free_lo, free_hi), maxfev=5000)

    all_vals = list(p0_all)
    for i, fi in enumerate(free_idx):
        all_vals[fi] = float(popt[i])
    for name, fv in fixed.items():
        all_vals[param_names.index(name)] = fv
    A_fit, kL_fit, p_fit, kox_fit = all_vals

    def predict(t):
        return _model(t, A_fit, kL_fit, p_fit, kox_fit)

    return predict, {'model': 'connectivity',
                     'A': round(A_fit, 4), 'k_L': round(kL_fit, 4),
                     'p': round(p_fit, 4), 'k_ox': round(kox_fit, 4)}


def _fit_oj_linear(t_oj, y_oj, params):
    """F(t) = a + b*t.  Simple linear extrapolation through the O-J rise."""
    coeffs = np.polyfit(t_oj, y_oj, 1)  # [slope, intercept]
    slope, intercept = float(coeffs[0]), float(coeffs[1])

    def predict(t):
        return intercept + slope * np.asarray(t, dtype=float)

    return predict, {'model': 'linear',
                     'slope': round(slope, 6), 'intercept': round(intercept, 6)}


_OJ_FITTERS = {
    'exponential':   _fit_oj_exponential,
    'biexponential': _fit_oj_biexponential,
    'connectivity':  _fit_oj_connectivity,
    'linear':        _fit_oj_linear,
}


def _oj_densify(x_log, y, fj_hi_log,
                model='exponential', model_params=None, synth_weight=0.3):
    """Inject model-based synthetic points into sparse O-J gaps.

    Fits the selected model to double-normalised O-J data and generates
    synthetic fill points within any gap > 0.5 decades.  Synthetic points
    are assigned a reduced weight for the downstream LSQ spline fit.

    Models: 'exponential' (default), 'biexponential', 'connectivity', 'linear'.

    Returns
    -------
    (x_aug, y_aug, w_aug, fit_info)
    fit_info is a dict with 'model' key + fitted params, or None.
    """
    GAP_THRESH = 0.5          # decades
    FILL_STEP  = 0.1          # decades between synthetic points

    oj_mask = x_log <= fj_hi_log
    if oj_mask.sum() < 3:
        return x_log, y, np.ones(len(x_log)), None

    oj_x = x_log[oj_mask]
    gaps = np.diff(oj_x)
    big_gaps = np.where(gaps > GAP_THRESH)[0]
    if len(big_gaps) == 0:
        return x_log, y, np.ones(len(x_log)), None

    t_oj = 10.0 ** oj_x
    y_oj = y[oj_mask]
    params = model_params or {}

    fitter = _OJ_FITTERS.get(model, _fit_oj_exponential)
    try:
        predict_fn, fit_info = fitter(t_oj, y_oj, params)
    except (RuntimeError, ValueError):
        # Fallback to exponential if the selected model fails
        if model != 'exponential':
            try:
                predict_fn, fit_info = _fit_oj_exponential(
                    t_oj, y_oj, {'tau_ms': params.get('tau_ms')})
                fit_info['_fallback'] = True
                fit_info['_original_model'] = model
            except (RuntimeError, ValueError):
                return x_log, y, np.ones(len(x_log)), None
        else:
            return x_log, y, np.ones(len(x_log)), None

    # Generate synthetic fill points in each gap
    fill_x_log = []
    for gi in big_gaps:
        lo_g, hi_g = oj_x[gi], oj_x[gi + 1]
        n_fill = max(1, int(np.round((hi_g - lo_g) / FILL_STEP)) - 1)
        fill_x_log.append(np.linspace(lo_g, hi_g, n_fill + 2)[1:-1])
    fill_x_log = np.concatenate(fill_x_log)

    if len(fill_x_log) == 0:
        return x_log, y, np.ones(len(x_log)), fit_info

    fill_t = 10.0 ** fill_x_log
    fill_y = predict_fn(fill_t)

    x_aug = np.concatenate([x_log, fill_x_log])
    y_aug = np.concatenate([y, fill_y])
    w_aug = np.concatenate([np.ones(len(x_log)),
                            np.full(len(fill_x_log), synth_weight)])
    order = np.argsort(x_aug)
    return x_aug[order], y_aug[order], w_aug[order], fit_info


def _fit_splines_log(double_norm_df: pd.DataFrame, x_col: str,
                     n_interior_knots: int = 10,
                     trim_first: int = 0, trim_last: int = 0,
                     knot_placement: str = 'hybrid',
                     oj_densify: bool = False,
                     oj_model: str = 'exponential',
                     oj_model_params: 'dict | None' = None) -> tuple:
    """
    Like _fit_splines but fits in log10(time) space.

    Knots are placed uniformly in log10 space so that each OJIP phase
    (O-J, J-I, I-P, P-decline) receives equal spline flexibility regardless
    of the linear-time density of the data points.  This greatly improves
    reconstruction quality for logarithmically-sampled instruments such as
    OJIPImaging (60 points over 0.1–2900 ms).

    A QUINTIC (k=5) spline is used so the analytic 2nd derivative is a smooth
    cubic and the 3rd derivative a smooth quadratic.  A cubic (k=3) spline would
    give a piecewise-LINEAR D2 and a piecewise-CONSTANT D3 — the latter renders
    as a visible staircase no matter how dense the evaluation grid, because the
    underlying function really is a step function.

    Two numerical safeguards keep the high-order derivatives realistic,
    especially in the tail used for FP detection:
      1. Interior knots are spread across the FULL log range (not pushed up
         against the first/last data point), so the boundary intervals stay wide
         and their derivatives don't explode.
      2. Derivatives are evaluated STRICTLY INSIDE the boundary knots.  Those
         knots sit at the data extremes with multiplicity k+1, and evaluating a
         high-order derivative exactly on them yields a huge numerical spike
         (order 1e13) that would otherwise hijack the FP D2-minimum search.

    trim_first / trim_last: exclude the first / last N data points from the
    spline fit (useful when the tail of OJIPImaging curves is poorly fitted).

    Returns an 8-tuple: the same 7 elements as _fit_splines, plus
    densify_info (dict mapping fname → fit_info when oj_densify is used).
    """
    dn = double_norm_df
    cols = dn.columns
    n_files = len(cols) - 1
    _t_start = _fit_start_index(dn, trim_first)
    _t_end   = len(dn) - trim_last if trim_last > 0 else len(dn)
    n_fit = _t_end - _t_start

    # Quintic where enough points support it; degrade for very short curves so
    # LSQUnivariateSpline always has k+1 points and valid derivative orders.
    k = 5 if n_fit >= 12 else (3 if n_fit >= 5 else max(1, n_fit - 2))
    # Cap interior knots so every knot interval keeps data (Schoenberg-Whitney).
    n_knots = max(1, min(n_interior_knots, n_fit // (k + 2)))

    x_fit = dn.iloc[_t_start:_t_end, 0].values.astype(float)
    x_log = np.log10(x_fit)
    lo, hi = float(x_log[0]), float(x_log[-1])

    # Interior knot placement (safeguard 1).
    HYBRID_GAP = 0.3   # decades — gap threshold triggering hybrid crossover

    if knot_placement == 'uniform':
        # Evenly spread across the full log range.
        knots_log = np.linspace(lo, hi, n_knots + 2)[1:-1]
    elif knot_placement == 'hybrid':
        # Auto-detect crossover: right edge of last big gap.  Below the
        # crossover, quantile knots cluster where sparse data exists; above,
        # uniform knots give even coverage for the well-sampled I-P tail.
        # On data with no large gap (AquaPen, MC-PAM), falls back to uniform.
        diffs = np.diff(x_log)
        big = np.where(diffs > HYBRID_GAP)[0]
        if len(big) == 0:
            knots_log = np.linspace(lo, hi, n_knots + 2)[1:-1]
        else:
            x_cross = float(x_log[big[-1] + 1])
            frac_below = (x_cross - lo) / (hi - lo)
            n_below = max(1, round(n_knots * frac_below))
            n_above = n_knots - n_below

            # Below crossover: quantile on sparse data
            mask_lo = x_log <= x_cross
            min_sep = (hi - lo) / (n_knots + 2) * 0.1
            if mask_lo.sum() >= 2 and n_below >= 1:
                raw_q = np.quantile(x_log[mask_lo],
                                    np.linspace(0, 1, n_below + 2)[1:-1])
                q_knots = [raw_q[0]]
                for kv in raw_q[1:]:
                    if kv - q_knots[-1] >= min_sep:
                        q_knots.append(kv)
            else:
                q_knots = []

            # Above crossover: uniform in well-sampled tail
            if n_above >= 1:
                u_knots = np.linspace(x_cross, hi,
                                      n_above + 2)[1:-1].tolist()
            else:
                u_knots = []

            merged = sorted(set(q_knots + u_knots))
            knots_log = (np.array(merged) if merged
                         else np.linspace(lo, hi, 3)[1:2])
    else:
        # 'quantile': place at data quantiles — for well-sampled instruments
        # whose data is ~uniform in log-space this ≈ np.linspace; for sparse
        # TOMI-3 data, knots cluster where data exists and avoid the
        # 0.1→0.5 ms gap where an unconstrained knot causes oscillation.
        raw_knots = np.quantile(x_log, np.linspace(0, 1, n_knots + 2)[1:-1])
        # Deduplicate: ensure minimum separation so LSQUnivariateSpline
        # doesn't get coincident knots (can happen with repeated timepoints).
        min_sep = (hi - lo) / (n_knots + 2) * 0.1
        knots_log = [raw_knots[0]]
        for kv in raw_knots[1:]:
            if kv - knots_log[-1] >= min_sep:
                knots_log.append(kv)
        knots_log = np.array(knots_log)

    # Dense grid for smooth chart lines and fine zero-crossing resolution, and
    # kept strictly inside the boundary knots (safeguard 2): small epsilon
    # offsets from the data extremes avoid derivative spikes at boundary knots
    # while covering almost the full fitted range — including the O-J region
    # needed for FJ detection on sparse TOMI-3 data.
    n_eval = max(600, n_fit)
    eps_lo = (float(x_log[1]) - float(x_log[0])) * 0.01
    eps_hi = (float(x_log[-1]) - float(x_log[-2])) * 0.01
    log_time = pd.Series(
        np.geomspace(10**(lo + eps_lo), 10**(hi - eps_hi), num=n_eval),
        name=cols[0])
    x_eval = np.log10(log_time.values.astype(float))

    Raw_recon_list, D1_list, D2_list, D3_list, Infl_list = [], [], [], [], []
    densify_info = {}

    for i in range(1, n_files + 1):
        fname = cols[i]
        y     = dn.iloc[_t_start:_t_end, i].values.astype(float)

        if oj_densify:
            fj_hi_log = np.log10(10.0)   # 10 ms — covers FJ for all instruments
            x_aug, y_aug, w_aug, fit_info = _oj_densify(
                x_log, y, fj_hi_log,
                model=oj_model, model_params=oj_model_params)
            model = LSQUnivariateSpline(x_aug, y_aug, knots_log, k=k, w=w_aug)
            if fit_info is not None:
                densify_info[fname] = fit_info
        else:
            model = LSQUnivariateSpline(x_log, y, knots_log, k=k)

        recon  = model(x_eval)
        # Analytic derivatives of the spline in log10(t) space.
        # Using model.derivative() avoids the compounding numerical noise of
        # np.gradient on a sparse grid, giving reliable D2 zero-crossing detection.
        # Guard the order for degraded (very short curve) fits where k < 2/3.
        d1 = model.derivative(1)(x_eval) if k >= 1 else np.zeros_like(x_eval)
        d2 = model.derivative(2)(x_eval) if k >= 2 else np.zeros_like(x_eval)
        d3 = model.derivative(3)(x_eval) if k >= 3 else np.zeros_like(x_eval)
        # Upward zero-crossings only: D2 goes from - to + (D3 > 0 at crossing)
        # These are the inflection points per Akinyemi et al. 2025
        zc   = np.where(np.diff(np.sign(d2)) > 0)[0]
        infl = pd.Series(log_time.values[zc]).reset_index(drop=True).rename(fname)

        Raw_recon_list.append(pd.Series(recon, name=fname))
        D1_list.append(pd.Series(d1,    name=fname))
        D2_list.append(pd.Series(d2,    name=fname))
        D3_list.append(pd.Series(d3,    name=fname))
        Infl_list.append(infl)

    def _assemble(series_list, time_series, col_names):
        df = pd.concat([time_series] + series_list, axis=1)
        df.columns = col_names
        return df

    Raw_recon_DF = _assemble(Raw_recon_list, log_time, cols)
    D1_DF        = _assemble(D1_list,        log_time, cols)
    D2_DF        = _assemble(D2_list,        log_time, cols)
    D3_DF        = _assemble(D3_list,        log_time, cols)
    Infl_DF      = pd.concat(Infl_list, axis=1)

    Resid_list = []
    for i in range(1, n_files + 1):
        interp = np.interp(
            np.array(dn.iloc[:, 0], dtype=float),
            np.array(log_time, dtype=float),
            np.array(Raw_recon_DF.iloc[:, i], dtype=float))
        Resid_list.append(pd.Series(
            np.array(dn.iloc[:, i], dtype=float) - interp, name=cols[i]))
    Resid_DF = pd.concat([dn.iloc[:, 0].reset_index(drop=True)] + Resid_list, axis=1)
    Resid_DF.columns = cols

    return Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time, densify_info


def _fit_splines_pchip(double_norm_df: pd.DataFrame, x_col: str,
                       trim_first: int = 0, trim_last: int = 0) -> tuple:
    """
    Monotone PCHIP interpolation in log10(time) space.

    PchipInterpolator passes through every data point and preserves
    monotonicity between points.  Its analytic derivatives are used
    directly for detection (inflection points, FJ/FI/FP).  For display,
    D2 and D3 are additionally smoothed with a light Gaussian filter
    (sigma=3) because PCHIP is only C¹ and its higher derivatives are
    inherently discontinuous.  Residuals are trivially zero by construction.

    trim_first / trim_last: exclude the first / last N data points from the
    interpolation (useful when the tail of OJIPImaging curves is unreliable).

    Returns an 8-tuple: the same 7 as _fit_splines, plus D2_smooth_DF.
    The 8th element (D2_smooth_DF) is also a DataFrame; callers that
    unpack only 7 values (the _fit_curves dispatcher) will need updating.
    """
    dn = double_norm_df
    cols = dn.columns
    n_files = len(cols) - 1
    _t_start = _fit_start_index(dn, trim_first)
    _t_end   = len(dn) - trim_last if trim_last > 0 else len(dn)
    n_fit = _t_end - _t_start

    # Dense evaluation grid matching the logspline approach (≥ 600 points)
    n_eval = max(600, n_fit)
    log_time = pd.Series(
        np.geomspace(dn.iloc[_t_start:_t_start + 1, 0].values.astype(float)[0],
                     dn.iloc[_t_end - 1:_t_end, 0].values.astype(float)[0],
                     num=n_eval),
        name=cols[0])

    Raw_recon_list, D1_list, D2_list, D3_list, Infl_list = [], [], [], [], []
    D2_smooth_list, D3_smooth_list = [], []

    for i in range(1, n_files + 1):
        fname = cols[i]
        x_raw = dn.iloc[_t_start:_t_end, 0].values.astype(float)
        y     = dn.iloc[_t_start:_t_end, i].values.astype(float)
        x_log = np.log10(x_raw)

        interp_fn = PchipInterpolator(x_log, y)
        x_eval = np.log10(log_time.values.astype(float))

        recon = interp_fn(x_eval)
        d1    = interp_fn.derivative(1)(x_eval)
        d2    = interp_fn.derivative(2)(x_eval)
        d3    = interp_fn.derivative(3)(x_eval)

        # Gaussian-smoothed D2/D3 for display (detection uses raw d2/d3)
        d2_smooth = gaussian_filter1d(d2, sigma=3)
        d3_smooth = gaussian_filter1d(d3, sigma=3)

        # Upward zero-crossings only: D2 goes from - to + (D3 > 0 at crossing)
        # These are the inflection points per Akinyemi et al. 2025
        zc   = np.where(np.diff(np.sign(d2)) > 0)[0]
        infl = pd.Series(log_time.values[zc]).reset_index(drop=True).rename(fname)

        Raw_recon_list.append(pd.Series(recon, name=fname))
        D1_list.append(pd.Series(d1,    name=fname))
        D2_list.append(pd.Series(d2,    name=fname))
        D3_list.append(pd.Series(d3,    name=fname))
        D2_smooth_list.append(pd.Series(d2_smooth, name=fname))
        D3_smooth_list.append(pd.Series(d3_smooth, name=fname))
        Infl_list.append(infl)

    def _assemble(series_list, time_series, col_names):
        df = pd.concat([time_series] + series_list, axis=1)
        df.columns = col_names
        return df

    Raw_recon_DF = _assemble(Raw_recon_list, log_time, cols)
    D1_DF        = _assemble(D1_list,        log_time, cols)
    D2_DF        = _assemble(D2_list,        log_time, cols)
    D3_DF        = _assemble(D3_list,        log_time, cols)
    D2_smooth_DF = _assemble(D2_smooth_list, log_time, cols)
    D3_smooth_DF = _assemble(D3_smooth_list, log_time, cols)
    Infl_DF      = pd.concat(Infl_list, axis=1)

    # PCHIP passes through all points — residuals at raw time axis are ~0 by construction
    Resid_list = []
    for i in range(1, n_files + 1):
        interp_vals = np.interp(
            np.array(dn.iloc[:, 0], dtype=float),
            np.array(log_time, dtype=float),
            np.array(Raw_recon_DF.iloc[:, i], dtype=float))
        Resid_list.append(pd.Series(
            np.array(dn.iloc[:, i], dtype=float) - interp_vals, name=cols[i]))
    Resid_DF = pd.concat([dn.iloc[:, 0].reset_index(drop=True)] + Resid_list, axis=1)
    Resid_DF.columns = cols

    return Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time, D2_smooth_DF, D3_smooth_DF


def _fit_curves(double_norm_df: pd.DataFrame, x_col: str, kr: int,
                method: str = 'spline',
                trim_first: int = 0, trim_last: int = 0,
                knot_placement: str = 'quantile',
                oj_densify: bool = False,
                oj_model: str = 'exponential',
                oj_model_params: 'dict | None' = None) -> tuple:
    """Dispatcher: route to the requested fitting method.

    Always returns a 10-tuple:
      Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time,
      D2_smooth_DF, D3_smooth_DF, densify_info
    D2_smooth_DF / D3_smooth_DF are None for non-PCHIP methods.
    densify_info is {} unless logspline + oj_densify produced results.
    """
    if method == 'logspline':
        result = _fit_splines_log(double_norm_df, x_col,
                                  trim_first=trim_first, trim_last=trim_last,
                                  knot_placement=knot_placement,
                                  oj_densify=oj_densify,
                                  oj_model=oj_model,
                                  oj_model_params=oj_model_params)
        # result is 8-tuple: 7 + densify_info
        return result[:-1] + (None, None, result[-1])
    if method == 'pchip':
        return _fit_splines_pchip(double_norm_df, x_col,
                                  trim_first=trim_first, trim_last=trim_last) + ({},)
    result = _fit_splines(double_norm_df, x_col, kr,
                          trim_first=trim_first, trim_last=trim_last)
    return result + (None, None, {})


def _safe(v):
    """Convert to float, returning None for NaN/None."""
    try:
        f = float(v)
        return None if np.isnan(f) else round(f, 8)
    except Exception:
        return None


def _fscalar(v):
    """Plain float from a single Series element.

    Deliberately untyped: indexing a pandas Series is typed by the stubs as a
    ``Series | scalar`` union, so a bare ``float(series[key])`` trips the type
    checker even though the value is always scalar at runtime. Routing through
    an untyped param launders that union (same trick as ``_safe``)."""
    return float(v)


def _fit_quality(y_raw: np.ndarray, y_recon_at_raw: np.ndarray,
                 d2: np.ndarray, method: str = 'spline') -> dict:
    """
    Per-curve fit quality metric.

    For spline / logspline: the 'poor' flag is driven by R² — the fraction of the
    transient's variance the reconstruction explains. R² is scale-, grid- and
    noise-level-independent and matches visual "does the fit follow the data"
    judgement: a smooth spline over noisy data still scores R² ≈ 0.99+ because the
    noise variance is tiny next to the O-J-I-P dynamic range, while a spline that
    misses real structure drops R² visibly. Flag 'poor' when R² < ``FIT_R2_POOR``.

    History of this flag: (1) fixed nRMSE > 3 % — mass-flagged noisy-but-good
    curves (smooth spline doesn't chase noise, so residual ≈ noise ≈ >3 % on
    per-pixel OJIPImaging). (2) noise-adaptive misfit_ratio = rmse / sigma_noise
    with sigma_noise from 2nd differences — but that high-pass estimator assumes
    WHITE noise; OJIPImaging residuals are smooth/correlated, so sigma_noise
    collapsed and misfit_ratio blew up to ~10 on every curve (nRMSE stayed ~2 %),
    mass-flagging again. nRMSE, R², sigma_noise and misfit_ratio are all still
    returned as diagnostics, but only R² drives the flag now.

    For pchip: residuals are trivially 0 by construction, so instead
    the *derivative roughness* (normalised RMS of diff(d2)) is computed.
    High roughness means noisy data propagates into timing detection.
    A 'poor' flag is raised when roughness > 0.15.
    """
    if method == 'pchip':
        dd2    = np.diff(d2)
        denom  = float(np.max(np.abs(d2))) or 1.0
        roughness = float(np.sqrt(np.mean(dd2 ** 2))) / denom
        flag   = 'poor' if roughness > 0.15 else 'good'
        return {
            'fit_nrmse':     None,
            'fit_r2':        None,
            'fit_roughness': round(roughness, 6),
            'fit_flag':      flag,
            'fit_method':    method,
        }
    res    = y_raw - y_recon_at_raw
    ss_res = float(np.sum(res ** 2))
    ss_tot = float(np.sum((y_raw - float(y_raw.mean())) ** 2))
    rmse   = float(np.sqrt(np.mean(res ** 2)))
    span   = float(y_raw.max() - y_raw.min()) or 1.0
    nrmse  = rmse / span
    r2     = 1.0 - ss_res / ss_tot if ss_tot > 0 else 1.0

    # Robust per-curve noise floor from 2nd differences (cancels the smooth
    # signal's slope; the median resists the few high-curvature OJIP steps):
    #   for white noise σ, std(2nd diff) = √6·σ and MAD ≈ 0.6745·std,
    #   so σ ≈ median|2nd diff| · 1.4826 / √6.
    if y_raw.size >= 3:
        sigma_noise = float(np.median(np.abs(np.diff(y_raw, n=2)))) * 1.4826 / np.sqrt(6.0)
    else:
        sigma_noise = 0.0
    # Ratio of residual to the noise floor: ≈1 when the fit is as good as the
    # noise allows, ≫1 when the spline systematically misses real structure.
    misfit_ratio = rmse / sigma_noise if sigma_noise > 1e-9 else float('inf')

    # Flag on R² alone: it explains-variance, so it is robust to the noise level
    # and grid non-uniformity that broke the nRMSE and misfit_ratio approaches.
    # Threshold is generous (a good fit is ≈0.99+); it fires only on fits that
    # visibly miss the transient. misfit_ratio/nRMSE are kept for diagnostics.
    # Kept LOCAL (not a module global) so the flag never depends on module
    # load/reload order — raise toward 0.98 to make the flag more sensitive.
    FIT_R2_POOR = 0.95
    flag = 'poor' if (r2 is not None and r2 < FIT_R2_POOR) else 'good'
    return {
        'fit_nrmse':        round(nrmse, 6),
        'fit_r2':           round(r2, 6),
        'fit_noise':        round(sigma_noise / span, 6),
        'fit_misfit_ratio': (round(misfit_ratio, 3) if np.isfinite(misfit_ratio) else None),
        'fit_roughness':    None,
        'fit_flag':         flag,
        'fit_method':       method,
    }


def _t_safe(v, ms_factor):
    """Convert native-unit time value to ms, None on NaN."""
    try:
        f = float(v)
        return None if np.isnan(f) else f * ms_factor
    except Exception:
        return None


def analyze_one_curve(time_native, values, fname, fluorometer, fj_time_ms, fi_time_ms, kr,
                      include_curves=False, fit_method='logspline',
                      trim_first: int = 0, trim_last: int = 0,
                      background_mode='auto', background_n=1,
                      f0_source='instrument',
                      bckg: 'float | None' = None, fo_footer: 'float | None' = None,
                      knot_placement: str = 'hybrid',
                      oj_densify: bool = False,
                      oj_model: str = 'exponential',
                      oj_model_params: 'dict | None' = None,
                      f0_time_ms: 'float | None' = None,
                      use_deriv_timing: bool = False):
    """
    Full OJIP analysis pipeline for a single curve.

    Reuses the exact same normalisation, spline-fitting, derivative, polynomial
    inflection-point, JIP-parameter and complementary-area logic as the
    multi-file ``ojip_process`` route, but wrapped so it can be called per-curve
    from the batch endpoint.

    Parameters
    ----------
    time_native : array-like
        Time axis in native units (µs for AquaPen, ms for MC-PAM, s for FL6000).
    values : array-like
        Fluorescence intensity values (same length as *time_native*).
    fname : str
        Curve name / label (used as the DataFrame column name).
    fluorometer : str
        Fluorometer type string (one of the values accepted by ``_axis_cfg``).
    fj_time_ms, fi_time_ms : float
        User-specified FJ / FI step times in milliseconds.
    kr : int
        Knot-reduction factor for the LSQ univariate spline.
    include_curves : bool
        If *True*, append full transient arrays (4 normalisation modes,
        reconstructed spline, 1st/2nd derivatives, residuals, polynomial
        fits) to the returned dict.  Set to *False* for the fast params-only
        batch pass.

    Returns
    -------
    dict
        Always contains scalar JIP parameters and key timing values.
        When *include_curves* is True, also contains ``time_raw_ms``,
        ``time_log_ms`` and a ``curves`` sub-dict.
    """
    ms    = _ms_factor(fluorometer)
    x_col = _axis_cfg(fluorometer)[0]
    ranges = _axis_cfg(fluorometer)[5]

    # Build 2-column DataFrame — identical structure to Summary_file with 1 col
    sf = pd.DataFrame({
        x_col: np.asarray(time_native, dtype=float),
        fname: np.asarray(values, dtype=float),
    })
    data_cols = [fname]
    FJ_time = fj_time_ms / ms
    FI_time = fi_time_ms / ms

    # ── background / first-point correction (AquaPen/FluorPen only) ───────────
    bg_applied = False
    if fluorometer == 'Aquapen':
        bckg_map = {fname: bckg} if bckg is not None else None
        sf, bg_applied = _bg_correct(sf, x_col, data_cols,
                                     background_mode, background_n, bckg_map)

    # ── normalise (mirrors ojip_process lines 419-440) ────────────────────────
    # Always auto-detect F0 first (value at the instrument's reference point).
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F0_index = sf[x_col].sub(0.01).abs().idxmin()
    else:
        F0_index = sf[x_col].sub(0).abs().idxmin()
    F0 = sf[data_cols].loc[F0_index].copy()

    # When the transient is background-subtracted, F0 may be set from the
    # instrument's own reported Fo (footer) so the result matches the FluorPen
    # readout; otherwise F0 stays the first measured point after the background.
    # Skip when user explicitly overrides F0 timing.
    if (f0_time_ms is None and f0_source == 'instrument' and bg_applied
            and fo_footer is not None and np.isfinite(fo_footer)):
        F0[fname] = float(fo_footer)

    # F0 time override: move ONLY the F0 data point to the user-specified time.
    # All other time points stay at their original positions.  This changes
    # the gap between F0 and the next point but preserves the rest of the
    # time axis, so spline fitting and derivatives reflect the corrected
    # F0 position without distorting the remainder of the curve.
    if f0_time_ms is not None:
        f0_native_target = f0_time_ms / ms
        sf.loc[F0_index, x_col] = f0_native_target
    FM = sf[data_cols].max()

    shifted_zero_df = pd.concat([
        sf.iloc[:, 0],
        sf[data_cols].subtract(F0, axis=1),
    ], axis=1)
    shifted_max_df = pd.concat([
        sf.iloc[:, 0],
        sf[data_cols].add(abs(FM - FM.max()), axis=1),
    ], axis=1)
    FMFORNORM = shifted_zero_df[data_cols].max()
    dn_df = pd.concat([
        sf.iloc[:, 0],
        shifted_zero_df[data_cols].div(FMFORNORM, axis=1),
    ], axis=1)

    # ── spline fitting ────────────────────────────────────────────────────────
    _spline_method = 'logspline' if fit_method in _DETECTION_METHODS else fit_method
    Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time, \
        D2_smooth_DF, D3_smooth_DF, densify_info = \
        _fit_curves(dn_df, x_col, kr, method=_spline_method,
                    trim_first=trim_first, trim_last=trim_last,
                    knot_placement=knot_placement,
                    oj_densify=oj_densify,
                    oj_model=oj_model,
                    oj_model_params=oj_model_params)

    poly_oj = _fit_oj_polynomial(dn_df, x_col, ms)
    poly_oi = _fit_oj_polynomial(dn_df, x_col, ms,
                                  oj_lo_ms=9.0, oj_hi_ms=100.0)

    # ── method-specific fits ─────────────────────────────────────────────────
    three_exp_res = _fit_three_exponential(dn_df, x_col, ms) if fit_method == 'three_exp' else None
    piecewise_res = _fit_piecewise_linear(dn_df, x_col, ms)  if fit_method == 'piecewise' else None
    gauss_d1_res  = _fit_d1_gaussians(D1_DF, x_col, ms)      if fit_method == 'gaussian_d1' else None

    # ── find FJ / FI / FP ────────────────────────────────────────────────────
    _use_poly = (fit_method == 'polynomial')
    _fjfi_method = fit_method if fit_method in _DETECTION_METHODS else 'd2_trough'
    FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl, fjifp_conf, method_extras = \
        _find_fjfifp(D2_DF, D3_DF, x_col, ranges, data_cols, Infl_DF,
                     fj_expect=FJ_time, fi_expect=FI_time,
                     poly_oj=poly_oj if _use_poly else None,
                     poly_oi=poly_oi if _use_poly else None,
                     ms_factor=ms,
                     method=_fjfi_method,
                     D1_DF=D1_DF,
                     three_exp_results=three_exp_res,
                     piecewise_results=piecewise_res,
                     gaussian_d1_results=gauss_d1_res)

    # ── reference time indexes ────────────────────────────────────────────────
    def tidx(t):
        return sf[x_col].sub(t).abs().idxmin()

    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F50us_idx  = tidx(0.05);   FK_idx     = tidx(0.3)
        F50ms_idx  = tidx(50);     F100ms_idx = tidx(100)
        F200ms_idx = tidx(200);    F300ms_idx = tidx(300)
    elif fluorometer == 'Aquapen':
        F50us_idx  = tidx(50);     FK_idx     = tidx(300)
        F50ms_idx  = tidx(50000);  F100ms_idx = tidx(100000)
        F200ms_idx = tidx(200000); F300ms_idx = tidx(300000)
    elif fluorometer == 'FL6000':
        F50us_idx  = tidx(5e-5);   FK_idx     = tidx(3e-4)
        F50ms_idx  = tidx(0.05);   F100ms_idx = tidx(0.1)
        F200ms_idx = tidx(0.2);    F300ms_idx = tidx(0.3)
    elif fluorometer == 'OJIPImaging':
        F50us_idx  = tidx(0.05);   FK_idx     = tidx(0.3)
        F50ms_idx  = tidx(50);     F100ms_idx = tidx(100)
        F200ms_idx = tidx(200);    F300ms_idx = tidx(300)
    else:
        raise ValueError(f'Unknown fluorometer: {fluorometer}')

    # Guard: if F50us and FK resolve to the same index (time axis lacks
    # sub-ms resolution, common for imaging-PAM), widen the interval so M₀
    # reflects the actual initial slope over the nearest available time points.
    if F50us_idx == FK_idx:
        FK_idx = min(int(F50us_idx) + 2, len(sf) - 1)
        if FK_idx == F50us_idx:          # only 1-2 data points total
            FK_idx = len(sf) - 1

    # ── optionally override FJ/FI with derivative-detected times ─────────────
    _deriv_timing_used = False
    if use_deriv_timing:
        _fj_d = _t_safe(FJ_deriv.get(fname), ms)
        _fi_d = _t_safe(FI_deriv.get(fname), ms)
        if _fj_d is not None and _fi_d is not None and _fj_d < _fi_d:
            fj_time_ms = _fj_d
            fi_time_ms = _fi_d
            FJ_time = fj_time_ms / ms
            FI_time = fi_time_ms / ms
            _deriv_timing_used = True

    FJ_idx = tidx(FJ_time)
    FI_idx = tidx(FI_time)

    F50 = sf[data_cols].loc[F50us_idx]
    FK  = sf[data_cols].loc[FK_idx]
    FJ  = sf[data_cols].loc[FJ_idx]
    FI  = sf[data_cols].loc[FI_idx]

    # ── JIP parameters ────────────────────────────────────────────────────────
    FV      = FM - F0
    FVFM    = FV / FM
    M0      = 4 * (FK - F50) / FV
    VJ      = (FJ - F0) / FV
    VI      = (FI - F0) / FV
    OJ      = FJ - F0
    JI      = FI - FJ
    IP      = FM - FI
    PSIE0   = 1 - VJ
    PSIR0   = 1 - VI
    DELTAR0 = PSIR0 / PSIE0
    PHIE0   = FVFM * PSIE0
    PHIR0   = FVFM * PSIR0
    TR0RC   = M0 / VJ
    ABSRC   = TR0RC / FVFM
    ET0RC   = TR0RC * PSIE0
    RE0RC   = TR0RC * PSIR0
    DI0RC   = ABSRC - TR0RC

    # ── phase slopes (fluorescence rise rate between O-J, J-I, I-P) ──────────
    # Use derived (detected) timings for phase boundaries; fall back to user times.
    _fj_t = _t_safe(FJ_deriv.get(fname), ms) or fj_time_ms
    _fi_t = _t_safe(FI_deriv.get(fname), ms) or fi_time_ms
    _fp_t = _t_safe(FP_deriv.get(fname), ms)
    _f0_t = float(sf[x_col].iloc[int(F50us_idx)]) * ms  # F0 time in ms
    _f0_v = _fscalar(F0[fname])
    _fj_v = _fscalar(FJ[fname])
    _fi_v = _fscalar(FI[fname])
    _fm_v = _fscalar(FM[fname])

    def _slope(f_end, f_start, t_end, t_start):
        dt = t_end - t_start
        if dt <= 0 or not np.isfinite(dt):
            return np.nan
        return (f_end - f_start) / dt

    slope_OJ = pd.Series([_slope(_fj_v, _f0_v, _fj_t, _f0_t)], index=[fname], name=fname)
    slope_JI = pd.Series([_slope(_fi_v, _fj_v, _fi_t, _fj_t)], index=[fname], name=fname)
    slope_IP = pd.Series([_slope(_fm_v, _fi_v, _fp_t if _fp_t else _fi_t + 100,
                                  _fi_t)], index=[fname], name=fname)

    # ── FI-FP dip detection ────────────────────────────────────────────────────
    # Look for a local D1 minimum (fluorescence decrease) between FI and FP.
    _lt_ms = log_time.values.astype(float) * ms  # log-time grid in ms
    _d1_vals = np.array(D1_DF[fname].values, dtype=float)
    _recon_vals = np.array(Raw_recon_DF[fname].values, dtype=float)
    _fi_t_actual = _fi_t if _fi_t else fi_time_ms
    _fp_t_actual = _fp_t if _fp_t else float(_lt_ms[-1])

    dip_IP_present = False
    dip_IP_time_ms_val = None
    dip_IP_amplitude_val = None
    dip_IP_d1_min_val = None

    _ip_mask = (_lt_ms >= _fi_t_actual) & (_lt_ms <= _fp_t_actual)
    if np.any(_ip_mask):
        _d1_ip = _d1_vals[_ip_mask]
        _t_ip  = _lt_ms[_ip_mask]
        _r_ip  = _recon_vals[_ip_mask]
        _d1_min_idx = np.argmin(_d1_ip)
        _d1_min_val = float(_d1_ip[_d1_min_idx])

        if _d1_min_val < 0:
            dip_IP_present = True
            dip_IP_time_ms_val = float(_t_ip[_d1_min_idx])
            dip_IP_d1_min_val = _d1_min_val
            # Amplitude: depth below the straight line FI → FP at the dip time
            _fi_recon = float(np.interp(_fi_t_actual, _lt_ms, _recon_vals))
            _fp_recon = float(np.interp(_fp_t_actual, _lt_ms, _recon_vals))
            _dt_total = _fp_t_actual - _fi_t_actual
            if _dt_total > 0:
                _frac = (dip_IP_time_ms_val - _fi_t_actual) / _dt_total
                _line_val = _fi_recon + _frac * (_fp_recon - _fi_recon)
                _curve_val = float(_r_ip[_d1_min_idx])
                dip_IP_amplitude_val = _line_val - _curve_val  # positive = dip below line

    # ── areas + FM timing ─────────────────────────────────────────────────────
    AREAOJ, AREAJI, AREAIP, AREAOP, FM_timings = _calc_areas_fm_timing(
        sf, data_cols, FJ_idx, FI_idx, ms,
        F50ms_idx, F100ms_idx, F200ms_idx, F300ms_idx)
    SM    = AREAOP / FV
    N_val = SM * M0 * (1 / VJ)

    # ── fit quality (evaluated on fitted / non-trimmed region only) ───────────
    _y_raw_dn      = np.array(dn_df[fname].values, dtype=float)
    _y_recon_at_raw = _y_raw_dn - np.array(Resid_DF[fname].values, dtype=float)
    _d2_vals       = np.array(D2_DF[fname].values, dtype=float)
    _q_start = _fit_start_index(dn_df, trim_first)
    _q_end   = len(_y_raw_dn) - trim_last if trim_last > 0 else len(_y_raw_dn)
    fq = _fit_quality(_y_raw_dn[_q_start:_q_end],
                      _y_recon_at_raw[_q_start:_q_end],
                      _d2_vals, method=fit_method)

    # ── build result dict ─────────────────────────────────────────────────────
    result = {
        'F0':  _safe(F0[fname]),  'FM': _safe(FM[fname]),
        'FK':  _safe(FK[fname]),  'F50': _safe(F50[fname]),
        'FJ':  _safe(FJ[fname]),  'FI':  _safe(FI[fname]),
        'FJ_time_user_ms':    fj_time_ms,
        'FI_time_user_ms':    fi_time_ms,
        'FJ_time_deriv_ms':   _t_safe(FJ_deriv.get(fname), ms),
        'FI_time_deriv_ms':   _t_safe(FI_deriv.get(fname), ms),
        'FP_time_deriv_ms':   _t_safe(FP_deriv.get(fname), ms),
        'FJ_time_inflect_ms': _t_safe(FJ_infl.get(fname),  ms),
        'FI_time_inflect_ms': _t_safe(FI_infl.get(fname),  ms),
        'FP_time_inflect_ms': _t_safe(FP_infl.get(fname),  ms),
        'FJ_conf': _safe(fjifp_conf['FJ'].get(fname)),
        'FI_conf': _safe(fjifp_conf['FI'].get(fname)),
        'FP_conf': _safe(fjifp_conf['FP'].get(fname)),
        'FM_time_ms':  _safe(FM_timings.get(fname)),
        'Area_OJ': _safe(AREAOJ[fname]),  'Area_JI': _safe(AREAJI[fname]),
        'Area_IP': _safe(AREAIP[fname]),  'Area_OP': _safe(AREAOP[fname]),
        'poly_infl_ms':    poly_oj[fname]['poly_infl_ms'],
        'poly_fi_infl_ms': poly_oi[fname]['poly_infl_ms'],
        # JIP computed parameters
        'FVFM': _safe(FVFM[fname]),    'VJ': _safe(VJ[fname]),
        'VI': _safe(VI[fname]),        'M0': _safe(M0[fname]),
        'PSIE0': _safe(PSIE0[fname]),  'PSIR0': _safe(PSIR0[fname]),
        'DELTAR0': _safe(DELTAR0[fname]),
        'PHIE0': _safe(PHIE0[fname]),  'PHIR0': _safe(PHIR0[fname]),
        'ABSRC': _safe(ABSRC[fname]),  'TR0RC': _safe(TR0RC[fname]),
        'ET0RC': _safe(ET0RC[fname]),  'RE0RC': _safe(RE0RC[fname]),
        'DI0RC': _safe(DI0RC[fname]),
        'OJ': _safe(OJ[fname]),  'JI': _safe(JI[fname]),  'IP': _safe(IP[fname]),
        'SM': _safe(SM[fname]),  'N':  _safe(N_val[fname]),
        # Phase slopes (fluorescence rise rate, r.u. / ms)
        'slope_OJ': _safe(slope_OJ[fname]),
        'slope_JI': _safe(slope_JI[fname]),
        'slope_IP': _safe(slope_IP[fname]),
        # FI-FP dip detection
        'dip_IP_present':   dip_IP_present,
        'dip_IP_time_ms':   _safe(dip_IP_time_ms_val) if dip_IP_time_ms_val is not None else None,
        'dip_IP_amplitude': _safe(dip_IP_amplitude_val) if dip_IP_amplitude_val is not None else None,
        'dip_IP_d1_min':    _safe(dip_IP_d1_min_val) if dip_IP_d1_min_val is not None else None,
        'deriv_timing_used': _deriv_timing_used,
        **fq,
    }

    if include_curves:
        # Drop the non-positive pre-illumination baseline from the raw-axis
        # display arrays (see ojip_process for the full rationale): a t ≤ 0 point
        # makes the log-axis charts fall back to a ~10 ms floor and hide the rise.
        _t_disp = sf[x_col].values.astype(float)
        _disp0  = int(np.argmax(_t_disp > 0)) if (_t_disp > 0).any() else 0
        result['time_raw_ms'] = (sf[x_col].iloc[_disp0:].astype(float) * ms).tolist()
        result['time_log_ms'] = (log_time.astype(float) * ms).tolist()
        result['curves'] = {
            'raw':           [_safe(v) for v in sf[fname].iloc[_disp0:]],
            'shifted_F0':    [_safe(v) for v in shifted_zero_df[fname].iloc[_disp0:]],
            'shifted_FM':    [_safe(v) for v in shifted_max_df[fname].iloc[_disp0:]],
            'double_norm':   [_safe(v) for v in dn_df[fname].iloc[_disp0:]],
            'residuals':     [_safe(v) for v in Resid_DF[fname].iloc[_disp0:]],
            'reconstructed': [_safe(v) for v in Raw_recon_DF[fname]],
            'd1':            [_safe(v) for v in D1_DF[fname]],
            'd2':            [_safe(v) for v in D2_DF[fname]],
            'd3':            [_safe(v) for v in D3_DF[fname]],
            'd2_smooth':     [_safe(v) for v in D2_smooth_DF[fname]] if D2_smooth_DF is not None else None,
            'd3_smooth':     [_safe(v) for v in D3_smooth_DF[fname]] if D3_smooth_DF is not None else None,
            'poly_oj_time_ms': poly_oj[fname]['poly_oj_time_ms'],
            'poly_oj_d2':      poly_oj[fname]['poly_oj_d2'],
            'poly_oi_time_ms': poly_oi[fname]['poly_oj_time_ms'],
            'poly_oi_d2':      poly_oi[fname]['poly_oj_d2'],
        }
        if fname in method_extras:
            result['curves']['method_fit'] = method_extras[fname]

    if densify_info:
        result['densify_info'] = densify_info

    # Method-specific key_values (analyze_one_curve result is flat, not nested by fname)
    _me = method_extras.get(fname)
    if _me and fit_method == 'three_exp':
        for k in ('A_OJ', 'A_JI', 'A_IP', 'tau_OJ_ms', 'tau_JI_ms', 'tau_IP_ms'):
            result[k] = _me.get(k)
    elif _me and fit_method == 'gaussian_d1':
        for g in range(_me.get('n_gaussians', 0)):
            result[f'gauss_center_{g+1}_ms'] = _me['centers_ms'][g] if g < len(_me.get('centers_ms', [])) else None
            result[f'gauss_sigma_{g+1}']     = _me['sigmas'][g]     if g < len(_me.get('sigmas', []))     else None
            result[f'gauss_amp_{g+1}']       = _me['amplitudes'][g] if g < len(_me.get('amplitudes', [])) else None

    return result


# ─── routes ─────────────────────────────────────────────────────────────────

@OJIP_data_analysis.route('/OJIP', methods=['GET'])
def ojip_page():
    return render_template('OJIP_analysis.html')

@OJIP_data_analysis.route('/OJIP_data_analysis', methods=['GET'])
def ojip_page_redirect():
    return redirect(url_for('OJIP_data_analysis.ojip_page'), 301)


@OJIP_data_analysis.route('/api/ojip_process', methods=['POST'])
def ojip_process():
    upload_folder = UPLOAD_FOLDER
    if not os.path.isdir(upload_folder):
        os.mkdir(upload_folder)

    if 'OJIP_files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files received.'}), 400

    files = request.files.getlist('OJIP_files')
    if not files or secure_filename(files[0].filename or '') == '':
        return jsonify({'status': 'error', 'message': 'Please select one or more files.'}), 400

    fluorometer = request.form.get('fluorometer', '')
    kr = int(request.form.get('knots_reduction_factor', 10))
    fit_method_proc = request.form.get('fit_method', 'logspline')
    trim_first_proc = int(request.form.get('trim_first', 0))
    trim_last_proc  = int(request.form.get('trim_last',  0))
    background_mode = request.form.get('background_mode', 'auto')
    background_n    = int(request.form.get('background_n', 1) or 1)
    f0_source       = request.form.get('f0_source', 'instrument')
    knot_placement  = request.form.get('knot_placement', 'quantile')
    _f0_raw_proc    = request.form.get('f0_time_ms', None)
    f0_time_ms_proc = float(_f0_raw_proc) if _f0_raw_proc not in (None, '') else None
    reduce_size = request.form.get('checkbox_reduce_file_size') == 'checked'
    FJ_time_ms = float(request.form.get('FJ_time', 2.0))
    FI_time_ms = float(request.form.get('FI_time', 30.0))

    if FJ_time_ms >= FI_time_ms:
        return jsonify({'status': 'error', 'message': 'FJ time must be less than FI time.'}), 400
    if len(files) > 100:
        return jsonify({'status': 'error', 'message': 'Maximum 100 files allowed.'}), 400

    try:
        x_col, x_unit, y_unit, _, allowed_ext, ranges = _axis_cfg(fluorometer)
    except ValueError:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 400

    ms = _ms_factor(fluorometer)
    FJ_time = FJ_time_ms / ms   # native units
    FI_time = FI_time_ms / ms

    # ── parse files ──────────────────────────────────────────────────────────
    Summary_file = pd.DataFrame()
    file_names_list = []
    last_fname = 'ojip'

    for file_number, file in enumerate(files):
        fname_no_ext = str.lower(os.path.splitext(file.filename or '')[0])
        ext = str.lower(os.path.splitext(file.filename or '')[1])
        fname_full = secure_filename(file.filename or '')
        last_fname = fname_no_ext

        if ext not in allowed_ext:
            return jsonify({'status': 'error',
                            'message': f'Wrong file type for {fname_full}. Expected: {allowed_ext}'}), 400

        if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
            df = pd.read_csv(file.stream, sep=';', engine='python')
            if str(df.columns[0]) != 'time/ms':
                return jsonify({'status': 'error',
                                'message': f'{fname_full}: first column must be "time/ms".'}), 400
            if file_number == 0:
                Summary_file = df.iloc[:, 0:2].rename(columns={df.columns[1]: fname_no_ext})
            else:
                tmp = df.iloc[:, 0:2].rename(columns={df.columns[1]: fname_no_ext})
                Summary_file = pd.merge(Summary_file, tmp, on='time/ms', how='outer')

        elif fluorometer in ('Aquapen', 'FL6000'):
            raw = file.read().decode('utf-8', errors='replace').splitlines(keepends=True)
            df = pd.DataFrame(raw)
            df = df[0].str.split('\t', expand=True).iloc[:, :2]

            if fluorometer == 'Aquapen':
                if not df[0].astype(str).str.strip().str.contains('FluorPen|AquaPen', case=False).any():
                    return jsonify({'status': 'error',
                                    'message': f'{fname_full}: not a valid AquaPen/FluorPen file.'}), 400
                if df.shape[1] < 2:
                    return jsonify({'status': 'error',
                                    'message': f'{fname_full}: file contains no measurement data (empty recording).'}), 400
                # Strip header rows and convert types before merging to avoid Cartesian product
                good = df[0].astype(str).str.isnumeric()
                df_clean = df[good].rename(columns={df.columns[0]: 'time_us', df.columns[1]: fname_no_ext}).copy()
                df_clean['time_us'] = df_clean['time_us'].astype(int)
                df_clean[fname_no_ext] = pd.to_numeric(df_clean[fname_no_ext], errors='coerce')
                # NB: the first ~11 µs point is a background reading; dropping /
                # subtracting it is now handled uniformly by _bg_correct() in the
                # normalize block below (was a hard-coded iloc[1:] here, which
                # never reached the batch/wide FluorPen path).
                if file_number == 0:
                    Summary_file = df_clean
                else:
                    Summary_file = pd.merge(Summary_file, df_clean[['time_us', fname_no_ext]], on='time_us', how='outer')
            else:  # FL6000
                if not df[0].astype(str).str.strip().str.contains('Fluorometer', case=False).any():
                    return jsonify({'status': 'error',
                                    'message': f'{fname_full}: not a valid FL6000 file.'}), 400
                start = df[df[0].str.strip() == 'Time'].index[0] + 1
                df = df.iloc[start:].reset_index(drop=True)
                if file_number == 0:
                    Summary_file = df.rename(columns={df.columns[0]: 'time_s', df.columns[1]: fname_no_ext})
                else:
                    tmp = df.rename(columns={df.columns[0]: 'time_s', df.columns[1]: fname_no_ext})
                    Summary_file = pd.merge(Summary_file, tmp[['time_s', fname_no_ext]], on='time_s', how='outer')

        file_names_list.append(fname_no_ext)

    # ── clean data ───────────────────────────────────────────────────────────
    # (Aquapen: header stripping and type conversion done per-file inside the loop above)
    if fluorometer == 'FL6000':
        Summary_file = Summary_file.rename(columns={Summary_file.columns[0]: 'time_s'})
        for col in Summary_file.columns:
            Summary_file[col] = pd.to_numeric(Summary_file[col], errors='coerce')
        Summary_file = Summary_file.dropna(subset=['time_s'])

    # Sort by time and interpolate any NaN values introduced by outer-join merging
    # of files with different time axes.
    # - interpolate(method='index'): fills interior gaps using actual time values as weights
    # - ffill/bfill: fills trailing/leading NaN (beyond a file's measurement range)
    #   with the last/first real value so spline fitting never receives NaN inputs.
    time_col_name = Summary_file.columns[0]
    Summary_file = (Summary_file
                    .sort_values(time_col_name)
                    .set_index(time_col_name)
                    .interpolate(method='index')
                    .ffill()
                    .bfill()
                    .reset_index()
                    .reset_index(drop=True))

    # ── reduce MC-PAM data ───────────────────────────────────────────────────
    # MC-PAM records at 0.01 ms resolution (~73 000 pts/file). For OJIP analysis,
    # 0.1 ms in the fast region and ~1 ms in the slow region is more than sufficient.
    # Target: ≤ 2 000 points per file regardless of original length.
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)' and reduce_size:
        # Clip pre-illumination baseline (keep from t ≈ 0.01 ms onward)
        F0_i = int(Summary_file['time/ms'].sub(0.01).abs().idxmin())
        Summary_file = Summary_file.iloc[F0_i:].reset_index(drop=True)
        # Three-zone downsampling preserving the initial steep rise:
        #   0 – 0.5 ms : full 0.01 ms resolution  (~50 pts)  — initial O→J rise intact
        #   0.5 – 30 ms: every 5th pt → 0.05 ms   (~590 pts) — J phase well sampled
        #   30 ms+      : ~200 pts                            — I→P slow phase
        t05_i   = int(Summary_file['time/ms'].sub(0.5).abs().idxmin())
        FI30_i  = int(Summary_file['time/ms'].sub(30).abs().idxmin())
        n_slow  = max(1, len(Summary_file) - FI30_i)
        s_factor = max(1, n_slow // 200)
        very_fast = Summary_file.iloc[:t05_i]
        mid       = Summary_file.iloc[t05_i:FI30_i:5]
        slow      = Summary_file.iloc[FI30_i::s_factor]
        Summary_file = pd.concat([very_fast, mid, slow]).reset_index(drop=True)

    data_cols = list(Summary_file.columns[1:])
    n_files = len(data_cols)

    # Background / first-point correction (AquaPen/FluorPen only). Single-file
    # uploads carry no Bckg/Fo footer (the loader keeps only numeric rows), so
    # 'auto' here means "drop the first point" — identical to the old iloc[1:].
    if fluorometer == 'Aquapen':
        Summary_file, _ = _bg_correct(Summary_file, x_col, data_cols,
                                      background_mode, background_n, None)

    # ── normalize ────────────────────────────────────────────────────────────
    # Always auto-detect F0 first.
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F0_index = Summary_file[x_col].sub(0.01).abs().idxmin()
    else:
        F0_index = Summary_file[x_col].sub(0).abs().idxmin()
    F0 = Summary_file[data_cols].loc[F0_index]

    # F0 time override: move only the F0 data point to the user-specified time.
    if f0_time_ms_proc is not None:
        f0_native_target = f0_time_ms_proc / ms
        Summary_file.loc[F0_index, x_col] = f0_native_target
    FM = Summary_file[data_cols].max()

    OJIP_shifted_to_zero = pd.concat([
        Summary_file.iloc[:, 0],
        Summary_file[data_cols].subtract(F0, axis=1)
    ], axis=1)
    OJIP_shifted_to_max = pd.concat([
        Summary_file.iloc[:, 0],
        Summary_file[data_cols].add(abs(FM - FM.max()), axis=1)
    ], axis=1)
    FMFORNORM = OJIP_shifted_to_zero[data_cols].max()
    OJIP_double_normalized = pd.concat([
        Summary_file.iloc[:, 0],
        OJIP_shifted_to_zero[data_cols].div(FMFORNORM, axis=1)
    ], axis=1)

    # ── spline fitting ───────────────────────────────────────────────────────
    _spline_method_proc = 'logspline' if fit_method_proc in _DETECTION_METHODS else fit_method_proc
    try:
        Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time, \
            D2_smooth_DF, D3_smooth_DF, _densify = _fit_curves(
            OJIP_double_normalized, x_col, kr, method=_spline_method_proc,
            trim_first=trim_first_proc, trim_last=trim_last_proc,
            knot_placement=knot_placement)
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'Spline fitting failed.'}), 400

    poly_oj = _fit_oj_polynomial(OJIP_double_normalized, x_col, ms)                          # FJ window 0.5–5 ms
    poly_oi = _fit_oj_polynomial(OJIP_double_normalized, x_col, ms, oj_lo_ms=9.0, oj_hi_ms=100.0)   # FI window 9–100 ms

    # ── method-specific fits ─────────────────────────────────────────────────
    three_exp_res_proc = _fit_three_exponential(OJIP_double_normalized, x_col, ms) if fit_method_proc == 'three_exp' else None
    piecewise_res_proc = _fit_piecewise_linear(OJIP_double_normalized, x_col, ms)  if fit_method_proc == 'piecewise' else None
    gauss_d1_res_proc  = _fit_d1_gaussians(D1_DF, x_col, ms)                      if fit_method_proc == 'gaussian_d1' else None

    # ── find FJ/FI/FP ────────────────────────────────────────────────────────
    _use_poly_proc = (fit_method_proc == 'polynomial')
    _fjfi_method_proc = fit_method_proc if fit_method_proc in _DETECTION_METHODS else 'd2_trough'
    try:
        FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl, fjifp_conf, method_extras_proc = _find_fjfifp(
            D2_DF, D3_DF, x_col, ranges, data_cols, Infl_DF,
            fj_expect=FJ_time, fi_expect=FI_time,
            poly_oj=poly_oj if _use_poly_proc else None,
            poly_oi=poly_oi if _use_poly_proc else None,
            ms_factor=ms,
            method=_fjfi_method_proc,
            D1_DF=D1_DF,
            three_exp_results=three_exp_res_proc,
            piecewise_results=piecewise_res_proc,
            gaussian_d1_results=gauss_d1_res_proc)
    except ValueError:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 400

    # ── reference time indexes ───────────────────────────────────────────────
    def tidx(t): return Summary_file[x_col].sub(t).abs().idxmin()

    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F50us_idx  = tidx(0.05);  FK_idx     = tidx(0.3)
        F50ms_idx  = tidx(50);    F100ms_idx = tidx(100)
        F200ms_idx = tidx(200);   F300ms_idx = tidx(300)
    elif fluorometer == 'Aquapen':
        F50us_idx  = tidx(50);    FK_idx     = tidx(300)
        F50ms_idx  = tidx(50000); F100ms_idx = tidx(100000)
        F200ms_idx = tidx(200000);F300ms_idx = tidx(300000)
    elif fluorometer == 'FL6000':
        F50us_idx  = tidx(5e-5);  FK_idx     = tidx(3e-4)
        F50ms_idx  = tidx(0.05);  F100ms_idx = tidx(0.1)
        F200ms_idx = tidx(0.2);   F300ms_idx = tidx(0.3)
    elif fluorometer == 'OJIPImaging':
        F50us_idx  = tidx(0.05);  FK_idx     = tidx(0.3)
        F50ms_idx  = tidx(50);    F100ms_idx = tidx(100)
        F200ms_idx = tidx(200);   F300ms_idx = tidx(300)
    else:
        raise ValueError(f'Unknown fluorometer: {fluorometer}')

    if F50us_idx == FK_idx:
        FK_idx = min(int(F50us_idx) + 2, len(Summary_file) - 1)
        if FK_idx == F50us_idx:
            FK_idx = len(Summary_file) - 1

    FJ_idx = tidx(FJ_time)
    FI_idx = tidx(FI_time)

    F50 = Summary_file[data_cols].loc[F50us_idx]
    FK  = Summary_file[data_cols].loc[FK_idx]
    FJ  = Summary_file[data_cols].loc[FJ_idx]
    FI  = Summary_file[data_cols].loc[FI_idx]

    # ── JIP parameters ───────────────────────────────────────────────────────
    FV      = FM - F0
    FVFM    = FV / FM
    M0      = 4 * (FK - F50) / FV
    VJ      = (FJ - F0) / FV
    VI      = (FI - F0) / FV
    OJ      = FJ - F0
    JI      = FI - FJ
    IP      = FM - FI
    PSIE0   = 1 - VJ
    PSIR0   = 1 - VI
    DELTAR0 = PSIR0 / PSIE0
    PHIE0   = FVFM * PSIE0
    PHIR0   = FVFM * PSIR0
    TR0RC   = M0 / VJ
    ABSRC   = TR0RC / FVFM
    ET0RC   = TR0RC * PSIE0
    RE0RC   = TR0RC * PSIR0
    DI0RC   = ABSRC - TR0RC

    # ── areas + FM timing ────────────────────────────────────────────────────
    AREAOJ, AREAJI, AREAIP, AREAOP, FM_timings_series = _calc_areas_fm_timing(
        Summary_file, data_cols, FJ_idx, FI_idx, ms,
        F50ms_idx, F100ms_idx, F200ms_idx, F300ms_idx)
    SM = AREAOP / FV
    N  = SM * M0 * (1 / VJ)

    # ── build xlsx ───────────────────────────────────────────────────────────
    params_to_concat = [
        F0, FK, FJ, FI, FM, OJ, JI, IP, VJ, VI, M0, PSIE0, PSIR0, DELTAR0, FVFM,
        PHIE0, PHIR0, ABSRC, TR0RC, ET0RC, RE0RC, DI0RC,
        pd.Series(AREAOJ), pd.Series(AREAJI), pd.Series(AREAIP), pd.Series(AREAOP),
        SM, N, FJ_infl, FI_infl, FP_infl, FM_timings_series, FJ_deriv, FI_deriv, FP_deriv
    ]
    OJIP_param_all = pd.concat(params_to_concat, axis=1)
    OJIP_param_all.columns = [
        'Fin', 'FK', 'FJ', 'FI', 'Fmax', 'Amplitude(0-J)', 'Amplitude(J-I)', 'Amplitude(I-P)',
        'VJ', 'VI', 'M0', 'ψE0', 'ψR0', 'δR0', 'ψP0 (Fv/Fm)', 'φE0', 'φR0',
        'ABS/RC', 'TR0/RC', 'ET0/RC', 'RE0/RC', 'DI0/RC',
        'Complementary area O-J', 'Complementary area J-I',
        'Complementary area I-P', 'Complementary area (O-P)',
        'Normalized complementary area Sm', 'N (turn-over number QA)',
        'Time FJ', 'Time FI', 'Time FP', 'Time FM',
        'Time Min 2nd Deriv Pre-FJ', 'Time Min 2nd Deriv Pre-FI', 'Time Min 2nd Deriv Pre-FP'
    ]

    # ── build JSON payload ───────────────────────────────────────────────────
    # Ensure all data-column labels are plain Python strings so JSON serialises
    # them as strings (not numbers). Pandas may carry numpy scalar column names
    # when filenames consist of digits only, causing the JS to receive numeric
    # file identifiers and breaking curve_id hashing in the annotation tool.
    data_cols = [str(c) for c in data_cols]
    new_cols = [Summary_file.columns[0]] + data_cols   # preserve time-axis name
    Summary_file.columns = new_cols

    # Display arrays are plotted on a LOG time axis, which cannot render the
    # non-positive pre-illumination baseline (MC-PAM records t ≤ 0 ms before the
    # light pulse). A single x ≤ 0 point makes Chart.js abandon data-driven
    # auto-scaling and fall back to its default ~10 ms floor, hiding the whole
    # O–J rise. Drop the leading non-positive samples from the raw-axis display
    # arrays so every log chart auto-scales from the first real timepoint.
    # (Reconstructed / derivative arrays live on the already-positive log_time
    # grid and are left untouched. Analysis/JIP params are computed above from
    # the full Summary_file and are unaffected.)
    _t_disp = Summary_file.iloc[:, 0].values.astype(float)
    _disp0  = int(np.argmax(_t_disp > 0)) if (_t_disp > 0).any() else 0

    time_raw_ms = (Summary_file.iloc[_disp0:, 0].astype(float) * ms).tolist()
    time_log_ms = (log_time.astype(float) * ms).tolist()

    curves = {}
    for i, fname in enumerate(data_cols, start=1):
        curves[fname] = {
            'raw':          [_safe(v) for v in Summary_file.iloc[_disp0:, i]],
            'shifted_F0':   [_safe(v) for v in OJIP_shifted_to_zero.iloc[_disp0:, i]],
            'shifted_FM':   [_safe(v) for v in OJIP_shifted_to_max.iloc[_disp0:, i]],
            'double_norm':  [_safe(v) for v in OJIP_double_normalized.iloc[_disp0:, i]],
            'residuals':    [_safe(v) for v in Resid_DF.iloc[_disp0:, i]],
            'reconstructed':[_safe(v) for v in Raw_recon_DF.iloc[:, i]],
            'd1':              [_safe(v) for v in D1_DF.iloc[:, i]],
            'd2':              [_safe(v) for v in D2_DF.iloc[:, i]],
            'd3':              [_safe(v) for v in D3_DF.iloc[:, i]],
            'd2_smooth':       [_safe(v) for v in D2_smooth_DF.iloc[:, i]] if D2_smooth_DF is not None else None,
            'd3_smooth':       [_safe(v) for v in D3_smooth_DF.iloc[:, i]] if D3_smooth_DF is not None else None,
            'poly_oj_time_ms': poly_oj[fname]['poly_oj_time_ms'],
            'poly_oj_d2':      poly_oj[fname]['poly_oj_d2'],
            'poly_oi_time_ms': poly_oi[fname]['poly_oj_time_ms'],
            'poly_oi_d2':      poly_oi[fname]['poly_oj_d2'],
        }
        if fname in method_extras_proc:
            curves[fname]['method_fit'] = method_extras_proc[fname]

    # Pre-compute log-time grid in ms for slope/dip calculations
    _lt_ms_multi = log_time.values.astype(float) * ms

    key_values = {}
    for i, fname in enumerate(data_cols, start=1):
        _y_raw_p      = np.array(OJIP_double_normalized[fname].values, dtype=float)
        _y_recon_p    = _y_raw_p - np.array(Resid_DF[fname].values, dtype=float)
        _d2_p         = np.array(D2_DF.iloc[:, i].values, dtype=float)
        _qp_start = _fit_start_index(OJIP_double_normalized, trim_first_proc)
        _qp_end   = len(_y_raw_p) - trim_last_proc if trim_last_proc > 0 else len(_y_raw_p)
        fq_p = _fit_quality(_y_raw_p[_qp_start:_qp_end],
                            _y_recon_p[_qp_start:_qp_end],
                            _d2_p, method=fit_method_proc)

        # ── per-file slopes ───────────────────────────────────────────────
        _fj_t_p = _t_safe(FJ_deriv.get(fname), ms) or FJ_time_ms
        _fi_t_p = _t_safe(FI_deriv.get(fname), ms) or FI_time_ms
        _fp_t_p = _t_safe(FP_deriv.get(fname), ms)
        _f0_t_p = float(Summary_file[x_col].iloc[int(F50us_idx)]) * ms
        _f0_v_p = float(F0[fname]); _fj_v_p = float(FJ[fname])
        _fi_v_p = float(FI[fname]); _fm_v_p = float(FM[fname])
        def _slope_p(fe, fs, te, ts):
            dt = te - ts
            return (fe - fs) / dt if dt > 0 and np.isfinite(dt) else np.nan
        _sl_oj = _slope_p(_fj_v_p, _f0_v_p, _fj_t_p, _f0_t_p)
        _sl_ji = _slope_p(_fi_v_p, _fj_v_p, _fi_t_p, _fj_t_p)
        _sl_ip = _slope_p(_fm_v_p, _fi_v_p,
                          _fp_t_p if _fp_t_p else _fi_t_p + 100, _fi_t_p)

        # ── per-file dip detection (FI-FP) ────────────────────────────────
        _d1_p = np.array(D1_DF.iloc[:, i].values, dtype=float)
        _recon_p = np.array(Raw_recon_DF.iloc[:, i].values, dtype=float)
        _fi_t_act = _fi_t_p if _fi_t_p else FI_time_ms
        _fp_t_act = _fp_t_p if _fp_t_p else float(_lt_ms_multi[-1])
        _dip_present = False; _dip_time = None; _dip_amp = None; _dip_d1 = None
        _ip_mask_p = (_lt_ms_multi >= _fi_t_act) & (_lt_ms_multi <= _fp_t_act)
        if np.any(_ip_mask_p):
            _d1_ip_p = _d1_p[_ip_mask_p]
            _t_ip_p = _lt_ms_multi[_ip_mask_p]
            _r_ip_p = _recon_p[_ip_mask_p]
            _d1_mi = np.argmin(_d1_ip_p)
            _d1_mv = float(_d1_ip_p[_d1_mi])
            if _d1_mv < 0:
                _dip_present = True
                _dip_time = float(_t_ip_p[_d1_mi])
                _dip_d1 = _d1_mv
                _fi_r = float(np.interp(_fi_t_act, _lt_ms_multi, _recon_p))
                _fp_r = float(np.interp(_fp_t_act, _lt_ms_multi, _recon_p))
                _dt_t = _fp_t_act - _fi_t_act
                if _dt_t > 0:
                    _frac_p = (_dip_time - _fi_t_act) / _dt_t
                    _dip_amp = (_fi_r + _frac_p * (_fp_r - _fi_r)) - float(_r_ip_p[_d1_mi])

        key_values[fname] = {
            'F0':  _safe(F0[fname]),  'FM': _safe(FM[fname]),
            'FK':  _safe(FK[fname]),  'F50': _safe(F50[fname]),
            'FJ':  _safe(FJ[fname]),  'FI':  _safe(FI[fname]),
            'FJ_time_user_ms':    FJ_time_ms,
            'FI_time_user_ms':    FI_time_ms,
            'FJ_time_deriv_ms':   _t_safe(FJ_deriv.get(fname), ms),
            'FI_time_deriv_ms':   _t_safe(FI_deriv.get(fname), ms),
            'FP_time_deriv_ms':   _t_safe(FP_deriv.get(fname), ms),
            'FJ_time_inflect_ms': _t_safe(FJ_infl.get(fname),  ms),
            'FI_time_inflect_ms': _t_safe(FI_infl.get(fname),  ms),
            'FP_time_inflect_ms': _t_safe(FP_infl.get(fname),  ms),
            'FJ_conf': _safe(fjifp_conf['FJ'].get(fname)),
            'FI_conf': _safe(fjifp_conf['FI'].get(fname)),
            'FP_conf': _safe(fjifp_conf['FP'].get(fname)),
            'FM_time_ms':  _safe(FM_timings_series.get(fname)),
            'Area_OJ': _safe(AREAOJ[fname]), 'Area_JI': _safe(AREAJI[fname]),
            'Area_IP': _safe(AREAIP[fname]), 'Area_OP': _safe(AREAOP[fname]),
            'poly_infl_ms':    poly_oj[fname]['poly_infl_ms'],
            'poly_fi_infl_ms': poly_oi[fname]['poly_infl_ms'],
            # Phase slopes
            'slope_OJ': _safe(_sl_oj), 'slope_JI': _safe(_sl_ji), 'slope_IP': _safe(_sl_ip),
            # FI-FP dip
            'dip_IP_present': _dip_present,
            'dip_IP_time_ms': _safe(_dip_time) if _dip_time is not None else None,
            'dip_IP_amplitude': _safe(_dip_amp) if _dip_amp is not None else None,
            'dip_IP_d1_min': _safe(_dip_d1) if _dip_d1 is not None else None,
            **fq_p,
        }
        # Method-specific parameters
        _me_proc = method_extras_proc.get(fname)
        if _me_proc and fit_method_proc == 'three_exp':
            for k in ('A_OJ', 'A_JI', 'A_IP', 'tau_OJ_ms', 'tau_JI_ms', 'tau_IP_ms'):
                key_values[fname][k] = _me_proc.get(k)
        elif _me_proc and fit_method_proc == 'gaussian_d1':
            for g in range(_me_proc.get('n_gaussians', 0)):
                key_values[fname][f'gauss_center_{g+1}_ms'] = _me_proc['centers_ms'][g] if g < len(_me_proc.get('centers_ms', [])) else None
                key_values[fname][f'gauss_sigma_{g+1}']     = _me_proc['sigmas'][g]     if g < len(_me_proc.get('sigmas', []))     else None
                key_values[fname][f'gauss_amp_{g+1}']       = _me_proc['amplitudes'][g] if g < len(_me_proc.get('amplitudes', [])) else None

    return jsonify({
        'status':      'success',
        'fluorometer': fluorometer,
        'kr':          kr,
        'fj_time_ms':  FJ_time_ms,
        'fi_time_ms':  FI_time_ms,
        'files':       data_cols,
        'file_stem':   last_fname,
        'time_raw_ms': time_raw_ms,
        'time_log_ms': time_log_ms,
        'curves':      curves,
        'key_values':  key_values,
    })


@OJIP_data_analysis.route('/api/ojip_refit', methods=['POST'])
def ojip_refit():
    """
    Re-fit splines with a new kr value.
    Receives JSON: {fluorometer, kr, fj_time_ms, fi_time_ms, time_raw_ms, double_norm: {file:[...]}}
    Returns: {curves: {file:{reconstructed,d1,d2}}, key_timings: {file:{...}}}
    """
    data = request.get_json(force=True)
    fluorometer = data.get('fluorometer', '')
    kr = int(data.get('kr', 10))
    fit_method_refit = data.get('fit_method', 'logspline')
    trim_first_refit = int(data.get('trim_first', 0))
    trim_last_refit  = int(data.get('trim_last',  0))
    knot_placement_refit = data.get('knot_placement', 'quantile')
    oj_densify_refit = bool(data.get('oj_densify', False))
    oj_model_refit = data.get('oj_model', 'exponential')
    oj_model_params_refit = data.get('oj_model_params', None)
    # Legacy compat: old payloads send oj_tau_ms directly
    if oj_model_params_refit is None:
        _tau_raw = data.get('oj_tau_ms', None)
        if _tau_raw is not None and _tau_raw != '':
            oj_model_params_refit = {'tau_ms': float(_tau_raw)}
    FJ_time_ms = float(data.get('fj_time_ms', 2.0))
    FI_time_ms = float(data.get('fi_time_ms', 30.0))
    time_raw_ms = data['time_raw_ms']
    double_norm_dict = data['double_norm']  # {file: [y values]}
    file_names = list(double_norm_dict.keys())

    try:
        x_col, _, _, _, _, ranges = _axis_cfg(fluorometer)
    except ValueError:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 400

    ms = _ms_factor(fluorometer)
    time_native = [t / ms for t in time_raw_ms]

    # Reconstruct double_norm DataFrame in native time units
    dn_df = pd.DataFrame({x_col: time_native})
    for fname, vals in double_norm_dict.items():
        dn_df[fname] = vals

    _spline_method_refit = 'logspline' if fit_method_refit in _DETECTION_METHODS else fit_method_refit
    try:
        Raw_recon_DF, D1_DF, D2_DF, D3_DF, Resid_DF, Infl_DF, log_time, \
            D2_smooth_DF, D3_smooth_DF, densify_info = \
            _fit_curves(dn_df, x_col, kr, method=_spline_method_refit,
                        trim_first=trim_first_refit, trim_last=trim_last_refit,
                        knot_placement=knot_placement_refit,
                        oj_densify=oj_densify_refit,
                        oj_model=oj_model_refit,
                        oj_model_params=oj_model_params_refit)
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'Refit failed.'}), 400

    poly_oj = _fit_oj_polynomial(dn_df, x_col, ms)
    poly_oi = _fit_oj_polynomial(dn_df, x_col, ms, oj_lo_ms=9.0, oj_hi_ms=100.0)

    # ── method-specific fits ─────────────────────────────────────────────────
    three_exp_res_r = _fit_three_exponential(dn_df, x_col, ms) if fit_method_refit == 'three_exp' else None
    piecewise_res_r = _fit_piecewise_linear(dn_df, x_col, ms)  if fit_method_refit == 'piecewise' else None
    gauss_d1_res_r  = _fit_d1_gaussians(D1_DF, x_col, ms)      if fit_method_refit == 'gaussian_d1' else None

    _use_poly_refit = (fit_method_refit == 'polynomial')
    _fjfi_method_refit = fit_method_refit if fit_method_refit in _DETECTION_METHODS else 'd2_trough'
    try:
        FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl, fjifp_conf, method_extras_r = _find_fjfifp(
            D2_DF, D3_DF, x_col, ranges, file_names, Infl_DF,
            fj_expect=FJ_time_ms / ms, fi_expect=FI_time_ms / ms,
            poly_oj=poly_oj if _use_poly_refit else None,
            poly_oi=poly_oi if _use_poly_refit else None,
            ms_factor=ms,
            method=_fjfi_method_refit,
            D1_DF=D1_DF,
            three_exp_results=three_exp_res_r,
            piecewise_results=piecewise_res_r,
            gaussian_d1_results=gauss_d1_res_r)
    except ValueError:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 400

    time_log_ms = (log_time.astype(float) * ms).tolist()
    updated_curves = {}
    for i, fname in enumerate(file_names, start=1):
        curve_entry = {
            'reconstructed': [_safe(v) for v in Raw_recon_DF.iloc[:, i]],
            'd1':            [_safe(v) for v in D1_DF.iloc[:, i]],
            'd2':            [_safe(v) for v in D2_DF.iloc[:, i]],
            'd3':            [_safe(v) for v in D3_DF.iloc[:, i]],
            'd2_smooth':     [_safe(v) for v in D2_smooth_DF.iloc[:, i]] if D2_smooth_DF is not None else None,
            'd3_smooth':     [_safe(v) for v in D3_smooth_DF.iloc[:, i]] if D3_smooth_DF is not None else None,
            'residuals':       [_safe(v) for v in Resid_DF.iloc[:, i]],
            'poly_oj_time_ms': poly_oj[fname]['poly_oj_time_ms'],
            'poly_oj_d2':      poly_oj[fname]['poly_oj_d2'],
            'poly_oi_time_ms': poly_oi[fname]['poly_oj_time_ms'],
            'poly_oi_d2':      poly_oi[fname]['poly_oj_d2'],
        }
        # Method-specific visualisation data
        if fname in method_extras_r:
            curve_entry['method_fit'] = method_extras_r[fname]
        updated_curves[fname] = curve_entry

    key_timings = {}
    for i, fname in enumerate(file_names, start=1):
        _y_raw_dn_r      = np.array(dn_df[fname].values, dtype=float)
        _y_recon_at_raw_r = _y_raw_dn_r - np.array(Resid_DF[fname].values, dtype=float)
        _d2_vals_r       = np.array(D2_DF.iloc[:, i].values, dtype=float)
        _qr_start = 1 + trim_first_refit
        _qr_end   = len(_y_raw_dn_r) - trim_last_refit if trim_last_refit > 0 else len(_y_raw_dn_r)
        fq_r = _fit_quality(_y_raw_dn_r[_qr_start:_qr_end],
                            _y_recon_at_raw_r[_qr_start:_qr_end],
                            _d2_vals_r, method=fit_method_refit)
        kt_entry = {
            'FJ_time_deriv_ms':   _t_safe(FJ_deriv.get(fname), ms),
            'FI_time_deriv_ms':   _t_safe(FI_deriv.get(fname), ms),
            'FP_time_deriv_ms':   _t_safe(FP_deriv.get(fname), ms),
            'FJ_time_inflect_ms': _t_safe(FJ_infl.get(fname),  ms),
            'FI_time_inflect_ms': _t_safe(FI_infl.get(fname),  ms),
            'FP_time_inflect_ms': _t_safe(FP_infl.get(fname),  ms),
            'FJ_conf': _safe(fjifp_conf['FJ'].get(fname)),
            'FI_conf': _safe(fjifp_conf['FI'].get(fname)),
            'FP_conf': _safe(fjifp_conf['FP'].get(fname)),
            'poly_infl_ms':    poly_oj[fname]['poly_infl_ms'],
            'poly_fi_infl_ms': poly_oi[fname]['poly_infl_ms'],
            **fq_r,
        }
        # Method-specific parameters
        me = method_extras_r.get(fname)
        if me and fit_method_refit == 'three_exp':
            for k in ('A_OJ', 'A_JI', 'A_IP', 'tau_OJ_ms', 'tau_JI_ms', 'tau_IP_ms'):
                kt_entry[k] = me.get(k)
        elif me and fit_method_refit == 'gaussian_d1':
            for g in range(me.get('n_gaussians', 0)):
                kt_entry[f'gauss_center_{g+1}_ms'] = me['centers_ms'][g] if g < len(me.get('centers_ms', [])) else None
                kt_entry[f'gauss_sigma_{g+1}']     = me['sigmas'][g]     if g < len(me.get('sigmas', []))     else None
                kt_entry[f'gauss_amp_{g+1}']       = me['amplitudes'][g] if g < len(me.get('amplitudes', [])) else None
        key_timings[fname] = kt_entry

    resp = {
        'status':       'success',
        'time_log_ms':  time_log_ms,
        'curves':       updated_curves,
        'key_timings':  key_timings,
    }
    if densify_info:
        resp['densify_info'] = densify_info
    return jsonify(resp)


@OJIP_data_analysis.route('/api/ojip_add_charts', methods=['POST'])
def ojip_add_charts():
    """
    Create a compact summary xlsx: Parameters + Charts + Group Statistics.
    All data is received from the client — no server-side xlsx file is read.
    """
    data         = request.get_json(force=True)
    charts       = data.get('charts', [])
    file_stem    = secure_filename(data.get('file_stem', 'ojip'))
    group_export = data.get('group_export')
    params_table = data.get('params_table')   # {header: [...], rows: [[...]]}

    summary_fname  = f'{file_stem}_summary.xlsx'
    summary_full   = os.path.join(UPLOAD_FOLDER, summary_fname)
    summary_static = f'uploads/{summary_fname}'

    try:
        wb = Workbook()

        # ── 1. Parameters sheet (data passed from client) ─────────────────────
        ws_params = wb.worksheets[0]
        ws_params.title = 'Parameters'
        if params_table:
            ws_params.append(params_table.get('header', []))
            for r in params_table.get('rows', []):
                ws_params.append(r)

        # ── 2. Charts sheet ────────────────────────────────────────────────────
        ws_charts = wb.create_sheet('Charts')
        row = 1
        for c in charts:
            url = c.get('data_url', '')
            if not url or ',' not in url:
                continue
            b64 = url.split(',', 1)[1]
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64)
            except Exception:
                continue
            if len(img_bytes) < 8:
                continue
            img_buf = io.BytesIO(img_bytes)
            try:
                xl_img = Image(img_buf)
            except Exception:
                continue
            img_buf.seek(0)
            TARGET_W = 700
            orig_w, orig_h = xl_img.width, xl_img.height  # type: ignore[attr-defined]
            if orig_w > 0:
                scale = TARGET_W / orig_w
                xl_img.width  = TARGET_W                      # type: ignore[attr-defined]
                xl_img.height = round(orig_h * scale)         # type: ignore[attr-defined]
            else:
                xl_img.width, xl_img.height = TARGET_W, 400  # type: ignore[attr-defined]

            title = c.get('title', '')
            if title:
                ws_charts.cell(row=row, column=1, value=title)
                row += 1

            xl_img.anchor = f'A{row}'
            ws_charts.add_image(xl_img)
            row += round(xl_img.height / 20) + 2  # type: ignore[attr-defined]

        # ── 3. Group statistics sheets ─────────────────────────────────────────
        if group_export:
            grp_stats    = group_export.get('stats', {})
            samples      = group_export.get('samples', [])
            param_order  = group_export.get('param_order', [])
            param_labels = group_export.get('param_labels', {})
            grp_names    = list(grp_stats.keys())

            if grp_stats and param_order:
                ws_st = wb.create_sheet('Group_Statistics')
                hdr = ['Parameter']
                for g in grp_names:
                    hdr += [f'{g} mean', f'{g} SD', f'{g} N']
                ws_st.append(hdr)
                for p in param_order:
                    stat_row = [param_labels.get(p, p)]
                    for g in grp_names:
                        s = grp_stats.get(g, {}).get('params', {}).get(p)
                        if s:
                            stat_row += [round(s['mean'], 6), round(s['sd'], 6), s.get('n')]
                        else:
                            stat_row += [None, None, None]
                    ws_st.append(stat_row)

            if samples and param_order:
                ws_sp = wb.create_sheet('Group_Samples')
                ws_sp.append(['Sample', 'Group'] + [param_labels.get(p, p) for p in param_order])
                for sr in samples:
                    sample_row = [sr.get('sample'), sr.get('group')]
                    for p in param_order:
                        v = sr.get(p)
                        sample_row.append(round(v, 6) if v is not None else None)
                    ws_sp.append(sample_row)

        # ── 4. Raw curve data sheets ───────────────────────────────────────────
        curve_data = data.get('curve_data')
        if curve_data:
            time_raw = curve_data.get('time_raw_ms', [])
            time_log = curve_data.get('time_log_ms', [])
            cd_files = curve_data.get('files', [])
            curves   = curve_data.get('curves', {})

            def make_curve_sheet(sheet_name, time_arr, key):
                ws = wb.create_sheet(sheet_name)
                ws.append(['time_ms'] + cd_files)
                for i, t in enumerate(time_arr):
                    row = [t]
                    for f in cd_files:
                        arr = curves.get(f, {}).get(key)
                        row.append(arr[i] if arr and i < len(arr) else None)
                    ws.append(row)

            make_curve_sheet('OJIP_raw',           time_raw, 'raw')
            make_curve_sheet('OJIP_to_zero',       time_raw, 'shifted_F0')
            make_curve_sheet('OJIP_to_max',        time_raw, 'shifted_FM')
            make_curve_sheet('OJIP_norm',          time_raw, 'double_norm')
            make_curve_sheet('OJIP_reconstructed', time_log, 'reconstructed')
            make_curve_sheet('1st_derivatives',    time_log, 'd1')
            make_curve_sheet('2nd_derivatives',    time_log, 'd2')
            make_curve_sheet('Residuals',          time_raw, 'residuals')

        # ── 5. Methods sheet ───────────────────────────────────────────────────
        methods_text = data.get('methods_text', '')
        if methods_text:
            ws_meth = wb.create_sheet('Methods')
            ws_meth.column_dimensions['A'].width = 120
            for line in methods_text.split('\n'):
                ws_meth.append([line])

        wb.save(summary_full)
        return jsonify({'status': 'success', 'xlsx_path': summary_static})
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 500


@OJIP_data_analysis.route('/api/ojip_interpret', methods=['POST'])
@csrf.exempt
def ojip_interpret():
    """
    Biological interpretation of JIP-test parameters for one sample.

    Receives JSON:
      {
        "params":          { FVFM, VJ, VI, M0, PSIE0, … F0, FM, FK, FJ, FI, … },
        "organism_class":  "cyanobacteria"|"green_alga"|"plant"|"unknown"|null,
        "reference":       { same keys as params } | null,
        "measurement": {
          "instrument":           "aquapen"|"fl6000"|"multicolor_pam"|…,
          "uses_measuring_light": bool,
          "sp_intensity_umol":    float,
          "culture_density":      float,
          "acclimation":          "dark"|"light"|"partial_dark",
          "dcmu":                 bool,
          "fj_timing_ms":         float
        }
      }

    Returns JSON:
      {
        "status":    "success",
        "findings":  { … structured interpret_ojip output … },
        "narrative": "Plain-language text (LLM or template fallback)"
      }

    Only named scalar parameter values are sent to any LLM; raw transient
    data is never transmitted.
    """
    data = request.get_json(force=True) or {}
    params        = data.get('params', {}) or {}
    organism_class = data.get('organism_class')
    reference     = data.get('reference')
    measurement   = data.get('measurement', {}) or {}

    if not params:
        return jsonify({'status': 'error', 'message': 'No params supplied.'}), 400

    try:
        # Sample interpretation (standalone — no reference for its own assessment)
        sample_findings = interpret_ojip(
            params,
            organism_class=organism_class,
            reference=None,
            measurement=measurement,
        )
        # Comparative interpretation (sample vs reference, when reference is present)
        comparison_findings = interpret_ojip(
            params,
            organism_class=organism_class,
            reference=reference,
            measurement=measurement,
        )
        # suppress "no reference provided" sentence in standalone cards when comparison IS present
        _standalone = reference is not None
        sample_narrative = summarise_findings(sample_findings, suppress_no_ref_caveat=_standalone)

        # Reference interpretation standalone (only when reference is present)
        ref_findings = None
        ref_narrative = None
        comparison_paragraph = None
        if reference:
            ref_findings = interpret_ojip(
                reference,
                organism_class=organism_class,
                reference=None,
                measurement=measurement,
            )
            ref_narrative = summarise_findings(ref_findings, suppress_no_ref_caveat=True)
            comparison_paragraph = compare_ojip_params(params, reference, organism_class)

        narrative = generate_narrative(comparison_findings, params, reference)
        if narrative is None:
            narrative = summarise_findings(comparison_findings)

        return jsonify({
            'status':               'success',
            'findings':             comparison_findings,   # backward-compat: comparison context embedded
            'sample_findings':      sample_findings,
            'sample_narrative':     sample_narrative,
            'ref_findings':         ref_findings,
            'ref_narrative':        ref_narrative,
            'narrative':            narrative,
            'has_reference':        reference is not None,
            'comparison_paragraph': comparison_paragraph,
        })
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'An internal server error occurred.'}), 500


@OJIP_data_analysis.route('/api/ojip_process_batch', methods=['POST'])
@csrf.exempt
def ojip_process_batch():
    """
    Batch-process multiple OJIP curves (params-only or full detail).

    Designed for the multi-curve orchestrator: the client fans batches of
    ~30 curves here with ``include_curves=false`` for the fast params pass,
    then fetches individual curves with ``include_curves=true`` on demand.

    Receives JSON::

        {
          "fluorometer":            str,
          "time_native":            [float, ...],
          "curves":                 [
            {"slot": int, "name": str, "values": [float, ...]}, ...
          ],
          "FJ_time":                float,          // ms
          "FI_time":                float,          // ms
          "knots_reduction_factor": int,
          "include_curves":         bool            // false = params only (fast)
        }

    Returns JSON::

        {
          "status":  "success",
          "results": [
            {slot, name, F0, FM, FVFM, VJ, ...},
            {slot, name, "error": "reason"},
            ...
          ]
        }

    One bad curve never fails the whole batch.
    """
    payload = request.get_json(force=True)
    fluorometer  = payload.get('fluorometer', '')
    time_native  = payload.get('time_native', [])
    curves       = payload.get('curves', [])
    fj_time_ms   = float(payload.get('FJ_time', 2.0))
    fi_time_ms   = float(payload.get('FI_time', 30.0))
    kr           = int(payload.get('knots_reduction_factor', 10))
    include      = bool(payload.get('include_curves', False))
    fit_method   = payload.get('fit_method', 'logspline')
    trim_first   = int(payload.get('trim_first', 0))
    trim_last    = int(payload.get('trim_last',  0))
    background_mode = payload.get('background_mode', 'auto')
    background_n    = int(payload.get('background_n', 1) or 1)
    f0_source       = payload.get('f0_source', 'instrument')
    knot_placement  = payload.get('knot_placement', 'quantile')
    oj_densify      = bool(payload.get('oj_densify', False))
    oj_model        = payload.get('oj_model', 'exponential')
    oj_model_params = payload.get('oj_model_params', None)
    # Legacy compat: old payloads send oj_tau_ms directly
    if oj_model_params is None:
        _tau_raw_b = payload.get('oj_tau_ms', None)
        if _tau_raw_b is not None and _tau_raw_b != '':
            oj_model_params = {'tau_ms': float(_tau_raw_b)}
    f0_raw          = payload.get('f0_time_ms', None)
    f0_time_ms      = float(f0_raw) if f0_raw is not None and f0_raw != '' else None
    use_deriv_timing = bool(payload.get('use_deriv_timing', False))

    if not time_native or not curves:
        return jsonify({'status': 'error',
                        'message': 'Missing time_native or curves.'}), 400
    if fj_time_ms >= fi_time_ms:
        return jsonify({'status': 'error',
                        'message': 'FJ time must be less than FI time.'}), 400
    try:
        _axis_cfg(fluorometer)
    except ValueError:
        return jsonify({'status': 'error',
                        'message': f'Unknown fluorometer: {fluorometer}'}), 400

    time_arr = np.asarray(time_native, dtype=float)
    results  = []
    for c in curves:
        slot = c.get('slot', 0)
        name = c.get('name', '')
        vals = c.get('values', [])
        bckg = c.get('bckg', None)
        fo_footer = c.get('fo_footer', None)
        try:
            r = analyze_one_curve(
                time_arr, vals, name, fluorometer,
                fj_time_ms, fi_time_ms, kr,
                include_curves=include,
                fit_method=fit_method,
                trim_first=trim_first,
                trim_last=trim_last,
                background_mode=background_mode,
                background_n=background_n,
                f0_source=f0_source,
                bckg=bckg,
                fo_footer=fo_footer,
                knot_placement=knot_placement,
                oj_densify=oj_densify,
                oj_model=oj_model,
                oj_model_params=oj_model_params,
                f0_time_ms=f0_time_ms,
                use_deriv_timing=use_deriv_timing,
            )
            r['slot'] = slot
            r['name'] = name
            results.append(r)
        except Exception as exc:
            results.append({'slot': slot, 'name': name,
                            'error': str(exc)})

    return jsonify({'status': 'success', 'results': results})


# ─── batch ZIP export ──────────────────────────────────────────────────────
@OJIP_data_analysis.route('/api/ojip_export_batch', methods=['POST'])
@csrf.exempt
def ojip_export_batch():
    """
    Build a ZIP archive containing the batch analysis results.

    Receives JSON::

        {
          "params_header": [str, ...],        // column names
          "params_rows":   [[val, ...], ...],  // one row per curve
          "charts":        {canvasId: dataURL, ...},  // base64 PNGs
          "stem":          str,                // base filename
          "include_plots": {                   // per-plot-type flags
            "raw": bool, "shifted_F0": bool, "shifted_FM": bool,
            "double_norm": bool, "reconstructed": bool,
            "d2": bool, "d3": bool, "residuals": bool
          },
          "curve_data":    {name: {time_raw_ms, time_log_ms, curves, key_values}, ...}
        }

    Returns a ZIP file as a binary download.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    payload = request.get_json(force=True)
    params_header = payload.get('params_header', [])
    params_rows   = payload.get('params_rows', [])
    charts        = payload.get('charts', {})
    stem          = payload.get('stem', 'ojip_batch')
    curve_data    = payload.get('curve_data', {})

    # Granular plot flags (with backwards compat for old include_curves/include_diag)
    inc = payload.get('include_plots', {})
    if not inc:
        # Legacy fallback
        ic = payload.get('include_curves', False)
        id_ = payload.get('include_diag', False)
        inc = {
            'raw': ic, 'shifted_F0': ic, 'shifted_FM': ic, 'double_norm': ic,
            'reconstructed': id_, 'd2': id_, 'd3': id_, 'residuals': id_,
        }

    try:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

            # ── 1. Parameters XLSX ────────────────────────────────────────
            xlsx_buf = io.BytesIO()
            wb = Workbook()
            ws = cast(Worksheet, wb.active)  # a new Workbook always has an active sheet
            ws.title = 'Parameters'
            if params_header:
                ws.append(params_header)
            for row in params_rows:
                ws.append(row)
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            zf.writestr('params_summary.xlsx', xlsx_buf.getvalue())

            # ── 2. Summary chart PNGs (captured from client canvases) ─────
            # Dynamic chart keys: param-FVFM, mc-panel-WT, mc-compare-chart, etc.
            for cid, data_url in charts.items():
                if not data_url or ',' not in data_url:
                    continue
                b64 = data_url.split(',', 1)[1]
                try:
                    img_bytes = base64.b64decode(b64)
                except Exception:
                    continue
                # Route each chart type to an appropriate subdirectory
                if cid.startswith('param-'):
                    fname = f'parameters/{cid[6:]}.png'
                elif cid.startswith('mc-panel-'):
                    fname = f'panels/{cid[9:]}.png'
                elif cid == 'mc-compare-chart':
                    fname = 'compare.png'
                elif cid == 'mc-aggregate-chart':
                    fname = 'aggregate_curves.png'
                else:
                    fname = f'{cid}.png'
                zf.writestr(f'summary_plots/{fname}', img_bytes)

            # ── 3+4. Individual per-curve plots (flat folder per type) ─────
            if curve_data:
                _render_per_curve(zf, curve_data, inc)

        zip_buf.seek(0)
        dl_name = stem.replace(' ', '_') + '_export.zip'
        return send_file(zip_buf, mimetype='application/zip',
                         as_attachment=True, download_name=dl_name)
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error',
                        'message': 'An internal error occurred during export.'}), 500


# ─── Incremental (chunked) batch export ────────────────────────────────────
# DEPRECATED (2026-08): Client-side Canvas 2D rendering + JSZip now handles
# all plot generation and ZIP assembly in the browser.  These three endpoints
# are retained for rollback only — the JS no longer calls them.
#
# Original purpose: for large datasets (>100 curves) the monolithic
# ojip_export_batch can timeout or hit payload limits.  These three endpoints
# let the client build the ZIP incrementally:
#   1. _start  → create a temp ZIP with params + summary charts
#   2. _add    → render a batch of ~50 curves and append PNGs to the ZIP
#   3. _finish → return the completed ZIP and clean up


@OJIP_data_analysis.route('/api/ojip_export_start', methods=['POST'])
@csrf.exempt
def ojip_export_start():
    """Create a temp ZIP and write params XLSX + summary chart PNGs."""
    import matplotlib
    matplotlib.use('Agg')

    payload = request.get_json(force=True)
    params_header = payload.get('params_header', [])
    params_rows   = payload.get('params_rows', [])
    charts        = payload.get('charts', {})
    stem          = payload.get('stem', 'ojip_batch')
    method_info   = payload.get('method_info', {})

    export_id = uuid.uuid4().hex[:12]
    zip_path  = os.path.join(UPLOAD_FOLDER, f'_export_{export_id}.zip')

    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Params XLSX
            xlsx_buf = io.BytesIO()
            wb = Workbook()
            ws = cast(Worksheet, wb.active)
            ws.title = 'Parameters'
            if params_header:
                ws.append(params_header)
            for row in params_rows:
                ws.append(row)
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            zf.writestr('params_summary.xlsx', xlsx_buf.getvalue())

            # Method info text file
            if method_info:
                zf.writestr('method_info.txt',
                            _format_method_info(method_info))

            # Summary chart PNGs
            for cid, data_url in charts.items():
                if not data_url or ',' not in data_url:
                    continue
                b64 = data_url.split(',', 1)[1]
                try:
                    img_bytes = base64.b64decode(b64)
                except Exception:
                    continue
                if cid.startswith('param-'):
                    fname = f'parameters/{cid[6:]}.png'
                elif cid.startswith('mc-panel-'):
                    fname = f'panels/{cid[9:]}.png'
                elif cid == 'mc-compare-chart':
                    fname = 'compare.png'
                elif cid == 'mc-aggregate-chart':
                    fname = 'aggregate_curves.png'
                else:
                    fname = f'{cid}.png'
                zf.writestr(f'summary_plots/{fname}', img_bytes)

        # Log what was written for diagnostics
        with zipfile.ZipFile(zip_path, 'r') as zr:
            entries = zr.namelist()
        print(f'[export_start] id={export_id}, '
              f'params_rows={len(params_rows)}, '
              f'charts={len(charts)}, '
              f'method_info={"yes" if method_info else "no"}, '
              f'zip_entries={entries}')
        return jsonify({'status': 'success', 'export_id': export_id,
                        'stem': stem})
    except Exception:
        import traceback; traceback.print_exc()
        if os.path.exists(zip_path):
            os.remove(zip_path)
        return jsonify({'status': 'error',
                        'message': 'Failed to start export.'}), 500


@OJIP_data_analysis.route('/api/ojip_export_add', methods=['POST'])
@csrf.exempt
def ojip_export_add():
    """Render PNGs for a batch of curves and append them to the temp ZIP."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    payload   = request.get_json(force=True)
    export_id = payload.get('export_id', '')
    inc       = payload.get('include_plots', {})
    curve_data = payload.get('curve_data', {})

    zip_path = os.path.join(UPLOAD_FOLDER, f'_export_{export_id}.zip')
    if not os.path.isfile(zip_path):
        return jsonify({'status': 'error',
                        'message': 'Unknown export_id.'}), 404

    try:
        added = 0
        with zipfile.ZipFile(zip_path, 'a', zipfile.ZIP_DEFLATED) as zf:
            added = _render_per_curve(zf, curve_data, inc)

        return jsonify({'status': 'success', 'added': added})
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error',
                        'message': 'Failed to add curves to export.'}), 500


@OJIP_data_analysis.route('/api/ojip_export_finish/<export_id>', methods=['GET'])
def ojip_export_finish(export_id):
    """Return the completed ZIP and delete the temp file."""
    # Sanitise export_id (hex only)
    if not export_id or not all(c in '0123456789abcdef' for c in export_id):
        return jsonify({'status': 'error', 'message': 'Invalid export_id.'}), 400

    zip_path = os.path.join(UPLOAD_FOLDER, f'_export_{export_id}.zip')
    if not os.path.isfile(zip_path):
        return jsonify({'status': 'error', 'message': 'Export not found.'}), 404

    try:
        # Read into memory so we can delete the temp file
        with open(zip_path, 'rb') as f:
            data = f.read()
        os.remove(zip_path)

        buf = io.BytesIO(data)
        stem = request.args.get('stem', 'ojip_batch')
        dl_name = stem.replace(' ', '_') + '_export.zip'
        return send_file(buf, mimetype='application/zip',
                         as_attachment=True, download_name=dl_name)
    except Exception:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error',
                        'message': 'Failed to finalise export.'}), 500


# ─── helpers for batch ZIP export ──────────────────────────────────────────

def _render_per_curve(zf, curve_data, inc):
    """Render per-curve PNGs and write them into *zf* (a ZipFile).

    Folder layout is flat per plot type:
        raw/<name>.png, shifted_F0/<name>.png, …
        reconstructed/<name>.png, d2/<name>.png, d3/<name>.png, residuals/<name>.png

    Returns the number of PNGs added.
    """
    added = 0
    curve_types = [
        ('raw',         'Raw'),
        ('shifted_F0',  'Shifted F0'),
        ('shifted_FM',  'Shifted FM'),
        ('double_norm', 'Double normalised'),
    ]

    for name, cd in curve_data.items():
        safe = _safe_zip_name(name)
        time_raw = cd.get('time_raw_ms', [])
        time_log = cd.get('time_log_ms', [])
        curves_d = cd.get('curves', {})
        kv       = cd.get('key_values', {})

        # Curve plots  →  raw/<name>.png  etc.
        for norm_key, label in curve_types:
            if not inc.get(norm_key):
                continue
            y_data = curves_d.get(norm_key)
            if y_data and time_raw:
                png = _make_curve_png(time_raw, y_data, name, label, kv)
                zf.writestr(f'{norm_key}/{safe}.png', png)
                added += 1

        # Reconstructed  →  reconstructed/<name>.png
        if inc.get('reconstructed'):
            recon = curves_d.get('reconstructed')
            dnorm = curves_d.get('double_norm')
            if recon and time_log:
                png = _make_diag_png(time_log, recon, name,
                                     'Reconstructed', time_raw, dnorm, kv)
                zf.writestr(f'reconstructed/{safe}.png', png)
                added += 1

        # D2  →  d2/<name>.png
        if inc.get('d2'):
            d2 = curves_d.get('d2_smooth') or curves_d.get('d2')
            if d2 and time_log:
                png = _make_deriv_png(time_log, d2, name,
                                     'D2 (2nd derivative)', kv)
                zf.writestr(f'd2/{safe}.png', png)
                added += 1

        # D3  →  d3/<name>.png
        if inc.get('d3'):
            d3 = curves_d.get('d3_smooth') or curves_d.get('d3')
            if d3 and time_log:
                png = _make_deriv_png(time_log, d3, name,
                                     'D3 (3rd derivative)', kv)
                zf.writestr(f'd3/{safe}.png', png)
                added += 1

        # Residuals  →  residuals/<name>.png
        if inc.get('residuals'):
            resid = curves_d.get('residuals')
            if resid and time_raw:
                png = _make_deriv_png(time_raw, resid, name, 'Residuals', kv)
                zf.writestr(f'residuals/{safe}.png', png)
                added += 1

    return added


def _format_method_info(mi: dict) -> str:
    """Format the method_info dict into a human-readable text summary."""
    _METHOD_NAMES = {
        'logspline':   'Log-time spline (quintic, analytic D2/D3)',
        'spline':      'Standard LSQ spline (linear, numeric D2)',
        'pchip':       'PCHIP (monotone piecewise cubic)',
        'polynomial':  'Polynomial inflection (Akinyemi et al. 2023)',
        'd1_minima':   'D1 local minima',
        'three_exp':   '3-Exponential decomposition (Boisvert et al. 2006)',
        'piecewise':   'Piecewise-linear breakpoints',
        'gaussian_d1': 'Gaussian D1 deconvolution',
    }
    _SPLINE_METHODS = {'logspline', 'spline', 'pchip'}
    fm = mi.get('fit_method', 'logspline')
    lines = [
        'OJIP Batch Export — Analysis Method Summary',
        '=' * 46,
        '',
        f'Instrument:             {mi.get("fluorometer", "—")}',
        f'Total curves:           {mi.get("total_curves", "—")}',
        '',
        '— Curve fitting —',
        f'Fitting method:         {_METHOD_NAMES.get(fm, fm)}',
    ]
    if fm in _SPLINE_METHODS:
        lines.append(
            f'FJ / FI detection:      D2 troughs (2nd derivative minima)')
    else:
        lines.append(
            f'FJ / FI detection:      Method-specific '
            f'(reconstruction via log-time spline)')
    lines += [
        f'Knot reduction (kr):    {mi.get("knots_reduction", "—")}',
        f'Knot placement:         {mi.get("knot_placement", "—")}',
        f'FJ search window:       {mi.get("FJ_time_ms", "—")} ms',
        f'FI search window:       {mi.get("FI_time_ms", "—")} ms',
    ]
    tf = mi.get('trim_first', 0)
    tl = mi.get('trim_last', 0)
    if tf or tl:
        lines.append(f'Trim:                   first {tf}, last {tl} points')
    f0t = mi.get('f0_time_ms')
    if f0t:
        lines.append(f'F0 timing override:     {f0t} ms')
    lines += [
        '',
        '— Background / F0 —',
        f'Background mode:        {mi.get("background_mode", "—")}',
        f'Background points (n):  {mi.get("background_n", "—")}',
        f'F0 source:              {mi.get("f0_source", "—")}',
    ]

    # O-J densify
    lines += ['', '— O-J densification —']
    if mi.get('oj_densify'):
        model = mi.get('oj_model', 'exponential')
        lines.append(f'Enabled:                yes')
        lines.append(f'Model:                  {model}')
        mp = mi.get('oj_model_params') or {}
        if model == 'exponential':
            tau = mp.get('tau_ms', 'auto')
            lines.append(f'  tau:                  {tau} ms')
        elif model == 'biexponential':
            lines.append(f'  tau1:                 {mp.get("tau1_ms", "auto")} ms')
            lines.append(f'  tau2:                 {mp.get("tau2_ms", "auto")} ms')
        elif model == 'connectivity':
            lines.append(f'  p (connectivity):     {mp.get("p", "auto")}')
            lines.append(f'  k_L:                  {mp.get("k_L", "auto")} ms-1')
            lines.append(f'  k_ox:                 {mp.get("k_ox", 0)} ms-1')
        elif model == 'linear':
            lines.append(f'  (no tuneable params)')
    else:
        lines.append(f'Enabled:                no')

    lines += [
        '',
        '— Generated by cyano.tools OJIP analysis —',
        'https://www.cyano.tools',
        '',
    ]
    return '\n'.join(lines)


def _safe_zip_name(name: str) -> str:
    """Sanitise a curve name for use as a ZIP path component."""
    import re
    s = re.sub(r'[<>:"/\\|?*]', '_', str(name))
    return s.strip().strip('.') or 'unnamed'


def _make_curve_png(time_ms, y_data, title, norm_label, kv=None):
    """Render one OJIP curve plot as a PNG byte string (log x-axis)."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    t = np.asarray(time_ms, dtype=float)
    y = np.asarray(y_data, dtype=float)
    # Only plot positive times for log scale
    mask = t > 0
    t, y = t[mask], y[mask]

    fig, ax = plt.subplots(figsize=(5, 3))
    ax.semilogx(t, y, 'b-', linewidth=1)
    ax.set_xlabel('Time (ms)')
    ax.set_ylabel('Fluorescence')
    ax.set_title(f'{title} — {norm_label}')

    # Mark FJ, FI, FP if available
    if kv:
        for phase, marker, colour in [
            ('FJ', '^', '#e6550d'), ('FI', 'D', '#31a354'), ('FP', 's', '#756bb1')
        ]:
            t_key = f'{phase}_time_deriv_ms'
            v_key = phase if phase != 'FP' else 'FM'
            tv = kv.get(t_key)
            fv = kv.get(v_key) if phase != 'FP' else kv.get('FM')
            if tv and fv:
                ax.plot(tv, fv, marker=marker, color=colour, markersize=8,
                        label=phase, zorder=5)
        ax.legend(fontsize=8)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=72, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _make_diag_png(time_log, recon, title, label,
                   time_raw=None, raw_data=None, kv=None):
    """Render a reconstructed-vs-raw diagnostic plot with FJ/FI/FP markers."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(5, 3))
    tl = np.asarray(time_log, dtype=float)
    yr = np.asarray(recon, dtype=float)

    # Raw data (measured)
    if raw_data and time_raw:
        tr = np.asarray(time_raw, dtype=float)
        yd = np.asarray(raw_data, dtype=float)
        mask = tr > 0
        ax.semilogx(tr[mask], yd[mask], 'b.', markersize=2, alpha=0.5,
                     label='Measured')

    # Reconstructed (fitted) curve
    ax.semilogx(tl, yr, 'r-', linewidth=1.2, label='Fitted')

    # FJ/FI/FP markers (interpolated onto the reconstructed curve)
    _add_phase_markers(ax, tl, yr, kv)

    ax.set_xlabel('Time (ms)')
    ax.set_ylabel('Fluorescence (double norm.)')
    ax.set_title(title)
    ax.legend(fontsize=8)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=72, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _make_deriv_png(time_arr, y_data, title, label, kv=None):
    """Render a derivative or residual plot with FJ/FI/FP markers."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(5, 3))
    t = np.asarray(time_arr, dtype=float)
    y = np.asarray(y_data, dtype=float)
    mask = t > 0
    ax.semilogx(t[mask], y[mask], 'g-', linewidth=1)
    ax.axhline(0, color='grey', linewidth=0.5, linestyle='--')

    # FJ/FI/FP markers (interpolated onto this derivative curve)
    _add_phase_markers(ax, t[mask], y[mask], kv)

    ax.set_xlabel('Time (ms)')
    ax.set_ylabel(label)
    ax.set_title(title)
    if kv:
        ax.legend(fontsize=8)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=72, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _add_phase_markers(ax, t_arr, y_arr, kv):
    """Add FJ / FI / FP markers to an axes, interpolated onto (t_arr, y_arr)."""
    if not kv:
        return
    t_np = np.asarray(t_arr, dtype=float)
    y_np = np.asarray(y_arr, dtype=float)
    for phase, marker, colour, label in [
        ('FJ', '^', '#e6550d', 'J'),
        ('FI', 'D', '#31a354', 'I'),
        ('FP', 's', '#756bb1', 'P'),
    ]:
        tv = kv.get(f'{phase}_time_deriv_ms')
        if tv is None:
            continue
        # Linear interpolation onto the curve
        yv = float(np.interp(tv, t_np, y_np))
        ax.plot(tv, yv, marker=marker, color=colour, markersize=8,
                label=label, zorder=5, linestyle='none')

    