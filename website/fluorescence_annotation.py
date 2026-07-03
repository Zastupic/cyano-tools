"""
fluorescence_annotation.py — Fluorescence Data Annotation Tool (M1 back-end core).

Blueprint: /fluorescence_annotation  (GET — UI page)
           /api/fluorescence_annotation/schema       (GET  — field schema for JS)
           /api/fluorescence_annotation/ingest       (POST — parse files → grid rows)
           /api/fluorescence_annotation/export       (POST — grid rows → ZIP bundle)
           /api/fluorescence_annotation/load_bundle  (POST — ZIP → grid state)

Design:
  - Four metadata tiers (Investigation / Study / Assay / Per-curve) with inheritance.
  - Each field cell carries a provenance flag: typed | from_header | from_filename |
    inherited | computed.
  - Existing parsers (_ms_factor, _detect_fluorometer, _parse_transient) and JIP-test
    computation (_compute_jip_params → calls ojip_interpretation.interpret_ojip) are
    implemented here as private helpers; OJIP_data_analysis.py is not modified.
  - Bundle format: ZIP containing state.json + dataset.parquet + MANIFEST.json.
  - Parquet requires pyarrow; a CSV fallback is used when pyarrow is absent.
  - Raw transients are held in a temp JSON file in static/uploads/ (same 30-min
    cleanup daemon as the rest of the app) between /ingest and /export calls.

Run  `python website/fluorescence_annotation.py`  for a smoke test.
"""

from __future__ import annotations

import base64
import hashlib
import io
import json
import os
import re
import uuid
import zipfile
from dataclasses import dataclass, field as dc_field
from typing import Any

import numpy as np
import pandas as pd
from flask import Blueprint, jsonify, render_template, request
from werkzeug.utils import secure_filename

SCHEMA_VERSION = "2.0.0"

# ── Provenance flag constants ─────────────────────────────────────────────────
TYPED         = "typed"
FROM_HEADER   = "from_header"
FROM_FILENAME = "from_filename"
INHERITED     = "inherited"
COMPUTED      = "computed"
MISSING       = "missing"


# ── Field registry ────────────────────────────────────────────────────────────

@dataclass
class FieldDef:
    tier:        str              # investigation | study | fluorometer | per_curve
    required:    str              # mandatory | recommended | optional | conditional
    dtype:       str              # str | float | int
    label:       str = ""
    vocab:       list = dc_field(default_factory=list)
    condition:   str = ""         # legacy condition expression (unused)
    miappe:      str = ""         # MIAPPE 1.1 / MICF mapping
    weight:      float = 0.0      # completeness scoring (0 = not scored)
    group:       str = ""         # grid column group: identity|biological|treatment|conditions|replicate_qc|acquisition
    sample_cond: str = ""         # conditional on sample_type: "" (all) | "liquid_culture" | "plant" | "plant_or_leaf"


FIELDS: dict[str, FieldDef] = {

    # ── Investigation (once per project) ─────────────────────────────────────
    "project_title":          FieldDef("investigation", "mandatory", "str",
                                       "Project title", weight=3.0,
                                       miappe="Investigation title"),
    "contact_name":           FieldDef("investigation", "recommended", "str",
                                       "Contact name", weight=1.0,
                                       miappe="Investigation person name"),
    "contact_email":          FieldDef("investigation", "recommended", "str",
                                       "Contact e-mail", weight=1.0,
                                       miappe="Investigation person email"),
    "institution":            FieldDef("investigation", "recommended", "str",
                                       "Institution", weight=1.0,
                                       miappe="Investigation person affiliation"),
    "contributor_namespace":  FieldDef("investigation", "recommended", "str",
                                       "Contributor namespace (ORCID / ROR)",
                                       miappe="Investigation person id"),
    "license":                FieldDef("investigation", "optional", "str",
                                       "Licence",
                                       vocab=["CC-BY-4.0", "CC-BY-NC-4.0",
                                              "CC0-1.0", "Proprietary"],
                                       miappe="License"),
    "project_description":    FieldDef("investigation", "optional", "str",
                                       "Description",
                                       miappe="Investigation description"),

    # ── Study (once per experiment) ───────────────────────────────────────────
    "organism":               FieldDef("study", "mandatory", "str", "Organism",
                                       vocab=["Synechocystis sp. PCC 6803",
                                              "Synechococcus sp. PCC 7942",
                                              "Anabaena sp. PCC 7120",
                                              "Thermosynechococcus elongatus BP-1",
                                              "Chlamydomonas reinhardtii",
                                              "Chlorella vulgaris",
                                              "Arabidopsis thaliana",
                                              "Spinacia oleracea", "Other"],
                                       weight=2.0, miappe="Organism",
                                       group="biological"),
    "genotype":               FieldDef("study", "recommended", "str", "Genotype",
                                       weight=2.0,
                                       miappe="Biological material source ID",
                                       group="biological"),
    "sub_strain_cultivar":    FieldDef("study", "optional", "str",
                                       "Sub-strain / Cultivar",
                                       miappe="Infraspecific name",
                                       group="biological"),
    "medium":                 FieldDef("study", "recommended", "str", "Medium",
                                       vocab=["BG-11", "BG-11\u2080", "M2", "TAP",
                                              "f/2", "1/2 MS", "Hoagland", "Other"],
                                       weight=1.0, miappe="Growth facility",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "medium_modification":    FieldDef("study", "optional", "str",
                                       "Medium modification",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "trophic_mode":           FieldDef("study", "recommended", "str",
                                       "Trophic mode",
                                       vocab=["autotrophy", "mixotrophy",
                                              "heterotrophy"],
                                       weight=1.0, miappe="Growth condition",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "cultivator_type":        FieldDef("study", "optional", "str",
                                       "Cultivator type",
                                       vocab=["Erlenmeyer flask (orbital shaking)",
                                              "Erlenmeyer flask (static)",
                                              "flat-bottom flask",
                                              "flat-panel photobioreactor",
                                              "tubular photobioreactor",
                                              "bubble-column photobioreactor",
                                              "airlift bioreactor",
                                              "stirred-tank bioreactor",
                                              "multiwell plate (6-well)",
                                              "multiwell plate (12-well)",
                                              "multiwell plate (24-well)",
                                              "multiwell plate (96-well)",
                                              "Petri dish",
                                              "test tube / culture tube",
                                              "culture bag",
                                              "raceway pond",
                                              "thin-layer cascade",
                                              "trough / tray",
                                              "carboy",
                                              "custom / other"],
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "growth_light_intensity": FieldDef("study", "recommended", "float",
                                       "Growth light intensity (\u03bcmol m\u207b\u00b2 s\u207b\u00b9)",
                                       weight=1.0, miappe="Growth condition",
                                       group="conditions"),
    "growth_temperature":     FieldDef("study", "recommended", "float",
                                       "Growth temp (\u00b0C)", weight=1.0,
                                       miappe="Growth condition",
                                       group="conditions"),
    "growth_co2":             FieldDef("study", "optional", "float",
                                       "Growth CO\u2082 (%)",
                                       miappe="Growth condition",
                                       group="conditions"),
    "growth_light_type":      FieldDef("study", "optional", "str",
                                       "Growth light type",
                                       vocab=["monochromatic LED",
                                              "multi-color / broadband LED",
                                              "white LED",
                                              "fluorescent lamp",
                                              "metal halide",
                                              "sunlight / daylight",
                                              "other"],
                                       group="conditions"),
    "growth_light_peak_wl":   FieldDef("study", "optional", "int",
                                       "Growth light peak wavelength (nm)",
                                       group="conditions"),
    "growth_light_peak_width": FieldDef("study", "optional", "int",
                                        "Growth light peak width / FWHM (nm)",
                                        group="conditions"),
    "growth_light_color_cat": FieldDef("study", "optional", "str",
                                       "Growth light color category",
                                       vocab=["warm white (<3300 K)",
                                              "neutral white (3300\u20135000 K)",
                                              "cool white (>5000 K)"],
                                       group="conditions"),
    "growth_light_note":      FieldDef("study", "optional", "str",
                                       "Growth light source note",
                                       group="conditions"),
    "acclimation_min":        FieldDef("fluorometer", "recommended", "float",
                                       "Dark pre-acclimation time (min)", weight=1.0,
                                       miappe="MICF:darkAdaptationDuration",
                                       group="acquisition"),
    "actinic_preaccl_intensity":     FieldDef("fluorometer", "optional", "float",
                                       "Actinic light intensity prior to OJIP",
                                       group="acquisition"),
    "actinic_preaccl_wavelength_nm": FieldDef("fluorometer", "optional", "int",
                                       "Actinic light wavelength prior to OJIP (nm)",
                                       group="acquisition"),
    "preaccl_temperature":           FieldDef("fluorometer", "optional", "float",
                                       "Pre-acclimation temperature (\u00b0C)",
                                       group="acquisition"),
    "preaccl_co2":                   FieldDef("fluorometer", "optional", "float",
                                       "Pre-acclimation CO\u2082 (%)",
                                       group="acquisition"),
    "sample_type":            FieldDef("study", "mandatory", "str", "Sample type",
                                       vocab=["liquid culture", "vascular plant",
                                              "detached leaf / disc"],
                                       weight=2.0,
                                       miappe="Observation unit type",
                                       group="identity"),

    # ── Fluorometer settings (acquisition protocol, shared across a run) ──────
    "instrument":             FieldDef("fluorometer", "mandatory", "str", "Instrument",
                                       vocab=["MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)",
                                              "Aquapen", "FL6000"],
                                       weight=2.0, miappe="MICF:instrumentModel",
                                       group="acquisition"),
    "sat_pulse_intensity":    FieldDef("fluorometer", "recommended", "float",
                                       "Sat. pulse intensity (\u03bcmol m\u207b\u00b2 s\u207b\u00b9)",
                                       weight=1.0,
                                       miappe="MICF:saturatingPulseIntensity",
                                       group="acquisition"),
    "sat_pulse_wavelength_nm": FieldDef("fluorometer", "optional", "float",
                                        "Sat. pulse wavelength (nm)",
                                        miappe="MICF:saturatingPulseWavelength",
                                        group="acquisition"),
    "sat_pulse_duration_s":   FieldDef("fluorometer", "optional", "float",
                                       "Sat. pulse duration (s)",
                                       miappe="MICF:saturatingPulseDuration",
                                       group="acquisition"),
    "meas_light_intensity":   FieldDef("fluorometer", "optional", "float",
                                       "Measuring light (\u03bcmol m\u207b\u00b2 s\u207b\u00b9)",
                                       miappe="MICF:measuringLightIntensity",
                                       group="acquisition"),
    "meas_light_wavelength_nm": FieldDef("fluorometer", "optional", "float",
                                         "Measuring light peak \u03bb (nm)",
                                         miappe="MICF:measuringLightWavelength",
                                         group="acquisition"),
    "fo_timing":              FieldDef("fluorometer", "recommended", "str",
                                       "F0 timing convention",
                                       vocab=["20 \u00b5s", "50 \u00b5s"],
                                       weight=2.0,
                                       miappe="MICF:F0TimingConvention",
                                       group="acquisition"),

    # ── Per-curve (repeats per file) ──────────────────────────────────────────

    # Identity
    "curve_id":               FieldDef("per_curve", "mandatory", "str", "Curve ID",
                                       weight=3.0, miappe="MICF:curveID",
                                       group="identity"),
    "filename":               FieldDef("per_curve", "mandatory", "str", "Filename",
                                       weight=3.0, group="identity"),
    "sample_id":              FieldDef("per_curve", "recommended", "str", "Sample ID",
                                       weight=2.0, miappe="Sample name",
                                       group="identity"),

    # Treatment (template-assigned or manual per-curve)
    "treatment_label":        FieldDef("per_curve", "recommended", "str",
                                       "Treatment group",
                                       weight=3.0, miappe="Factor value",
                                       group="treatment"),
    "chem_treatment":         FieldDef("per_curve", "optional", "str",
                                       "Chemical treatment",
                                       vocab=["control", "DCMU",
                                              "methyl viologen", "KCN",
                                              "glycolaldehyde", "DBMIB",
                                              "hydroxylamine", "lincomycin",
                                              "other"],
                                       miappe="Factor value", group="treatment"),
    "chem_dose":              FieldDef("per_curve", "optional", "str",
                                       "Chemical dose",
                                       miappe="Factor value", group="treatment"),
    "chem_duration":          FieldDef("per_curve", "optional", "str",
                                       "Chemical duration",
                                       miappe="Factor value", group="treatment"),
    "chem_unit":              FieldDef("per_curve", "optional", "str",
                                       "Chemical dose unit",
                                       vocab=["\u00b5M", "mM", "M",
                                              "mg L\u207b\u00b9",
                                              "\u00b5g L\u207b\u00b9",
                                              "ng L\u207b\u00b9", "%", "other"],
                                       miappe="Factor value", group="treatment"),
    "chem_detail":            FieldDef("per_curve", "optional", "str",
                                       "Chemical treatment detail",
                                       miappe="Factor value", group="treatment"),
    "stress_treatment":       FieldDef("per_curve", "optional", "str",
                                       "Stress treatment",
                                       vocab=["high light", "low light",
                                              "heat", "cold", "UV-B",
                                              "nitrogen starvation",
                                              "phosphorus starvation",
                                              "sulfur starvation",
                                              "iron starvation",
                                              "salt stress", "drought",
                                              "other"],
                                       miappe="Factor value", group="treatment"),
    "stress_dose":            FieldDef("per_curve", "optional", "str",
                                       "Stress dose / intensity",
                                       miappe="Factor value", group="treatment"),
    "stress_duration":        FieldDef("per_curve", "optional", "str",
                                       "Stress duration",
                                       miappe="Factor value", group="treatment"),
    "stress_unit":            FieldDef("per_curve", "optional", "str",
                                       "Stress dose unit",
                                       vocab=["\u00b5mol photons m\u207b\u00b2 s\u207b\u00b9",
                                              "\u00b0C", "mM NaCl",
                                              "\u00b5W cm\u207b\u00b2",
                                              "% field capacity", "% RH",
                                              "other"],
                                       miappe="Factor value", group="treatment"),
    "stress_detail":          FieldDef("per_curve", "optional", "str",
                                       "Stress treatment detail",
                                       miappe="Factor value", group="treatment"),
    "other_treatment":        FieldDef("per_curve", "optional", "str",
                                       "Other treatment",
                                       miappe="Factor value", group="treatment"),
    "other_dose":             FieldDef("per_curve", "optional", "str",
                                       "Other treatment dose",
                                       miappe="Factor value", group="treatment"),
    "other_unit":             FieldDef("per_curve", "optional", "str",
                                       "Other treatment unit",
                                       miappe="Factor value", group="treatment"),
    "other_duration":         FieldDef("per_curve", "optional", "str",
                                       "Other treatment duration",
                                       miappe="Factor value", group="treatment"),
    "other_detail":           FieldDef("per_curve", "optional", "str",
                                       "Other treatment detail",
                                       miappe="Factor value", group="treatment"),
    "timepoint":              FieldDef("per_curve", "recommended", "float",
                                       "Time (h)", weight=2.0,
                                       miappe="Observation unit factor value",
                                       group="treatment"),

    # Conditions — shared across all sample types
    "temperature":            FieldDef("per_curve", "recommended", "float",
                                       "Temp \u00b0C", weight=2.0,
                                       miappe="MICF:measurementTemperature",
                                       group="conditions"),

    # Culture density (study tier — set once, inherited to all curves; override per-curve in grid)
    "culture_density_chla":   FieldDef("study", "recommended", "float",
                                       "Chl a (\u03bcg mL\u207b\u00b9)",
                                       weight=1.0,
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "culture_density_od":     FieldDef("study", "optional", "float",
                                       "OD (1 cm cuvette)",
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "culture_density_od_wl":  FieldDef("study", "optional", "int",
                                       "OD wavelength (nm)",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "culture_density_other":  FieldDef("study", "optional", "str",
                                       "Other density (value)",
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "culture_density_other_unit": FieldDef("study", "optional", "str",
                                       "Other density unit",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "co2":                    FieldDef("per_curve", "optional", "float",
                                       "CO\u2082 (%)",
                                       miappe="Growth condition",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "vessel":                 FieldDef("per_curve", "optional", "str",
                                       "Culture vessel",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "agitation":              FieldDef("per_curve", "optional", "str",
                                       "Agitation",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "growth_phase":           FieldDef("per_curve", "recommended", "str",
                                       "Growth phase",
                                       vocab=["lag", "exponential", "linear",
                                              "stationary", "decline"],
                                       weight=1.0, miappe="Growth condition",
                                       group="conditions",
                                       sample_cond="liquid_culture"),
    "culture_age_h":          FieldDef("per_curve", "optional", "float",
                                       "Culture age (h since dilution)",
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="liquid_culture"),

    # Conditions — vascular plant (MIAPPE-aligned)
    "growth_facility":        FieldDef("per_curve", "recommended", "str",
                                       "Growth facility",
                                       vocab=["field", "greenhouse",
                                              "growth chamber", "other"],
                                       weight=1.0, miappe="Growth facility",
                                       group="conditions", sample_cond="plant"),
    "dev_stage":              FieldDef("per_curve", "optional", "str",
                                       "Dev. stage",
                                       vocab=["seedling", "young",
                                              "adult / vegetative",
                                              "flowering", "senescent"],
                                       miappe="Developmental stage",
                                       group="conditions",
                                       sample_cond="plant_or_leaf"),
    "plant_organ":            FieldDef("per_curve", "optional", "str",
                                       "Plant organ",
                                       vocab=["whole leaf", "leaf disc",
                                              "cotyledon", "stem",
                                              "root", "whole plant"],
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="plant_or_leaf"),
    "leaf_position":          FieldDef("per_curve", "optional", "str",
                                       "Leaf position",
                                       vocab=["1st true leaf", "2nd true leaf",
                                              "3rd true leaf", "flag leaf",
                                              "rosette centre", "rosette outer"],
                                       miappe="Sample description",
                                       group="conditions",
                                       sample_cond="plant_or_leaf"),
    "leaf_surface":           FieldDef("per_curve", "optional", "str",
                                       "Leaf surface",
                                       vocab=["adaxial", "abaxial"],
                                       group="conditions",
                                       sample_cond="plant_or_leaf"),
    "photoperiod":            FieldDef("per_curve", "optional", "str",
                                       "Photoperiod",
                                       miappe="Growth condition",
                                       group="conditions", sample_cond="plant"),
    "humidity":               FieldDef("per_curve", "optional", "float",
                                       "Humidity (%)",
                                       miappe="Growth condition",
                                       group="conditions", sample_cond="plant"),
    "substrate":              FieldDef("per_curve", "optional", "str",
                                       "Substrate / soil",
                                       miappe="Growth condition",
                                       group="conditions", sample_cond="plant"),
    "watering_regime":        FieldDef("per_curve", "optional", "str",
                                       "Watering regime",
                                       miappe="Growth condition",
                                       group="conditions", sample_cond="plant"),

    # Replicate / QC
    "bio_rep":                FieldDef("per_curve", "recommended", "int",
                                       "Bio. rep.", weight=1.0,
                                       miappe="Biological replicate",
                                       group="replicate_qc"),
    "tech_rep":               FieldDef("per_curve", "recommended", "int",
                                       "Tech. rep.", weight=1.0,
                                       miappe="Technical replicate",
                                       group="replicate_qc"),
    "batch_id":               FieldDef("per_curve", "optional", "str",
                                       "Batch ID", weight=2.0,
                                       miappe="Sample description",
                                       group="replicate_qc"),
    "quality":                FieldDef("per_curve", "optional", "str",
                                       "Quality",
                                       vocab=["ok", "noisy",
                                              "saturation artifact",
                                              "low signal", "exclude"],
                                       weight=1.0, miappe="Quality",
                                       group="replicate_qc"),

    # Acquisition (per-curve; some inherited from fluorometer tier)
    "gain":                   FieldDef("per_curve", "optional", "float",
                                       "Gain", miappe="MICF:gain",
                                       group="acquisition"),
    "timestamp":              FieldDef("per_curve", "optional", "str",
                                       "Timestamp",
                                       miappe="MICF:acquisitionTimestamp",
                                       group="acquisition"),
}


# ── Schema helpers ────────────────────────────────────────────────────────────

def get_schema_json() -> dict:
    """Serialise FIELDS to a plain dict suitable for jsonify / JS consumption."""
    return {
        "schema_version": SCHEMA_VERSION,
        "fields": {
            name: {
                "tier":        f.tier,
                "required":    f.required,
                "dtype":       f.dtype,
                "label":       f.label,
                "vocab":       f.vocab,
                "condition":   f.condition,
                "miappe":      f.miappe,
                "weight":      f.weight,
                "group":       f.group,
                "sample_cond": f.sample_cond,
            }
            for name, f in FIELDS.items()
        },
    }


def validate_row(row: dict) -> list[str]:
    """
    Validate field values in a grid row.
    Returns a list of human-readable issue strings (empty = valid).
    """
    issues: list[str] = []
    for field_name, fdef in FIELDS.items():
        cell = row.get(field_name)
        if not isinstance(cell, dict):
            continue
        val = cell.get("value")
        if val is None or val == "":
            continue
        if fdef.vocab and str(val) not in fdef.vocab:
            issues.append(
                f"{fdef.label or field_name}: '{val}' is not in "
                f"the allowed vocabulary {fdef.vocab}"
            )
        if fdef.dtype == "float":
            try:
                float(val)
            except (ValueError, TypeError):
                issues.append(f"{fdef.label or field_name}: expected a number, got '{val}'")
        elif fdef.dtype == "int":
            try:
                int(float(str(val)))
            except (ValueError, TypeError):
                issues.append(f"{fdef.label or field_name}: expected an integer, got '{val}'")
    return issues


def _completeness_score(row: dict) -> float:
    """
    Completeness score 0–100: fraction of non-empty cells across all annotation
    grid fields (study + fluorometer + per_curve tiers).  Investigation-tier fields
    are excluded because they are entered in project-level forms, not in the per-curve
    grid, so including them in per-curve completeness would be misleading.
    Each field counts as 1 (simple field count, not weighted) so the score matches
    the user's visual impression of 'how many cells in this row are filled'.
    Conditional fields (sample_cond) are only counted when the row's sample_type matches.
    """
    st_cell = row.get("sample_type")
    sample_type: str = (st_cell.get("value", "") or "") if isinstance(st_cell, dict) else ""
    is_liquid = "liquid" in sample_type
    is_plant  = "vascular" in sample_type or "plant" in sample_type
    is_leaf   = "leaf" in sample_type or "disc" in sample_type

    total = earned = 0
    for field_name, fdef in FIELDS.items():
        if fdef.tier == "investigation":
            continue           # investigation fields are not in the per-curve grid
        if field_name == "completeness_score":
            continue
        sc = fdef.sample_cond
        if sc == "liquid_culture" and not is_liquid:
            continue
        if sc == "plant" and not is_plant:
            continue
        if sc == "plant_or_leaf" and not (is_plant or is_leaf):
            continue
        total += 1             # every field counts as 1 — no weights
        cell = row.get(field_name)
        if isinstance(cell, dict) and cell.get("value") not in (None, "", []):
            earned += 1
    return round(100.0 * earned / total, 1) if total > 0 else 0.0


# ── Instrument helpers ────────────────────────────────────────────────────────

def _ms_factor(fluorometer: str) -> float:
    """Multiplier from native time unit to milliseconds."""
    if fluorometer == "Aquapen":
        return 0.001    # µs → ms
    if fluorometer == "FL6000":
        return 1000.0   # s  → ms
    return 1.0          # MULTI-COLOR-PAM already in ms


def _detect_fluorometer(filename: str, raw_bytes: bytes) -> str | None:
    """
    Auto-detect instrument from file extension and content.
    Mirrors the detection logic in OJIP_data_analysis.py.
    """
    ext = os.path.splitext(filename)[1].lower()
    try:
        head = raw_bytes[:2000].decode("utf-8", errors="replace")
    except Exception:
        return None

    if ext == ".csv":
        if "time/ms" in head:
            return "MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)"
    elif ext == ".txt":
        if re.search(r"AquaPen|FluorPen", head, re.IGNORECASE):
            return "Aquapen"
        if re.search(r"Fluorometer", head, re.IGNORECASE):
            return "FL6000"
    return None


# ── Transient parser ──────────────────────────────────────────────────────────

def _parse_transient(
    raw_bytes: bytes, filename: str, fluorometer: str
) -> tuple[pd.DataFrame, str, float]:
    """
    Parse a single fluorescence file to a two-column DataFrame [x_col, 'fluorescence'].
    Returns (df, x_col, ms_factor).
    Raises ValueError on unrecognised format or empty data.
    """
    ms = _ms_factor(fluorometer)

    if fluorometer == "MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)":
        df = pd.read_csv(io.BytesIO(raw_bytes), sep=";", engine="python")
        if str(df.columns[0]) != "time/ms":
            raise ValueError(f"Expected 'time/ms' column, got '{df.columns[0]}'")
        df = (df.iloc[:, [0, 1]]
                .rename(columns={df.columns[1]: "fluorescence"})
                .apply(pd.to_numeric, errors="coerce")
                .dropna()
                .reset_index(drop=True))
        return df, "time/ms", ms

    elif fluorometer == "Aquapen":
        text = raw_bytes.decode("utf-8", errors="replace")
        rows = []
        for line in text.splitlines():
            parts = line.strip().split("\t")
            if len(parts) >= 2:
                try:
                    rows.append((int(parts[0]), float(parts[1])))
                except (ValueError, IndexError):
                    pass
        if len(rows) < 2:
            raise ValueError("No numeric data found in Aquapen/FluorPen file")
        df = pd.DataFrame(rows[1:], columns=["time_us", "fluorescence"])  # skip t=0
        return df, "time_us", ms

    elif fluorometer == "FL6000":
        text = raw_bytes.decode("utf-8", errors="replace")
        lines = text.splitlines()
        start = next(
            (i + 1 for i, ln in enumerate(lines) if ln.strip() == "Time"), None
        )
        if start is None:
            raise ValueError("'Time' header not found in FL6000 file")
        rows = []
        for line in lines[start:]:
            parts = line.strip().split("\t")
            if len(parts) >= 2:
                try:
                    rows.append((float(parts[0]), float(parts[1])))
                except (ValueError, IndexError):
                    pass
        if not rows:
            raise ValueError("No numeric data found in FL6000 file")
        df = pd.DataFrame(rows, columns=["time_s", "fluorescence"])
        return df, "time_s", ms

    else:
        raise ValueError(f"Unknown fluorometer: {fluorometer!r}")


def _downsample_transient(
    df: pd.DataFrame, x_col: str, ms_factor: float, max_points: int = 2000
) -> dict:
    """
    Downsample a transient DataFrame to at most max_points rows.
    Preserves the fast O→J region (t < 0.5 ms) at full resolution.
    Returns {'time_ms': [...], 'fluorescence': [...]}.
    """
    x_ms = df[x_col].astype(float) * ms_factor
    y    = df["fluorescence"].astype(float)

    if len(df) <= max_points:
        return {"time_ms": x_ms.tolist(), "fluorescence": y.tolist()}

    mask_fast = x_ms < 0.5
    df_fast = df[mask_fast]
    df_slow = df[~mask_fast]

    n_fast = min(len(df_fast), max_points // 4)
    n_slow = max_points - n_fast

    step_f = max(1, len(df_fast) // n_fast) if n_fast > 0 else 1
    step_s = max(1, len(df_slow) // n_slow) if n_slow > 0 else 1

    df_out = pd.concat([
        df_fast.iloc[::step_f],
        df_slow.iloc[::step_s],
    ]).reset_index(drop=True)

    x_out = df_out[x_col].astype(float) * ms_factor
    y_out = df_out["fluorescence"].astype(float)
    return {"time_ms": x_out.tolist(), "fluorescence": y_out.tolist()}


# ── JIP parameter computation ─────────────────────────────────────────────────

def _sf(v) -> float | None:
    """Safe float: returns None for NaN/Inf/non-numeric."""
    try:
        f = float(v)
        return None if (np.isnan(f) or np.isinf(f)) else round(f, 6)
    except Exception:
        return None


def _compute_jip_params(
    df: pd.DataFrame,
    x_col: str,
    ms_factor: float,
    FJ_time_ms: float = 2.0,
    FI_time_ms: float = 30.0,
) -> dict:
    """
    Compute core JIP-test parameters from a parsed transient.
    FJ_time_ms / FI_time_ms are in milliseconds.
    Returns a flat dict with jip_* keys matching FIELDS.
    All arithmetic mirrors OJIP_data_analysis.py; spline fitting is omitted
    (only scalar reference-point parameters are needed for metadata annotation).
    """
    if len(df) < 10:
        return {"jip_error": "too_few_points"}

    x = df[x_col].values.astype(float)
    y = df["fluorescence"].values.astype(float)
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]

    if len(x) < 10:
        return {"jip_error": "too_few_valid_points"}

    def at_ms(t_ms: float) -> float:
        """Fluorescence value at time t_ms (converts ms → native unit)."""
        t_native = t_ms / ms_factor
        return float(y[np.argmin(np.abs(x - t_native))])

    def idx_at_ms(t_ms: float) -> int:
        t_native = t_ms / ms_factor
        return int(np.argmin(np.abs(x - t_native)))

    # F0: at 0.01 ms for Walz (pre-illumination baseline removed), at t≈0 for others
    F0 = at_ms(0.01) if ms_factor == 1.0 else float(y[0])
    FM = float(np.max(y))
    FV = FM - F0

    if FV <= 0:
        return {"jip_F0": _sf(F0), "jip_FM": _sf(FM), "jip_error": "FV_zero_or_negative"}

    # Reference-time fluorescence values (times are the same in ms for all instruments)
    F50 = at_ms(0.05)   # 50 µs in ms
    FK  = at_ms(0.3)    # 300 µs in ms
    FJ  = at_ms(FJ_time_ms)
    FI  = at_ms(FI_time_ms)

    VJ     = (FJ - F0) / FV
    VI     = (FI - F0) / FV
    Fv_Fm  = FV / FM
    M0     = 4.0 * (FK - F50) / FV

    psiE0   = 1.0 - VJ
    psiR0   = 1.0 - VI
    deltaR0 = psiR0 / psiE0 if psiE0 != 0 else None
    phiE0   = Fv_Fm * psiE0
    phiR0   = Fv_Fm * psiR0
    TR0RC   = M0 / VJ if VJ != 0 else None
    ABS_RC  = TR0RC / Fv_Fm if (TR0RC is not None and Fv_Fm != 0) else None
    ET0_RC  = TR0RC * psiE0 if TR0RC is not None else None
    RE0_RC  = TR0RC * psiR0 if TR0RC is not None else None
    DI0_RC  = (ABS_RC - TR0RC) if (ABS_RC is not None and TR0RC is not None) else None

    # Complementary areas (rectangle − trapezoid, time axis in ms)
    x_ms   = x * ms_factor
    FJ_idx = idx_at_ms(FJ_time_ms)
    FI_idx = idx_at_ms(FI_time_ms)
    FM_idx = int(np.argmax(y))

    def safe_area(x_seg, y_seg, height) -> float | None:
        if len(x_seg) < 2:
            return None
        rect = float(x_seg[-1] - x_seg[0]) * height
        trap  = float(np.trapezoid(y_seg, x_seg))
        return rect - trap

    Area_OJ = safe_area(x_ms[:FJ_idx + 1], y[:FJ_idx + 1], FM) if FJ_idx > 0 else None
    Area_JI = safe_area(x_ms[FJ_idx:FI_idx + 1], y[FJ_idx:FI_idx + 1], FM) if FI_idx > FJ_idx else None
    Area_IP = safe_area(x_ms[FI_idx:FM_idx + 1], y[FI_idx:FM_idx + 1], FM) if FM_idx > FI_idx else None
    Area_OP = safe_area(x_ms[:FM_idx + 1], y[:FM_idx + 1], FM) if FM_idx > 0 else None

    Sm = Area_OP / FV if (Area_OP is not None and FV != 0) else None
    N  = Sm * M0 / VJ if (Sm is not None and VJ != 0) else None

    return {
        "jip_F0":      _sf(F0),
        "jip_FM":      _sf(FM),
        "jip_FK":      _sf(FK),
        "jip_FJ":      _sf(FJ),
        "jip_FI":      _sf(FI),
        "jip_VJ":      _sf(VJ),
        "jip_VI":      _sf(VI),
        "jip_Fv_Fm":   _sf(Fv_Fm),
        "jip_M0":      _sf(M0),
        "jip_psiE0":   _sf(psiE0),
        "jip_psiR0":   _sf(psiR0),
        "jip_deltaR0": _sf(deltaR0),
        "jip_phiE0":   _sf(phiE0),
        "jip_phiR0":   _sf(phiR0),
        "jip_ABS_RC":  _sf(ABS_RC),
        "jip_TR0_RC":  _sf(TR0RC),
        "jip_ET0_RC":  _sf(ET0_RC),
        "jip_RE0_RC":  _sf(RE0_RC),
        "jip_DI0_RC":  _sf(DI0_RC),
        "jip_Area_OJ": _sf(Area_OJ),
        "jip_Area_JI": _sf(Area_JI),
        "jip_Area_IP": _sf(Area_IP),
        "jip_Area_OP": _sf(Area_OP),
        "jip_Sm":      _sf(Sm),
        "jip_N":       _sf(N),
    }


# ── Filename token parser ─────────────────────────────────────────────────────

def _parse_filename(stem: str, token_dict: dict) -> dict[str, tuple[Any, str]]:
    """
    Apply a positional token convention to a filename stem.
    Returns {field_name: (value, FROM_FILENAME)} for every matched token.

    token_dict schema:
      {
        "separator": "_",          # split character (default "_")
        "tokens": [
          {"position": 0, "field": "strain"},
          {"position": 1, "field": "treatment",
           "vocab_map": {"HL": "high_light", "LL": "low_light"}},
          {"position": 2, "field": "timepoint_h", "strip_suffix": "h"},
          {"position": 3, "field": "replicate_id", "strip_prefix": "rep"}
        ]
      }
    """
    if not token_dict or not token_dict.get("tokens"):
        return {}

    sep   = token_dict.get("separator", "_")
    parts = stem.split(sep)
    result: dict[str, tuple[Any, str]] = {}

    for tok in token_dict["tokens"]:
        pos = tok.get("position", -1)
        if not (0 <= pos < len(parts)):
            continue
        raw: Any = parts[pos]
        field_name = tok.get("field", "")
        if not field_name:
            continue

        if "vocab_map" in tok:
            raw = tok["vocab_map"].get(raw, raw)
        if "strip_suffix" in tok and isinstance(raw, str) and raw.endswith(tok["strip_suffix"]):
            raw = raw[: -len(tok["strip_suffix"])]
        if "strip_prefix" in tok and isinstance(raw, str) and raw.startswith(tok["strip_prefix"]):
            raw = raw[len(tok["strip_prefix"]) :]

        fdef = FIELDS.get(field_name)
        if fdef:
            if fdef.dtype == "float":
                try:
                    raw = float(raw)
                except (ValueError, TypeError):
                    pass
            elif fdef.dtype == "int":
                try:
                    raw = int(float(str(raw)))
                except (ValueError, TypeError):
                    pass

        result[field_name] = (raw, FROM_FILENAME)

    return result


# ── Tier inheritance ──────────────────────────────────────────────────────────

def _inherit_tiers(
    investigation: dict, study: dict, fluor: dict
) -> dict[str, tuple[Any, str]]:
    """
    Merge tier blocks into a flat inherited dict.
    Order: investigation → study → fluorometer (later overrides earlier on collision).
    Returns {field_name: (value, INHERITED)}.
    """
    result: dict[str, tuple[Any, str]] = {}
    for tier_data in (investigation, study, fluor):
        for key, val in tier_data.items():
            if val not in (None, "", [], {}):
                result[key] = (val, INHERITED)
    return result


# ── Grid row builder ──────────────────────────────────────────────────────────

def _build_row(
    raw_bytes: bytes,
    filename: str,
    fluorometer: str | None,
    inherited: dict,
    token_dict: dict,
) -> dict:
    """
    Build a single reconciliation-grid row for one uploaded file.

    Each field in the returned dict has the shape:
        {field_name: {"value": <any>, "provenance": <flag_constant>}}

    Internal key:
        _meta — parse metadata (fluorometer, parse_error, curve_id)

    Note: No JIP parameters are computed here; this is a metadata-only
    annotation tool. Raw transient data is not stored.
    """
    stem = os.path.splitext(secure_filename(filename))[0].lower()
    row: dict[str, Any] = {}

    # 1. Apply inherited tier values first (lowest priority)
    for field_name, (val, prov) in inherited.items():
        row[field_name] = {"value": val, "provenance": prov}

    # 2. Apply filename tokens (override inherited for per_curve fields only)
    for field_name, (val, prov) in _parse_filename(stem, token_dict).items():
        if FIELDS.get(field_name, FieldDef("", "", "")).tier == "per_curve":
            row[field_name] = {"value": val, "provenance": prov}

    # 3. Auto-detect instrument from file content if not supplied in fluor tier
    fluo = fluorometer or _detect_fluorometer(filename, raw_bytes)
    if fluo:
        row["instrument"] = {"value": fluo, "provenance": FROM_HEADER}

    # 4. curve_id = sha-256 content hash (globally unique without coordination)
    curve_id = hashlib.sha256(raw_bytes).hexdigest()[:16]
    row["curve_id"] = {"value": curve_id, "provenance": COMPUTED}
    row["filename"] = {"value": filename,  "provenance": FROM_HEADER}

    # 5. Default tech_rep = 1 (overridable per-curve)
    if "tech_rep" not in row:
        row["tech_rep"] = {"value": 1, "provenance": COMPUTED}

    # 6. Completeness score (computed last so all fields are present)
    row["completeness_score"] = {"value": _completeness_score(row), "provenance": COMPUTED}

    # 7. Internal metadata (not a FIELDS entry)
    parse_error: str | None = None
    if not fluo:
        parse_error = "instrument not detected — set it in Fluorometer settings"
    row["_meta"] = {
        "filename":    filename,
        "fluorometer": fluo,
        "parse_error": parse_error,
        "curve_id":    curve_id,
    }

    return row


# ── Temporary transient store ─────────────────────────────────────────────────

def _get_upload_folder() -> str:
    try:
        from . import UPLOAD_FOLDER  # type: ignore[attr-defined]
        return UPLOAD_FOLDER
    except ImportError:
        return os.path.join(os.path.dirname(__file__), "static", "uploads")


def _transient_store_path(bundle_id: str) -> str:
    return os.path.join(_get_upload_folder(), f"_ann_{bundle_id}.json")


def _write_transient_store(bundle_id: str, store: dict) -> None:
    if not store:
        return
    try:
        folder = _get_upload_folder()
        os.makedirs(folder, exist_ok=True)
        with open(_transient_store_path(bundle_id), "w", encoding="utf-8") as fh:
            json.dump(store, fh)
    except Exception:
        pass  # non-fatal; export will lack transient data


def _read_transient_store(bundle_id: str) -> dict:
    try:
        path = _transient_store_path(bundle_id)
        if os.path.exists(path):
            with open(path, encoding="utf-8") as fh:
                return json.load(fh)
    except Exception:
        pass
    return {}


# ── Parquet builder ────────────────────────────────────────────────────────────

def _build_summary_parquet(state: dict) -> bytes | None:
    """
    Build a summary Parquet with one row per curve (no time-series).
    Per-field provenance is embedded in Parquet file-level metadata as JSON.
    Falls back to None when pyarrow is not installed.
    """
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        return None

    records:  list[dict]      = []
    prov_map: dict[str, dict] = {}

    for row in state.get("rows", []):
        cid       = row.get("curve_id", {}).get("value", "")
        meta_flat: dict[str, Any] = {}
        prov_flat: dict[str, str] = {}
        for field_name, cell in row.items():
            if field_name.startswith("_") or field_name == "completeness_score":
                continue
            if isinstance(cell, dict) and "value" in cell:
                meta_flat[field_name] = cell["value"]
                prov_flat[field_name] = cell.get("provenance", "")
        prov_map[cid] = prov_flat
        records.append(meta_flat)

    if not records:
        return None

    df_out = pd.DataFrame(records)
    table  = pa.Table.from_pandas(df_out, preserve_index=False)
    table  = table.replace_schema_metadata({
        b"schema_version": SCHEMA_VERSION.encode(),
        b"tool":           b"cyano.tools/fluorescence_annotation",
        b"provenance":     json.dumps(prov_map).encode(),
    })

    buf = io.BytesIO()
    pq.write_table(table, buf)
    return buf.getvalue()


# ── Bundle serialisation ──────────────────────────────────────────────────────

def bundle_serialise(state: dict) -> bytes:
    """
    Pack annotation state into a ZIP bundle:
        MANIFEST.json               — bundle identity and counts
        state.json                  — tier blocks + token dict + per-curve rows with provenance
        annot/<curve_id>.json       — per-curve sidecar (full metadata + provenance)
        summary.parquet             — one row per curve, all metadata (no time-series)
        summary.csv                 — CSV fallback when pyarrow is absent

    state schema:
        {
          'investigation': {field: value, ...},
          'study':         {field: value, ...},
          'fluor':         {field: value, ...},
          'token_dict':    { ... },
          'rows': [{field_name: {'value': ..., 'provenance': ...}, '_meta': {...}}, ...]
        }
    """
    bundle_id = uuid.uuid4().hex
    buf = io.BytesIO()

    rows = state.get("rows", [])

    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:

        # ── MANIFEST ──────────────────────────────────────────────────────────
        manifest: dict[str, Any] = {
            "schema_version": SCHEMA_VERSION,
            "bundle_id":      bundle_id,
            "created":        pd.Timestamp.now().isoformat(timespec="seconds"),
            "curve_count":    len(rows),
            "tool":           "cyano.tools/fluorescence_annotation",
        }
        zf.writestr("MANIFEST.json", json.dumps(manifest, indent=2))

        # ── STATE (provenance-preserving) ─────────────────────────────────────
        rows_clean = [
            {k: v for k, v in row.items() if not k.startswith("_")}
            for row in rows
        ]
        state_out = {
            "schema_version":     SCHEMA_VERSION,
            "investigation":      state.get("investigation", {}),
            "study":              state.get("study", {}),
            "fluor":              state.get("fluor", {}),
            "token_dict":         state.get("token_dict", {}),
            "treatment_templates": state.get("treatment_templates", []),
            "rows":               rows_clean,
        }
        zf.writestr("state.json",
                    json.dumps(state_out, indent=2, ensure_ascii=False))

        # ── PER-CURVE SIDECARS ────────────────────────────────────────────────
        for row in rows:
            cid = (row.get("curve_id") or {}).get("value") or ""
            fname = (row.get("filename") or {}).get("value") or cid
            sidecar = {
                "schema_version": SCHEMA_VERSION,
                "curve_id":       cid,
                "source_file":    fname,
                "investigation":  state.get("investigation", {}),
                "study":          state.get("study", {}),
                "fluor":          state.get("fluor", {}),
                "metadata":       {
                    k: v for k, v in row.items()
                    if not k.startswith("_") and k not in ("completeness_score",)
                },
            }
            safe = re.sub(r"[^a-zA-Z0-9_.-]", "_", fname)[:60]
            zf.writestr(f"annot/annot_{safe}.json",
                        json.dumps(sidecar, indent=2, ensure_ascii=False))

        # ── SUMMARY PARQUET (or CSV fallback) ─────────────────────────────────
        parquet_bytes = _build_summary_parquet(state)
        if parquet_bytes:
            zf.writestr("summary.parquet", parquet_bytes)
            manifest["parquet_included"] = True
        else:
            # CSV fallback: flat metadata only
            rows_meta = []
            for row in rows:
                flat = {k: v.get("value") for k, v in row.items()
                        if isinstance(v, dict) and "value" in v
                        and not k.startswith("_")}
                rows_meta.append(flat)
            if rows_meta:
                csv_buf = io.StringIO()
                pd.DataFrame(rows_meta).to_csv(csv_buf, index=False)
                zf.writestr("summary.csv", csv_buf.getvalue())
                manifest["csv_fallback"] = True

    return buf.getvalue()


def bundle_deserialise(data: bytes) -> dict:
    """
    Unpack a ZIP bundle and return the state dict.
    Raises ValueError on missing state.json or version mismatch.
    """
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()
        if "state.json" not in names:
            raise ValueError("Not a valid annotation bundle: state.json is missing")
        state = json.loads(zf.read("state.json").decode("utf-8"))
        if "MANIFEST.json" in names:
            state["_manifest"] = json.loads(zf.read("MANIFEST.json").decode("utf-8"))
        else:
            state["_manifest"] = {}

    bundle_ver = state.get("schema_version", "0.0.0")
    if bundle_ver != SCHEMA_VERSION:
        state["_version_warning"] = (
            f"Bundle schema version {bundle_ver} differs from "
            f"current tool version {SCHEMA_VERSION}."
        )
    # Back-compat: bundles saved with "assay" key (v1.x) → rename to "fluor"
    if "assay" in state and "fluor" not in state:
        state["fluor"] = state.pop("assay")
    return state


def _build_row_from_ojip_params(
    fname: str,
    fluorometer: str | None,
    inherited: dict,
    token_dict: dict,
    session_salt: str = "",
) -> dict:
    """
    Build an annotation grid row from the OJIP analysis context.
    No raw file bytes available; curve_id is a sha-256 of (session_salt + fname),
    where session_salt is the per-ingest bundle UUID — guaranteeing global uniqueness
    even when two sessions process identically-named files.
    No JIP parameters are stored in the annotation table (metadata only).
    """
    row: dict = {}

    # Apply inherited tier metadata (investigation + study + fluorometer)
    for k, (v, p) in inherited.items():
        if v not in (None, "", []):
            row[k] = {"value": v, "provenance": p}

    # Apply filename tokens for per_curve fields
    fname_tokens = _parse_filename(os.path.splitext(fname)[0], token_dict)
    for k, (v, p) in fname_tokens.items():
        if k in FIELDS and FIELDS[k].tier == "per_curve" and v not in (None, ""):
            row[k] = {"value": v, "provenance": p}

    # Core identifiers
    row["filename"] = {"value": fname, "provenance": FROM_HEADER}
    # sha-256 of (session_salt + fname) — session_salt is the per-ingest bundle UUID,
    # so two sessions that process the same filename still produce distinct curve_ids.
    row["curve_id"] = {
        "value":      hashlib.sha256((session_salt + fname).encode("utf-8")).hexdigest()[:16],
        "provenance": COMPUTED,
    }

    # Instrument — from OJIP auto-detection
    if fluorometer:
        row["instrument"] = {"value": fluorometer, "provenance": FROM_HEADER}

    # Default tech_rep = 1
    if "tech_rep" not in row:
        row["tech_rep"] = {"value": 1, "provenance": COMPUTED}

    # Completeness score
    row["completeness_score"] = {
        "value":      _completeness_score(row),
        "provenance": COMPUTED,
    }

    row["_meta"] = {"fluorometer": fluorometer, "source": "ojip_analysis"}
    return row


# ── Flask Blueprint + routes ──────────────────────────────────────────────────

fluorescence_annotation = Blueprint("fluorescence_annotation", __name__)


@fluorescence_annotation.route("/fluorescence_annotation", methods=["GET"])
def annotation_page():
    return render_template("fluorescence_annotation.html")


@fluorescence_annotation.route("/api/fluorescence_annotation/schema", methods=["GET"])
def schema_route():
    """Return the field schema as JSON (used by JS to build dropdowns)."""
    return jsonify(get_schema_json())


@fluorescence_annotation.route("/api/fluorescence_annotation/ingest_from_ojip", methods=["POST"])
def ingest_from_ojip():
    """
    Build annotation rows from already-computed OJIP results.
    Accepts a JSON body instead of raw file uploads, so no re-parsing is needed.
    No JIP parameters are stored — this is a metadata-only annotation.

    JSON body:
        ojip_results  — { files, fluorometer }
        tier_json     — { investigation, study, fluor, token_dict }
    """
    payload = request.get_json(force=True, silent=True) or {}
    ojip    = payload.get("ojip_results", {})
    tier    = payload.get("tier_json",    {})

    files       = ojip.get("files", [])
    fluorometer = ojip.get("fluorometer") or None

    if not files:
        return jsonify({"status": "error",
                        "message": "No files in ojip_results."}), 400

    investigation = tier.get("investigation", {})
    study         = tier.get("study",         {})
    # Accept both "fluor" (new) and "assay" (legacy) key names
    fluor         = tier.get("fluor") or tier.get("assay") or {}
    token_dict    = tier.get("token_dict",    {})
    treatment_templates = tier.get("treatment_templates", [])

    # If instrument not set in fluor tier, fill from OJIP detection.
    if fluorometer and not fluor.get("instrument"):
        fluor = dict(fluor)
        fluor["instrument"] = fluorometer

    inherited = _inherit_tiers(investigation, study, fluor)

    rows:     list[dict] = []
    warnings: list[str]  = []
    bundle_id: str       = str(uuid.uuid4())

    for fname in files:
        # Coerce to str — Pandas may produce numeric column labels for digit-only
        # filenames, which JSON serialises as numbers; str() restores them cleanly.
        row = _build_row_from_ojip_params(str(fname), fluorometer, inherited, token_dict,
                                          session_salt=bundle_id)

        issues = validate_row(row)
        row["_warnings"] = issues
        if issues:
            warnings.extend(issues)

        rows.append(row)

    return jsonify({
        "status":        "ok",
        "bundle_id":     bundle_id,
        "rows":          rows,
        "tier_defaults": {
            "investigation":      investigation,
            "study":              study,
            "fluor":              fluor,
            "token_dict":         token_dict,
            "treatment_templates": treatment_templates,
        },
        "warnings":       warnings,
        "schema_version": SCHEMA_VERSION,
    })


@fluorescence_annotation.route("/api/fluorescence_annotation/ingest", methods=["POST"])
def ingest():
    """
    Receive raw fluorescence files + tier JSON.
    Returns pre-filled reconciliation grid rows.

    Form fields:
        fluo_files   — one or more raw fluorescence files
        tier_json    — JSON string: {investigation, study, assay, token_dict,
                                     FJ_time_ms (opt), FI_time_ms (opt)}
    """
    if "fluo_files" not in request.files:
        return jsonify({"status": "error", "message": "No files received."}), 400

    files = request.files.getlist("fluo_files")
    if not files or secure_filename(files[0].filename or "") == "":
        return jsonify({"status": "error", "message": "Please select one or more files."}), 400

    if len(files) > 200:
        return jsonify({"status": "error",
                        "message": "Maximum 200 files per batch."}), 400

    try:
        tier_json = json.loads(request.form.get("tier_json", "{}"))
    except json.JSONDecodeError:
        return jsonify({"status": "error", "message": "Invalid tier_json."}), 400

    investigation = tier_json.get("investigation", {})
    study         = tier_json.get("study", {})
    # Accept both "fluor" (new) and "assay" (legacy) key names
    fluor         = tier_json.get("fluor") or tier_json.get("assay") or {}
    token_dict    = tier_json.get("token_dict", {})
    treatment_templates = tier_json.get("treatment_templates", [])

    # Instrument from fluor tier (user may have already chosen it)
    fluor_instrument = fluor.get("instrument") or None
    inherited = _inherit_tiers(investigation, study, fluor)

    rows: list[dict] = []
    warnings: list[str] = []

    for file in files:
        fname = file.filename or "unknown"
        raw   = file.read()
        fluo  = fluor_instrument or _detect_fluorometer(fname, raw)

        try:
            row = _build_row(raw, fname, fluo, inherited, token_dict)
        except Exception as exc:
            curve_id_err = hashlib.sha256(raw).hexdigest()[:16] if raw else fname + "_err"
            row = {
                "filename":           {"value": fname,        "provenance": FROM_HEADER},
                "curve_id":           {"value": curve_id_err, "provenance": COMPUTED},
                "completeness_score": {"value": 0.0,          "provenance": COMPUTED},
                "_meta": {"filename": fname, "fluorometer": fluo, "parse_error": str(exc)},
            }
            warnings.append(f"{fname}: {exc}")

        rows.append(row)

    bundle_id = uuid.uuid4().hex

    return jsonify({
        "status":        "success",
        "bundle_id":     bundle_id,
        "rows":          rows,
        "tier_defaults": {
            "investigation":      investigation,
            "study":              study,
            "fluor":              fluor,
            "token_dict":         token_dict,
            "treatment_templates": treatment_templates,
        },
        "warnings":       warnings,
        "schema_version": SCHEMA_VERSION,
    })


@fluorescence_annotation.route("/api/fluorescence_annotation/export", methods=["POST"])
def export_bundle():
    """
    Receive the reviewed grid (JS posts back after user edits).
    Returns a base64-encoded ZIP bundle for download.

    JSON body: {bundle_id, rows, tier_defaults}
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "No JSON body."}), 400

    bundle_id     = data.get("bundle_id", "")
    rows          = data.get("rows", [])
    tier_defaults = data.get("tier_defaults", {})

    state = {
        "investigation":      tier_defaults.get("investigation", {}),
        "study":              tier_defaults.get("study", {}),
        "fluor":              tier_defaults.get("fluor") or tier_defaults.get("assay") or {},
        "token_dict":         tier_defaults.get("token_dict", {}),
        "treatment_templates": tier_defaults.get("treatment_templates", []),
        "rows":               rows,
    }

    bundle_bytes = bundle_serialise(state)
    b64 = base64.b64encode(bundle_bytes).decode("ascii")

    project = tier_defaults.get("investigation", {}).get("project_title", "annotation")
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", project)[:40]

    return jsonify({
        "status":   "success",
        "bundle_b64": b64,
        "filename": f"{safe_name}_annotation_bundle.zip",
    })


@fluorescence_annotation.route("/api/fluorescence_annotation/load_bundle", methods=["POST"])
def load_bundle():
    """
    Receive an annotation bundle ZIP, deserialise, and return the grid state
    so the user can pick up where they left off or merge new files.
    """
    if "bundle_file" not in request.files:
        return jsonify({"status": "error", "message": "No bundle file received."}), 400

    try:
        state = bundle_deserialise(request.files["bundle_file"].read())
    except Exception as exc:
        return jsonify({"status": "error",
                        "message": f"Could not read bundle: {exc}"}), 400

    return jsonify({
        "status":    "success",
        "rows":      state.get("rows", []),
        "tier_defaults": {
            "investigation":      state.get("investigation", {}),
            "study":              state.get("study", {}),
            "fluor":              state.get("fluor", {}),
            "token_dict":         state.get("token_dict", {}),
            "treatment_templates": state.get("treatment_templates", []),
        },
        "schema_version":   state.get("schema_version", ""),
        "_manifest":        state.get("_manifest", {}),
        "_version_warning": state.get("_version_warning"),
    })


# ── NCBI Taxonomy resolver ────────────────────────────────────────────────────

@fluorescence_annotation.route("/api/fluorescence_annotation/ncbi_taxon", methods=["GET"])
def ncbi_taxon_search():
    """
    Proxy typeahead search against NCBI Taxonomy (Entrez esearch + esummary).
    Query param: q=<search term>
    Returns JSON list of {taxid, name, rank, division}.
    """
    import urllib.request
    import urllib.parse

    query = request.args.get("q", "").strip()
    if len(query) < 3:
        return jsonify([])

    try:
        # Step 1: esearch to get taxon IDs
        esearch_url = (
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?"
            + urllib.parse.urlencode({
                "db": "taxonomy",
                "term": query + "[Scientific Name]",
                "retmax": "8",
                "retmode": "json",
            })
        )
        with urllib.request.urlopen(esearch_url, timeout=6) as resp:
            search_data = json.loads(resp.read().decode())
        id_list = search_data.get("esearchresult", {}).get("idlist", [])
        if not id_list:
            return jsonify([])

        # Step 2: esummary to get organism names
        esummary_url = (
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?"
            + urllib.parse.urlencode({
                "db": "taxonomy",
                "id": ",".join(id_list),
                "retmode": "json",
            })
        )
        with urllib.request.urlopen(esummary_url, timeout=6) as resp:
            summary_data = json.loads(resp.read().decode())

        results = []
        for tid in id_list:
            info = summary_data.get("result", {}).get(tid, {})
            results.append({
                "taxid":    int(tid),
                "name":     info.get("scientificname", ""),
                "rank":     info.get("rank", ""),
                "division": info.get("division", ""),
            })
        return jsonify(results)

    except Exception as exc:
        return jsonify({"error": str(exc)}), 502


# ── XLSX template ─────────────────────────────────────────────────────────────

def _build_xlsx_template(tier_defaults: dict | None = None) -> bytes:
    """
    Generate a multi-sheet MIAPPE-style XLSX workbook for pre-filling
    annotation fields.

    Sheets:
      Investigation — 1 row, 7 columns
      Study         — 1 row, study fields
      Fluorometer   — 1 row, fluorometer fields
      Treatments    — N rows, one per treatment template
      Per-curve     — N rows, one per file/curve (user fills in)

    Data validation (dropdowns) are added for vocab fields.
    If tier_defaults is provided, pre-fill cells from existing form values.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX template generation. "
            "Install with: pip install openpyxl"
        )

    wb = Workbook()
    td = tier_defaults or {}

    def _add_sheet(name: str, tier: str, source_dict: dict | None = None):
        ws = wb.create_sheet(title=name)
        fields = [(k, f) for k, f in FIELDS.items() if f.tier == tier]
        # Header row
        for ci, (k, fdef) in enumerate(fields, 1):
            cell = ws.cell(row=1, column=ci, value=fdef.label or k)
            cell.font = cell.font.copy(bold=True)
            # Dropdown validation for vocab fields
            if fdef.vocab:
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(str(v) for v in fdef.vocab) + '"',
                    allow_blank=True,
                )
                dv.error = f"Please choose from: {', '.join(str(v) for v in fdef.vocab)}"
                dv.errorTitle = fdef.label or k
                ws.add_data_validation(dv)
                dv.add(ws.cell(row=2, column=ci))
            # Pre-fill from tier_defaults
            if source_dict and k in source_dict:
                val = source_dict[k]
                if val not in (None, "", []):
                    ws.cell(row=2, column=ci, value=str(val))
        # Auto-width
        for ci, (k, _) in enumerate(fields, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 18
        return ws

    # Remove default sheet
    wb.remove(wb.active)

    _add_sheet("Investigation", "investigation", td.get("investigation"))
    _add_sheet("Study", "study", td.get("study"))
    _add_sheet("Fluorometer", "fluorometer", td.get("fluor"))

    # Treatment sheets — one per treatment type
    def _add_treatment_sheet(title: str, treat_type: str, sub_fields: list):
        ws = wb.create_sheet(title=title)
        all_keys = ["treatment_label"] + sub_fields
        for ci, k in enumerate(all_keys, 1):
            fdef = FIELDS.get(k)
            label = (fdef.label if fdef else k) or k
            cell = ws.cell(row=1, column=ci, value=label)
            cell.font = cell.font.copy(bold=True)
            if fdef and fdef.vocab:
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(str(v) for v in fdef.vocab) + '"',
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                for ri in range(2, 51):
                    dv.add(ws.cell(row=ri, column=ci))
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 22
        # Pre-fill existing templates of matching type
        data_row = 2
        for tmpl in td.get("treatment_templates", []):
            if tmpl.get("treatment_type", "chemical") == treat_type:
                for ci, k in enumerate(all_keys, 1):
                    val = tmpl.get(k, "")
                    if val:
                        ws.cell(row=data_row, column=ci, value=str(val))
                data_row += 1

    _add_treatment_sheet(
        "Chem. treatments", "chemical",
        ["chem_treatment", "chem_dose", "chem_unit", "chem_duration", "chem_detail"],
    )
    _add_treatment_sheet(
        "Stress treatments", "stress",
        ["stress_treatment", "stress_dose", "stress_unit", "stress_duration", "stress_detail"],
    )
    _add_treatment_sheet(
        "Other treatments", "other",
        ["other_treatment", "other_dose", "other_unit", "other_duration", "other_detail"],
    )

    # Per-curve sheet
    ws_curve = wb.create_sheet(title="Per-curve")
    curve_fields = [(k, f) for k, f in FIELDS.items() if f.tier == "per_curve"]
    for ci, (k, fdef) in enumerate(curve_fields, 1):
        cell = ws_curve.cell(row=1, column=ci, value=fdef.label or k)
        cell.font = cell.font.copy(bold=True)
        if fdef.vocab:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(str(v) for v in fdef.vocab) + '"',
                allow_blank=True,
            )
            ws_curve.add_data_validation(dv)
            for ri in range(2, 201):
                dv.add(ws_curve.cell(row=ri, column=ci))
        ws_curve.column_dimensions[ws_curve.cell(row=1, column=ci).column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_xlsx_upload(xlsx_bytes: bytes) -> dict:
    """
    Parse an uploaded XLSX template back into tier dicts + treatment templates
    + per-curve overrides.

    Returns:
        {
          'investigation': {field: value},
          'study':         {field: value},
          'fluor':         {field: value},
          'treatment_templates': [{label, chem_*, stress_*}, ...],
          'per_curve_overrides':  [{field: value, ...}, ...],
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    result: dict[str, Any] = {}

    def _read_tier_sheet(sheet_name: str, tier: str) -> dict:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        tier_fields = {f.label or k: k for k, f in FIELDS.items() if f.tier == tier}
        data = {}
        if ws.max_row >= 2:
            for ci, header in enumerate(headers):
                if header and header in tier_fields:
                    val = ws.cell(row=2, column=ci + 1).value
                    if val is not None and str(val).strip():
                        data[tier_fields[header]] = str(val).strip()
        return data

    result["investigation"] = _read_tier_sheet("Investigation", "investigation")
    result["study"]         = _read_tier_sheet("Study", "study")
    result["fluor"]         = _read_tier_sheet("Fluorometer", "fluorometer")

    # Treatment templates — 3 separate sheets
    templates = []

    def _parse_treat_sheet(sheet_name: str, treat_type: str, sub_fields: list):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        all_keys = ["treatment_label"] + sub_fields
        lbl_to_key = {(FIELDS[k].label or k): k for k in all_keys if k in FIELDS}
        col_map = {ci: lbl_to_key[h]
                   for ci, h in enumerate(headers)
                   if h and h in lbl_to_key}
        for ri in range(2, ws.max_row + 1):
            row_data: dict[str, Any] = {"treatment_type": treat_type}
            for ci, fk in col_map.items():
                val = ws.cell(row=ri, column=ci + 1).value
                if val is not None and str(val).strip():
                    row_data[fk] = str(val).strip()
            if row_data.get("treatment_label"):
                templates.append(row_data)

    _parse_treat_sheet("Chem. treatments", "chemical",
                       ["chem_treatment", "chem_dose", "chem_unit",
                        "chem_duration", "chem_detail"])
    _parse_treat_sheet("Stress treatments", "stress",
                       ["stress_treatment", "stress_dose", "stress_unit",
                        "stress_duration", "stress_detail"])
    _parse_treat_sheet("Other treatments", "other",
                       ["other_treatment", "other_dose", "other_unit",
                        "other_duration", "other_detail"])
    # Back-compat: old single "Treatments" sheet
    if not templates and "Treatments" in wb.sheetnames:
        _parse_treat_sheet("Treatments", "chemical",
                           ["chem_treatment", "chem_dose", "chem_unit",
                            "chem_duration", "stress_treatment",
                            "stress_dose", "stress_duration"])

    result["treatment_templates"] = templates

    # Per-curve overrides
    overrides = []
    if "Per-curve" in wb.sheetnames:
        ws = wb["Per-curve"]
        headers = [cell.value for cell in ws[1]]
        curve_fields = {(f.label or k): k for k, f in FIELDS.items()
                        if f.tier == "per_curve"}
        col_map = {}
        for ci, h in enumerate(headers):
            if h and h in curve_fields:
                col_map[ci] = curve_fields[h]
        for ri in range(2, ws.max_row + 1):
            row_data = {}
            for ci, field_key in col_map.items():
                val = ws.cell(row=ri, column=ci + 1).value
                if val is not None and str(val).strip():
                    row_data[field_key] = str(val).strip()
            if row_data:
                overrides.append(row_data)
    result["per_curve_overrides"] = overrides

    return result


@fluorescence_annotation.route(
    "/api/fluorescence_annotation/xlsx_template", methods=["POST"]
)
def xlsx_template():
    """
    Generate and return a multi-sheet XLSX template.
    Accepts optional JSON body with tier_defaults to pre-fill cells.
    """
    tier_defaults = None
    data = request.get_json(silent=True)
    if data:
        tier_defaults = data.get("tier_defaults")

    try:
        xlsx_bytes = _build_xlsx_template(tier_defaults)
    except ImportError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

    b64 = base64.b64encode(xlsx_bytes).decode("ascii")
    return jsonify({
        "status":   "ok",
        "xlsx_b64": b64,
        "filename": "annotation_template.xlsx",
    })


@fluorescence_annotation.route(
    "/api/fluorescence_annotation/xlsx_upload", methods=["POST"]
)
def xlsx_upload():
    """
    Upload a filled-in XLSX template. Returns parsed tier data + per-curve
    overrides that the JS can merge into the current state.
    """
    if "xlsx_file" not in request.files:
        return jsonify({"status": "error",
                        "message": "No XLSX file received."}), 400
    try:
        parsed = _parse_xlsx_upload(request.files["xlsx_file"].read())
    except ImportError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"status": "error",
                        "message": f"Failed to parse XLSX: {exc}"}), 400

    return jsonify({"status": "ok", **parsed})


# ── Smoke test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import zipfile as _zf
    from pathlib import Path

    _root = Path(__file__).parent
    _zip_path = _root / "static" / "files" / "OJIP_example_files.zip"

    if not _zip_path.exists():
        print(f"Sample ZIP not found at {_zip_path}")
        sys.exit(1)

    # Extract first file from the sample ZIP
    with _zf.ZipFile(_zip_path) as z:
        names = [n for n in z.namelist() if not n.endswith("/")]
        if not names:
            print("No files inside the sample ZIP.")
            sys.exit(1)
        _fname = names[0]
        _raw   = z.read(_fname)

    print(f"Sample file: {_fname}  ({len(_raw):,} bytes)")

    # Minimal tier blocks
    _inv   = {"project_title": "Smoke test project", "institution": "CzechGlobe",
               "contact_name": "T. Zavrel"}
    _study = {"organism": "Synechocystis sp. PCC 6803",
               "genotype": "WT",
               "sample_type": "liquid culture",
               "acclimation_min": 30,
               "medium": "BG-11"}
    _fluor = {"instrument": None,    # will be auto-detected from file content
              "fo_timing": "20 µs"}
    _tdict = {
        "separator": "_",
        "tokens": [
            {"position": 0, "field": "treatment_label"},
            {"position": 1, "field": "timepoint", "strip_suffix": "h"},
            {"position": 2, "field": "bio_rep",   "strip_prefix": "rep"},
        ],
    }

    _inherited = _inherit_tiers(_inv, _study, _fluor)
    print(f"Inherited fields ({len(_inherited)}): {list(_inherited.keys())}\n")

    _row = _build_row(_raw, _fname, None, _inherited, _tdict)

    print("--- Grid row ---------------------------------------------------")
    for _k in sorted(_row.keys()):
        if _k.startswith("_"):
            continue
        _cell = _row[_k]
        _v, _p = _cell.get("value"), _cell.get("provenance", "")
        if _v is not None:
            print(f"  {_k:<35} = {str(_v):<22}  [{_p}]")

    _meta = _row.get("_meta", {})
    print("\n--- Parse metadata ---------------------------------------------")
    for _k, _v in _meta.items():
        print(f"  {_k}: {_v}")

    _score = _row.get("completeness_score", {}).get("value", "?")
    print(f"\nCompleteness score: {_score}%")

    if "_transient" in _row:
        _t = _row["_transient"]
        print(f"Transient: {len(_t['time_ms'])} points  "
              f"t=[{_t['time_ms'][0]:.4f} .. {_t['time_ms'][-1]:.4f}] ms")

    # Validation
    _issues = validate_row(_row)
    print(f"Validation issues: {_issues or 'none'}")

    # Bundle round-trip
    print("\n--- Bundle round-trip ------------------------------------------")
    _state = {
        "investigation": _inv,
        "study":         _study,
        "fluor":         _fluor,
        "token_dict":    _tdict,
        "rows":          [_row],
    }
    _bundle_bytes = bundle_serialise(_state)
    print(f"Bundle size: {len(_bundle_bytes):,} bytes")

    with _zf.ZipFile(io.BytesIO(_bundle_bytes)) as _z:
        print(f"Bundle contents: {_z.namelist()}")

    _state2 = bundle_deserialise(_bundle_bytes)
    assert _state2["investigation"]["project_title"] == "Smoke test project", \
        "Round-trip FAILED: project_title mismatch"
    assert len(_state2["rows"]) == 1, \
        "Round-trip FAILED: row count mismatch"
    _cid_orig  = _row["curve_id"]["value"]
    _cid_rt    = _state2["rows"][0].get("curve_id", {}).get("value")
    assert _cid_orig == _cid_rt, f"Round-trip FAILED: curve_id {_cid_orig!r} ≠ {_cid_rt!r}"

    print("Round-trip: OK")
    print("\nSmoke test passed.")
